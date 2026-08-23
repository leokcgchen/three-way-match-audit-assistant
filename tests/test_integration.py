"""三单匹配 + 截止性 Agent 联动集成测试。

优先使用 FastAPI TestClient（不依赖外部已启动进程）。
另含：Agent 不可达降级、合并优先级。
"""

from __future__ import annotations

from typing import Optional

import requests
from fastapi.testclient import TestClient

from src.api.main import app
from src.three_way_match import (
    Invoice,
    Order,
    ThreeWayMatcher,
    ThreeWayMatchRequest,
    WarehouseReceipt,
    merge_overall_status,
)

DEAD_URL = "http://127.0.0.1:59999/api/v1/cutoff"


def _sample_request(
    *,
    payment_terms: str = "签收后10日",
    receipt_date: str = "2026-06-01",
    posting_date: str = "2026-06-01",
    supplier: str = "甲供应商",
    amount: float = 500.0,
    quantity: float = 100.0,
) -> ThreeWayMatchRequest:
    return ThreeWayMatchRequest(
        order=Order(
            order_no="PO-INT-001",
            supplier_name=supplier,
            total_amount=amount,
            quantity=quantity,
            unit="吨",
            order_date="2026-05-20",
            payment_terms=payment_terms,
            contract_no="HT-INT-001",
        ),
        warehouse_receipt=WarehouseReceipt(
            receipt_no="WR-INT-001",
            order_no="PO-INT-001",
            supplier_name=supplier,
            total_amount=amount,
            quantity=quantity,
            receipt_date=receipt_date,
            receiver="李四",
        ),
        invoice=Invoice(
            invoice_no="INV-INT-001",
            order_no="PO-INT-001",
            supplier_name=supplier,
            total_amount=amount,
            quantity=quantity,
            invoice_date="2026-06-08",
            posting_date=posting_date,
        ),
    )


def test_merge_priority():
    assert merge_overall_status("PASS", "PASS") == "PASS"
    assert merge_overall_status("PASS", "WARNING") == "WARNING"
    assert merge_overall_status("WARNING", "FAIL") == "FAIL"
    assert merge_overall_status("FAIL", "PASS") == "FAIL"
    print("test_merge_priority: PASS")


def test_cutoff_agent_down_graceful():
    matcher = ThreeWayMatcher()
    result = matcher.match_and_cutoff(_sample_request(), cutoff_agent_url=DEAD_URL)
    assert result["match_result"].overall_status == "PASS"
    assert result["cutoff_result"] is None
    assert result["cutoff_available"] is False
    assert "截止性Agent未响应" in (result.get("cutoff_error") or "")
    assert result["overall_status"] == "WARNING"
    print("test_cutoff_agent_down_graceful: PASS")


def test_api_three_way_match_integration():
    client = TestClient(app)
    payload = _sample_request().model_dump()
    resp = client.post("/api/v1/three-way-match", json=payload)
    assert resp.status_code == 200, resp.text
    data = resp.json()
    assert "match_result" in data
    assert "cutoff_result" in data
    assert data.get("cutoff_available") is True
    assert data["match_result"]["overall_status"] == "PASS"
    assert data["cutoff_result"]["测试状态"] == "PASS"
    assert data["overall_status"] == "PASS"
    assert data["cutoff_result"]["业务编号"] == "PO-INT-001"
    print("test_api_three_way_match_integration: PASS")


def test_api_merge_fail_priority():
    client = TestClient(app)
    req = _sample_request()
    req.warehouse_receipt.supplier_name = "乙供应商"
    resp = client.post("/api/v1/three-way-match", json=req.model_dump())
    assert resp.status_code == 200, resp.text
    data = resp.json()
    assert data["match_result"]["overall_status"] == "FAIL"
    assert data["overall_status"] == "FAIL"
    print("test_api_merge_fail_priority: PASS")


def test_http_match_and_cutoff_when_server_up():
    """若本机 8000 已启动且为 TWM-2，再验 HTTP 客户端路径。"""
    try:
        health = requests.get("http://127.0.0.1:8000/health", timeout=2)
    except requests.RequestException:
        print("test_http_match_and_cutoff_when_server_up: SKIP (API 未启动)")
        return
    if health.status_code != 200 or "TWM-2" not in health.text:
        print("test_http_match_and_cutoff_when_server_up: SKIP (非 TWM-2 实例)")
        return
    matcher = ThreeWayMatcher()
    result = matcher.match_and_cutoff(
        _sample_request(),
        cutoff_agent_url="http://127.0.0.1:8000/api/v1/cutoff",
        inprocess=False,
    )
    assert result["cutoff_available"] is True
    assert result["cutoff_result"] is not None
    print("test_http_match_and_cutoff_when_server_up: PASS")


def test_cutoff_skipped_when_posting_date_missing():
    """入账日期缺失：不调用截止性；整体状态跟随三单匹配（PASS 不降级）。"""
    matcher = ThreeWayMatcher()
    req = _sample_request(posting_date="")
    req.invoice.posting_date = None
    # 即使有开票日，也不应兜底调用截止性
    req.invoice.invoice_date = "2026-06-08"
    result = matcher.match_and_cutoff(req, cutoff_agent_url=DEAD_URL, inprocess=False)
    assert result["match_result"].overall_status == "PASS"
    assert result["cutoff_result"] is None
    assert result["cutoff_available"] is False
    assert result["cutoff_skipped_reason"] == "入账日期缺失，无法执行截止性测试"
    assert result["overall_status"] == "PASS"
    assert result["match_result"].cutoff_available is False
    assert result["match_result"].cutoff_skipped_reason == (
        "入账日期缺失，无法执行截止性测试"
    )
    assert result["match_result"].cutoff_test_status == "SKIPPED"
    print("test_cutoff_skipped_when_posting_date_missing: PASS")


def test_payment_days_extraction_consistency():
    """账期提取在 matcher 组装与 cutoff_runner 解析中保持一致。"""
    from src.api.cutoff_runner import resolve_payment_days
    from src.models.schemas import CutoffRequest
    from src.utils.date_extractor import extract_days_from_description

    assert extract_days_from_description("票到30天") == 30
    assert extract_days_from_description("签收后10日") == 10

    matcher = ThreeWayMatcher()
    req = _sample_request(payment_terms="票到30天", posting_date="2026-07-01")
    payload = matcher.build_cutoff_payload(req)
    assert payload["合同账期天数"] == 30

    cutoff_req = CutoffRequest(
        业务编号="PO-X",
        签收日期="2026-06-01",
        入账日期="2026-07-01",
        入账金额=100.0,
        合同账期描述="票到30天",
        合同账期天数=None,
    )
    assert resolve_payment_days(cutoff_req) == 30
    print("test_payment_days_extraction_consistency: PASS")


def test_workbook_22_columns_with_match_and_cutoff_only():
    """CSV 列定义仍为 22 列；API 默认不再自动写底稿（改由菜单手动生成 xlsx）。"""
    from config.settings import settings
    from src.reporting import WORKBOOK_COLUMNS

    assert len(WORKBOOK_COLUMNS) == 22
    for col in (
        "三单匹配状态",
        "供应商一致性",
        "订单金额（万元）",
        "入库金额（万元）",
        "发票金额（万元）",
        "金额差异率（%）",
        "三单决策",
    ):
        assert col in WORKBOOK_COLUMNS

    wb_path = settings.get_workbook_path()
    if wb_path.exists():
        wb_path.unlink()

    client = TestClient(app)
    resp = client.post("/api/v1/three-way-match", json=_sample_request().model_dump())
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body.get("match_result", {}).get("overall_status") == "PASS"
    # 手动底稿模式：默认不落盘
    assert not body.get("底稿文件路径")
    assert not wb_path.exists()

    cutoff_payload = {
        "业务编号": "SO-E2-ONLY",
        "签收日期": "2026-06-01",
        "入账日期": "2026-06-01",
        "入账金额": 500,
        "合同账期天数": 10,
    }
    cresp = client.post("/api/v1/cutoff", json=cutoff_payload)
    assert cresp.status_code == 200, cresp.text
    assert cresp.json()["测试状态"] == "PASS"
    assert cresp.json()["应确认日期"] == "2026-06-01"
    assert not cresp.json().get("底稿文件路径")
    assert not wb_path.exists()
    print("test_workbook_22_columns_with_match_and_cutoff_only: PASS")


def test_manual_audit_workbook_xlsx(tmp_dir: Optional[str] = None):
    """手动 xlsx：按已跑测试生成多 sheet。"""
    import tempfile
    from pathlib import Path

    from src.reporting.audit_workbook_xlsx import (
        build_audit_workbook_payload,
        generate_audit_workbook_xlsx,
    )

    root = Path(tmp_dir) if tmp_dir else Path(tempfile.mkdtemp(prefix="wb_xlsx_"))
    payload = build_audit_workbook_payload(
        amount={
            "status": "PASS",
            "human_readable_summary": "金额一致",
            "accuracy_report": {"business_id": "B1"},
        }
    )
    out = generate_audit_workbook_xlsx(payload, root / "t.xlsx")
    assert out.exists()
    from openpyxl import load_workbook

    wb = load_workbook(out)
    assert "汇总" in wb.sheetnames
    assert "金额准确性" in wb.sheetnames
    print("test_manual_audit_workbook_xlsx: PASS")


def test_reports_dir_config_effective(tmp_dir: Optional[str] = None):
    """REPORTS_DIR 配置生效：CSV 输出到自定义目录。"""
    import tempfile
    from pathlib import Path

    from config.settings import Settings
    from src.models.schemas import CutoffResponse
    from src.reporting.workbook_generator import WorkbookGenerator

    root = Path(tmp_dir) if tmp_dir else Path(tempfile.mkdtemp(prefix="e2_reports_"))
    custom = root / "custom_out"
    s = Settings(
        REPORTS_DIR=str(custom),
        WORKBOOK_FILENAME="底稿_自定义.csv",
    )
    out = s.get_workbook_path()
    assert out == custom / "底稿_自定义.csv"

    WorkbookGenerator.append_to_workbook(
        CutoffResponse(
            报告ID="R1",
            业务编号="SO-CFG",
            测试状态="PASS",
            风险等级="无异常",
            问题描述="ok",
            计算依据="n/a",
            底稿回填={"审计结论": "无异常"},
        ),
        output_path=str(out),
    )
    assert out.exists()
    text = out.read_text(encoding="utf-8-sig")
    assert "三单匹配状态" in text.splitlines()[0]
    assert "SO-CFG" in text
    print("test_reports_dir_config_effective: PASS", out)


def test_human_readable_summary_contains_key_fields():
    """摘要包含供应商、金额、签收日、入账日、偏差等关键信息。"""
    matcher = ThreeWayMatcher()
    result = matcher.match_and_cutoff(_sample_request(), inprocess=True)
    summary = result.get("human_readable_summary") or result[
        "match_result"
    ].human_readable_summary
    assert summary
    assert "供应商一致" in summary
    assert "金额差异" in summary
    assert "控制权转移日（签收/验收）2026-06-01" in summary
    assert "2026-06-01" in summary  # 应确认/入账
    assert "偏差" in summary
    assert "综合结论" in summary
    assert result["overall_status"] == "PASS"
    assert "✅ PASS" in summary
    print("test_human_readable_summary_contains_key_fields: PASS")


def test_human_readable_summary_missing_posting_date():
    """入账日期缺失时摘要提示未执行截止性。"""
    matcher = ThreeWayMatcher()
    req = _sample_request()
    req.invoice.posting_date = None
    result = matcher.match_and_cutoff(req, cutoff_agent_url=DEAD_URL, inprocess=False)
    summary = result.get("human_readable_summary") or result[
        "match_result"
    ].human_readable_summary
    assert "截止性测试未执行" in summary
    assert "入账日期缺失" in summary
    assert "仅三单匹配" in summary or result["overall_status"] == "PASS"
    assert result["overall_status"] == "PASS"
    print("test_human_readable_summary_missing_posting_date: PASS")


if __name__ == "__main__":
    test_merge_priority()
    test_cutoff_agent_down_graceful()
    test_api_three_way_match_integration()
    test_api_merge_fail_priority()
    test_http_match_and_cutoff_when_server_up()
    test_cutoff_skipped_when_posting_date_missing()
    test_payment_days_extraction_consistency()
    test_workbook_22_columns_with_match_and_cutoff_only()
    test_manual_audit_workbook_xlsx()
    test_reports_dir_config_effective()
    test_human_readable_summary_contains_key_fields()
    test_human_readable_summary_missing_posting_date()
    print("全部集成测试通过。")
