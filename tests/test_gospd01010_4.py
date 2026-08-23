"""GOSPD01010.4 价格分摊底稿：配方 + 断言 + filler。"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from src.audit.gospd01010_4_assertions import (
    LABEL_O_NOT_MET,
    LABEL_P_NA,
    LABEL_V_YES,
    LABEL_W_YES,
    assert_discount_allocation_criteria,
    build_gospd01010_4_assertions,
)
from src.reporting.gospd01010_4_filler import (
    fill_gospd01010_4_workbook,
    resolve_template_path,
)
from src.workflow.pipeline import build_workbooks_for_job, selected_workbook_formats
from src.workflow.recipes import list_workpaper_goals, resolve_workflow_plan


def test_goal_listed():
    assert "gospd01010_4" in {g["goal_id"] for g in list_workpaper_goals()}
    plan = resolve_workflow_plan(["gospd01010_4"])
    assert "amount_test" in plan["required_steps"]
    assert "three_way_cutoff" not in plan["required_steps"]
    assert plan["workbook_formats"] == ["gospd01010_4"]


def test_criteria_defaults_not_met():
    out = assert_discount_allocation_criteria(
        docs_by_type={
            "contract": {
                "raw_text": "销售商品交付",
                "fields": {},
            }
        },
        contract_res={"extracted": {"issue_codes": []}},
    )
    assert out["verdict_label"] == LABEL_O_NOT_MET


def test_filler_smoke(tmp_path: Path):
    assert resolve_template_path().is_file()
    job = {
        "job_id": "t4",
        "goal_ids": ["gospd01010_4"],
        "plan": resolve_workflow_plan(["gospd01010_4"]),
        "classified": [
            {
                "file_name": "c.pdf",
                "doc_type": "contract",
                "raw_text": "销售合同 交付货物 签收转移控制权",
                "fields": {
                    "contractNo": "HT25-0104",
                    "buyerName": "客户B",
                    "totalAmount": 100000,
                    "paymentTerms": "签收后30日",
                },
            },
            {
                "file_name": "o.pdf",
                "doc_type": "order",
                "fields": {"orderNo": "SO25-0104", "totalAmount": 100000},
            },
        ],
        "contract_terms": {
            "status": "PASS",
            "extracted": {
                "dimension_statuses": {"交易对价": "CLEAR", "履约义务": "CLEAR"}
            },
        },
        "amount_test": {"status": "PASS"},
    }
    path = fill_gospd01010_4_workbook(job, tmp_path / "o4.xlsx", entity_name="测")
    wb = load_workbook(path)
    assert "底稿须知" in wb.sheetnames
    ws = wb["Sheet1"]
    assert ws["E5"].value == "GOSPD01010.4"
    assert ws["B23"].value == 1
    assert ws["O23"].value == LABEL_O_NOT_MET
    assert ws["P23"].value in {LABEL_P_NA, "Applicable 适用"}
    assert ws["V23"].value == LABEL_V_YES
    assert ws["W23"].value == LABEL_W_YES
    assert ws["U23"].value == "=F23-T23"
    assert ws["G23"].value == 100000


def test_multi_export(tmp_path: Path, monkeypatch):
    monkeypatch.setattr("src.workflow.pipeline.workbook_output_dir", lambda: tmp_path)
    job = {
        "job_id": "m4",
        "goal_ids": ["gospd01010_4", "gospd01010_3"],
        "plan": resolve_workflow_plan(["gospd01010_4", "gospd01010_3"]),
        "classified": [
            {
                "file_name": "c.pdf",
                "doc_type": "contract",
                "raw_text": "合同",
                "fields": {"contractNo": "HT1", "totalAmount": 100},
            }
        ],
        "contract_terms": {
            "status": "PASS",
            "extracted": {"dimension_statuses": {"交易对价": "CLEAR"}},
        },
        "amount_test": {"status": "PASS"},
    }
    assert selected_workbook_formats(job) == ["gospd01010_4", "gospd01010_3"]
    paths = build_workbooks_for_job(job)
    assert len(paths) == 2
    joined = " ".join(p.name.upper() for p in paths)
    assert "01010.4" in joined and "01010.3" in joined
