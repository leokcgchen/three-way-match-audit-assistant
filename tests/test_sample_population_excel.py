"""抽样清单：灵活表头 Excel + 文本业务号。"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook

from src.audit.sample_population import (
    build_sample_population,
    desk_sample_ids,
    ledger_patch_from_parsed,
    parse_sample_workbook,
)
from src.workflow.sample_required_fields import missing_required_fields, required_fields_for_docs


def _write_sap_like(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "SAP序时账"
    ws.append(
        [
            "凭证号",
            "过账日期",
            "客户名称",
            "销售订单号",
            "借方金额",
            "贷方金额",
        ]
    )
    ws.append(["SA25-0001", datetime(2025, 1, 11), "甲公司", "SO25-0001", 2015.67, 2015.67])
    ws.append(["SA25-0002", datetime(2025, 1, 16), "乙公司", "SO25-0002", 3089.02, 3089.02])
    ws2 = wb.create_sheet("金蝶U8序时账")
    ws2.append(["凭证字号", "入账日期", "往来单位名称", "订单编号", "借方发生额", "贷方发生额"])
    ws2.append(["SA25-0001", datetime(2025, 1, 11), "甲公司", "SO25-0001", 2015.67, 2015.67])
    ws2.append(["SA25-0099", datetime(2025, 2, 1), "丙公司", "SO25-0099", 100, 100])
    wb.save(path)


def test_parse_prefers_order_no_not_voucher(tmp_path: Path):
    xlsx = tmp_path / "sample.xlsx"
    _write_sap_like(xlsx)
    parsed = parse_sample_workbook(xlsx)
    assert parsed["business_ids"] == ["SO25-0001", "SO25-0002", "SO25-0099"]
    row = parsed["rows"][0]
    assert row["book_date"] == "2025-01-11"
    assert row["book_amount"] == 2015.67
    assert row["customer"] == "甲公司"
    assert parsed["ledger_mapping"]["biz_id"] == "business_id"
    assert parsed["ledger_mapping"]["posting_date"] == "book_date"
    assert len(parsed["ledger_rows"]) >= 3
    assert parsed["ledger_auto_ok"] is True
    patch = ledger_patch_from_parsed(parsed, path=str(xlsx))
    assert patch["ledger_mapping"]["biz_id"] == "business_id"
    assert patch["ledger_rows"][0]["book_date"] == "2025-01-11"


def test_parse_prefers_business_number_and_accepts_multi_segment_ids(tmp_path: Path):
    xlsx = tmp_path / "adaptive-primary-key.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "抽样清单"
    ws.append(
        [
            "业务编号",
            "账载日期",
            "凭证号",
            "客户名称",
            "销售订单号",
            "发票号码",
            "账载金额",
        ]
    )
    ws.append(
        [
            "YW-2025-3986",
            datetime(2026, 1, 16),
            "JZ-202601-0148",
            "甲公司",
            "SO-251218-7365",
            "FP-260116-8417",
            113000,
        ]
    )
    ws.append(
        [
            "YW-2025-3962",
            datetime(2026, 1, 2),
            "JZ-202601-0087",
            "乙公司",
            "SO-251209-7214",
            "FP-260102-8305",
            113000,
        ]
    )
    wb.save(xlsx)

    parsed = parse_sample_workbook(xlsx)

    assert parsed["business_ids"] == ["YW-2025-3986", "YW-2025-3962"]
    assert parsed["primary_key_column"] == "业务编号"
    assert parsed["primary_key_method"] == "keyword"
    assert parsed["primary_key_confidence"] > 0.9
    assert parsed["rows"][0]["book_date"] == "2026-01-16"
    assert parsed["rows"][0]["book_amount"] == 113000.0
    assert parsed["rows"][0]["order_numbers"] == ["SO-251218-7365"]
    assert parsed["rows"][1]["order_numbers"] == ["SO-251209-7214"]
    assert parsed["rows"][0]["invoice_numbers"] == ["FP-260116-8417"]
    assert parsed["rows"][0]["voucher_numbers"] == ["JZ-202601-0148"]
    assert parsed["ambiguous_aliases"] == []


def test_parse_keeps_content_matching_features(tmp_path: Path):
    xlsx = tmp_path / "content-index.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(
        [
            "业务编号",
            "销售订单号",
            "发票号码",
            "客户名称",
            "物料名称",
            "数量",
            "单位",
            "账载金额",
            "币种",
        ]
    )
    ws.append(
        [
            "YW-2025-3995",
            "SO-251229-7498",
            "CI-260119-0068",
            "NordWerk Verpackung GmbH",
            "NW-500",
            12,
            "套",
            73066.7,
            "EUR",
        ]
    )
    wb.save(xlsx)

    parsed = parse_sample_workbook(xlsx)
    row = parsed["rows"][0]

    assert row["invoice_numbers"] == ["CI-260119-0068"]
    assert row["material_names"] == ["NW-500"]
    assert row["quantities"] == [12.0]
    assert row["units"] == ["套"]
    assert row["currencies"] == ["EUR"]


def test_parse_merges_order_aliases_and_reports_cross_business_duplicates(tmp_path: Path):
    xlsx = tmp_path / "duplicate-order-alias.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "抽样清单"
    ws.append(["业务编号", "订单编号"])
    ws.append(["YW-2025-0001", "SO-251201-0001"])
    ws.append(["YW-2025-0001", "SO-251201-0002"])
    ws.append(["YW-2025-0002", "SO-251201-0002"])
    wb.save(xlsx)

    parsed = parse_sample_workbook(xlsx)

    assert parsed["business_ids"] == ["YW-2025-0001", "YW-2025-0002"]
    assert parsed["rows"][0]["order_numbers"] == [
        "SO-251201-0001",
        "SO-251201-0002",
    ]
    assert parsed["ambiguous_aliases"] == [
        {
            "type": "order_number",
            "value": "SO-251201-0002",
            "business_ids": ["YW-2025-0001", "YW-2025-0002"],
        }
    ]


def test_parse_infers_unknown_erp_identifier_from_unique_values(tmp_path: Path):
    xlsx = tmp_path / "unknown-erp.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Record Locator", "Posting Date", "Amount", "Customer"])
    ws.append(["CASE-A9", datetime(2025, 12, 30), 100, "甲公司"])
    ws.append(["CASE-B7", datetime(2025, 12, 31), 200, "乙公司"])
    wb.save(xlsx)

    parsed = parse_sample_workbook(xlsx)

    assert parsed["business_ids"] == ["CASE-A9", "CASE-B7"]
    assert parsed["primary_key_column"] == "Record Locator"
    assert parsed["primary_key_method"] == "uniqueness_fallback"


def test_parse_does_not_treat_row_date_amount_or_customer_as_primary_key(tmp_path: Path):
    xlsx = tmp_path / "no-safe-key.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["行号", "日期", "金额", "客户名称"])
    ws.append([1, datetime(2025, 12, 30), 100, "甲公司"])
    ws.append([2, datetime(2025, 12, 31), 200, "乙公司"])
    wb.save(xlsx)

    with pytest.raises(ValueError) as exc_info:
        parse_sample_workbook(xlsx)

    message = str(exc_info.value)
    assert "未能自动确定唯一业务索引列" in message
    assert "行号" in message
    assert "销售订单号" not in message


def test_build_population_keeps_book_rows():
    pop = build_sample_population(
        business_ids=["SO25-0001"],
        rows=[
            {
                "business_id": "SO25-0001",
                "book_date": "2025-01-11",
                "book_amount": 1,
                "order_numbers": ["ORDER-001"],
            }
        ],
        source="excel",
        ambiguous_aliases=[
            {
                "type": "order_number",
                "value": "ORDER-SHARED",
                "business_ids": ["SO25-0001", "SO25-0002"],
            }
        ],
    )
    assert pop["count"] == 1
    assert pop["rows"][0]["book_date"] == "2025-01-11"
    assert pop["rows"][0]["order_numbers"] == ["ORDER-001"]
    assert pop["ambiguous_aliases"][0]["value"] == "ORDER-SHARED"


def test_desk_chain_exposes_business_and_order_display_index():
    from src.workflow.sample_desk import build_desk_chains

    job = _gospd_job(
        [],
        pop_ids=["YW-2025-3962"],
        extra={
            "sample_population": {
                "business_ids": ["YW-2025-3962"],
                "rows": [
                    {
                        "business_id": "YW-2025-3962",
                        "order_numbers": ["SO-251209-7214"],
                    }
                ],
            }
        },
    )

    row = build_desk_chains(job)[0]

    assert row["order_numbers"] == ["SO-251209-7214"]
    assert row["display_index"] == "YW-2025-3962 & SO-251209-7214"


def test_desk_ids_are_strictly_bounded_by_population():
    job = {
        "sample_population": {"business_ids": ["SO25-0001"]},
        "classified": [
            {"file_name": "a.pdf", "doc_type": "invoice", "fields": {"orderNo": "SO25-0001"}},
            {"file_name": "b.pdf", "doc_type": "invoice", "fields": {"orderNo": "SO25-7777"}},
        ],
    }
    ids = desk_sample_ids(job)
    assert ids == ["SO25-0001"]


def test_desk_ids_drop_truncated_prefix_of_population():
    job = {
        "sample_population": {"business_ids": ["SO25-0021"]},
        "classified": [
            {
                "file_name": "ghost.pdf",
                "doc_type": "receipt",
                "fields": {"orderNo": "SO25-002"},
            },
            {
                "file_name": "inv.pdf",
                "doc_type": "invoice",
                "fields": {"orderNo": "SO25-0021", "invoiceNo": "1", "buyerName": "甲", "totalAmount": "10", "documentDate": "2026-01-01"},
            },
        ],
    }
    ids = desk_sample_ids(job)
    assert "SO25-0021" in ids
    assert "SO25-002" not in ids


def test_01030_invoice_does_not_require_posting_date():
    docs = [
        {
            "doc_type": "invoice",
            "fields": {
                "invoiceNo": "1",
                "buyerName": "甲",
                "totalAmount": "10",
                "documentDate": "2026-01-01",
            },
        }
    ]
    keys = {r["key"] for r in required_fields_for_docs(docs, goal_ids=["gospd01030"])}
    assert "postingDate" not in keys
    assert missing_required_fields(docs, goal_ids=["gospd01030"]) == []


def _full_invoice(order_no: str = "SO25-0001") -> dict:
    document = {
        "file_name": f"{order_no}.pdf",
        "doc_type": "invoice",
        "fields": {
            "orderNo": order_no,
            "invoiceNo": "INV-1",
            "documentNo": "INV-1",
            "buyerName": "甲公司",
            "supplierName": "乙公司",
            "totalAmount": "113",
            "amount": "100",
            "taxAmount": "13",
            "quantity": "1",
            "postingDate": "2026-01-02",
            "documentDate": "2026-01-01",
        },
    }
    document["raw_text"] = (
        f"增值税专用发票\n订单号 {order_no}\n发票号码 INV-1\n买方 甲公司\n卖方 乙公司\n"
        "价税合计 113\n不含税金额 100\n税额 13\n数量 1\n"
        "入账日期 2026-01-02\n开票日期 2026-01-01"
    )
    return document


def _full_pack(order_no: str = "SO25-0001") -> list[dict]:
    return [
        {
            "file_name": f"{order_no}_order.pdf",
            "doc_type": "order",
            "raw_text": (
                f"销售订单\n订单编号 {order_no}\n合同号 HT-1\n买方 甲公司\n"
                "数量 1\n价税合计 113"
            ),
            "fields": {
                "orderNo": order_no,
                "contractNo": "HT-1",
                "buyerName": "甲公司",
                "quantity": "1",
                "totalAmount": "113",
            },
        },
        {
            "file_name": f"{order_no}_receipt.pdf",
            "doc_type": "receipt",
            "raw_text": f"签收验收单\n关联订单号 {order_no}\n验收日期 2025-12-30\n实收数量 1",
            "fields": {"orderNo": order_no, "acceptanceDate": "2025-12-30", "quantity": "1"},
        },
        _full_invoice(order_no),
    ]


def _gospd_job(docs: list, *, pop_ids: list[str] | None = None, extra: dict | None = None) -> dict:
    job = {
        "goal_ids": ["gospd01030"],
        "plan": {
            "goal_ids": ["gospd01030"],
            "required_steps": [
                "field_confirm",
                "relations_gate4",
                "three_way_cutoff",
                "conclusion_gate5",
            ],
        },
        "classified": docs,
        "gospd_sample_results": {},
        "sample_population": {"business_ids": pop_ids or ["SO25-0001"]},
    }
    if extra:
        job.update(extra)
    return job


def test_required_fields_depend_on_present_docs():
    invoice_only = [
        {
            "doc_type": "invoice",
            "fields": {"invoiceNo": "1", "buyerName": "甲", "totalAmount": "10", "postingDate": "2026-01-02"},
        }
    ]
    keys = {r["key"] for r in required_fields_for_docs(invoice_only, goal_ids=["gospd01030"])}
    assert "invoiceNo" in keys
    assert "acceptanceDate" not in keys
    assert missing_required_fields(invoice_only, goal_ids=["gospd01030"])  # documentDate 等可能仍缺
    both = invoice_only + [{"doc_type": "receipt", "fields": {"acceptanceDate": "2025-12-30", "quantity": "1"}}]
    miss = missing_required_fields(both, goal_ids=["gospd01030"])
    assert "acceptanceDate" not in miss


def test_01030_three_way_pack_does_not_require_contract_number():
    docs = _full_pack()
    docs[0]["fields"].pop("contractNo", None)

    rows = required_fields_for_docs(docs, goal_ids=["gospd01030"])

    assert "contractNo" not in {row["key"] for row in rows}
    assert "contractNo" not in missing_required_fields(docs, goal_ids=["gospd01030"])


def test_auto_pass_fields_when_complete():
    from src.workflow.sample_desk import apply_auto_pass_on_job, desk_row_status

    job, passed = apply_auto_pass_on_job(_gospd_job(_full_pack()))
    assert passed == ["SO25-0001"]
    assert job["gospd_sample_results"]["SO25-0001"]["fields_confirmed"] is True
    st = desk_row_status(job, "SO25-0001")
    assert st["reason"] == "tests_pending"
    assert st["light"] == "wait"


def test_auto_pass_invoice_without_posting_date():
    from src.workflow.sample_desk import apply_auto_pass_on_job

    inv = _full_invoice()
    inv["fields"].pop("postingDate", None)
    pack = _full_pack()
    pack[-1] = inv
    job, passed = apply_auto_pass_on_job(_gospd_job(pack))
    assert passed == ["SO25-0001"]


def test_desk_row_missing_labels_chinese():
    from src.workflow.sample_desk import desk_row_status

    pack = _full_pack()
    pack[1] = {
        "file_name": "thin.pdf",
        "doc_type": "receipt",
        "fields": {"orderNo": "SO25-0001"},
    }
    st = desk_row_status(_gospd_job(pack), "SO25-0001")
    assert st["reason"] == "fields_gap"
    assert st["label"].startswith("缺：")
    assert st["missing_labels"]
    assert "文件日期" in st["missing_labels"] or "数量" in st["missing_labels"]


def test_desk_missing_invoice_is_not_field_ok():
    from src.workflow.sample_desk import apply_auto_pass_on_job, desk_row_status

    docs = [d for d in _full_pack() if d["doc_type"] != "invoice"]
    job = _gospd_job(docs)
    st = desk_row_status(job, "SO25-0001")
    assert st["reason"] == "missing_docs"
    assert "发票" in (st.get("missing_doc_labels") or [])
    nxt, passed = apply_auto_pass_on_job(job)
    assert passed == []
    assert not (nxt.get("gospd_sample_results") or {}).get("SO25-0001", {}).get("fields_confirmed")


def test_auto_pass_skips_missing_and_ambiguity():
    from src.workflow.sample_desk import apply_auto_pass_on_job

    thin = {
        "file_name": "thin.pdf",
        "doc_type": "invoice",
        "fields": {"orderNo": "SO25-0001", "invoiceNo": "1"},
    }
    job, passed = apply_auto_pass_on_job(_gospd_job([thin]))
    assert passed == []
    assert not (job.get("gospd_sample_results") or {}).get("SO25-0001", {}).get("fields_confirmed")


def test_auto_conclusion_pass_and_fail():
    from src.workflow.sample_desk import apply_auto_conclusions_on_job, desk_row_status

    base = _gospd_job(
        _full_pack(),
        extra={
            "gospd_sample_results": {
                "SO25-0001": {
                    "fields_confirmed": True,
                    "matching_confirmed": True,
                    "three_way": {"overall_status": "PASS"},
                }
            }
        },
    )
    job, passed = apply_auto_conclusions_on_job(base)
    assert "SO25-0001" in passed
    assert job["gospd_sample_results"]["SO25-0001"]["conclusion_confirmed"] is True
    assert desk_row_status(job, "SO25-0001")["light"] == "green"

    fail_job = _gospd_job(
        _full_pack(),
        extra={
            "gospd_sample_results": {
                "SO25-0001": {
                    "fields_confirmed": True,
                    "matching_confirmed": True,
                    "three_way": {"overall_status": "FAIL"},
                }
            }
        },
    )
    nxt, passed_fail = apply_auto_conclusions_on_job(fail_job)
    assert passed_fail == []
    assert not nxt["gospd_sample_results"]["SO25-0001"].get("conclusion_confirmed")
    st = desk_row_status(nxt, "SO25-0001")
    assert st["reason"] == "test_fail"
    assert st["light"] == "red"


def test_desk_fail_closed_stays_red():
    from src.workflow.sample_desk import desk_row_status

    job = _gospd_job(
        _full_pack(),
        extra={
            "gospd_sample_results": {
                "SO25-0001": {
                    "fields_confirmed": True,
                    "matching_confirmed": True,
                    "conclusion_confirmed": True,
                    "conclusion_disposition": "fail",
                    "three_way": {"overall_status": "FAIL"},
                }
            }
        },
    )
    st = desk_row_status(job, "SO25-0001")
    assert st["reason"] == "fail_closed"
    assert st["light"] == "red"
    assert "已人工确认" in st["label"]


def test_desk_not_green_until_tests_run():
    from src.workflow.sample_desk import desk_row_status

    job = _gospd_job(
        _full_pack(),
        extra={
            "gospd_sample_results": {
                "SO25-0001": {
                    "fields_confirmed": True,
                    "matching_confirmed": True,
                }
            }
        },
    )
    st = desk_row_status(job, "SO25-0001")
    assert st["light"] == "wait"
    assert st["reason"] == "tests_pending"


def test_export_blocks_unfinished_population():
    from src.workflow.export_readiness import build_export_readiness

    job = _gospd_job(
        [_full_invoice()],
        pop_ids=["SO25-0001", "SO25-0002"],
        extra={
            "gospd_sample_results": {
                "SO25-0001": {
                    "fields_confirmed": True,
                    "matching_confirmed": True,
                    "conclusion_confirmed": True,
                    "three_way": {"overall_status": "PASS"},
                }
            }
        },
    )
    ready = build_export_readiness(job)
    assert ready["ready"] is False
    conc = next(s for s in ready["stages"] if s["id"] == "conclusion")
    assert "SO25-0002" in conc["affected_groups"]


def test_replace_sample_resets_tests_quarantines_old_outside_ocr(tmp_path: Path):
    """中途换清单：旧 OCR 文件留在异常区，但不得进入新清单业务。"""
    from src.workflow.job_store import JOB_STORE
    from src.workflow.sample_desk import replay_after_sample_replace

    xlsx = tmp_path / "new.xlsx"
    _write_sap_like(xlsx)
    parsed = parse_sample_workbook(xlsx)
    parsed["business_ids"] = ["SO25-0002"]
    parsed["rows"] = [r for r in parsed["rows"] if r.get("business_id") == "SO25-0002"]
    parsed["ledger_rows"] = [r for r in parsed["ledger_rows"] if r.get("business_id") == "SO25-0002"]

    job = JOB_STORE.create(title="replace-sample")
    jid = job["job_id"]
    JOB_STORE.update(
        jid,
        goal_ids=["gospd01030"],
        plan={"goal_ids": ["gospd01030"], "required_steps": ["evidence"]},
        classified=[
            {
                "file_name": "SO25-0001_inv.pdf",
                "doc_type": "invoice",
                "fields": {"orderNo": "SO25-0001", "totalAmount": "100"},
                "ledger_match_ok": True,
                "ledger_matched_biz_id": "SO25-0001",
                "ledger_amount": 2015.67,
            }
        ],
        gospd_sample_results={
            "SO25-0001": {
                "fields_confirmed": True,
                "amount_test": {"overall_status": "PASS"},
            }
        },
        fields_confirmed=True,
        sample_population=build_sample_population(
            business_ids=parsed["business_ids"],
            rows=parsed["rows"],
            source="excel",
        ),
        **ledger_patch_from_parsed(parsed, path=str(xlsx)),
    )
    nxt = replay_after_sample_replace(jid)
    docs = nxt["classified"]
    assert docs == []
    [scope_exception] = nxt.get("scope_exceptions") or []
    assert scope_exception["file_name"] == "SO25-0001_inv.pdf"
    assert scope_exception["scope_status"] == "OUT_OF_SAMPLE"
    assert scope_exception["document"]["file_name"] == "SO25-0001_inv.pdf"
    assert scope_exception["document"].get("ledger_match_ok") is False
    assert not scope_exception["document"].get("ledger_matched_biz_id")
    sample = (nxt.get("gospd_sample_results") or {}).get("SO25-0001") or {}
    assert (sample.get("amount_test") or {}).get("overall_status") != "PASS"
    assert nxt.get("fields_confirmed") is not True
