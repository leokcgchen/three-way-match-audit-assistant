"""GOSPD01010.3 交易价格底稿：配方 + 断言 + filler。"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from src.audit.gospd01010_3_assertions import (
    LABEL_K_NO,
    LABEL_M_NO,
    LABEL_M_YES,
    assert_needs_price_calculation,
    assert_transaction_price_ok,
    build_gospd01010_3_assertions,
)
from src.reporting.gospd01010_3_filler import (
    fill_gospd01010_3_workbook,
    resolve_template_path,
)
from src.workflow.pipeline import build_workbooks_for_job, selected_workbook_formats
from src.workflow.recipes import list_workpaper_goals, resolve_workflow_plan


def test_goal_listed_and_requires_amount_not_three_way():
    assert "gospd01010_3" in {g["goal_id"] for g in list_workpaper_goals()}
    plan = resolve_workflow_plan(["gospd01010_3"])
    assert "contract_terms" in plan["required_steps"]
    assert "amount_test" in plan["required_steps"]
    assert "three_way_cutoff" not in plan["required_steps"]
    assert plan["workbook_formats"] == ["gospd01010_3"]


def test_price_ok_uses_consideration_dimension():
    ok = assert_transaction_price_ok(
        has_contract=True,
        contract_res={
            "status": "WARNING",
            "extracted": {
                "dimension_statuses": {"交易对价": "CLEAR", "支付条款": "AMBIGUOUS"},
            },
        },
        amount={"status": "PASS"},
    )
    assert ok["verdict_label"] == LABEL_M_YES

    bad = assert_transaction_price_ok(
        has_contract=True,
        contract_res={
            "status": "WARNING",
            "extracted": {
                "issue_codes": ["REBATE_TERM_AMBIGUOUS"],
                "dimension_statuses": {"交易对价": "AMBIGUOUS"},
            },
        },
        amount={"status": "PASS"},
    )
    assert bad["verdict_label"] == LABEL_M_NO


def test_needs_calc_with_discount():
    out = assert_needs_price_calculation(
        docs_by_type={
            "order": {"fields": {"discountRate": "10%", "totalAmount": 1000}},
        },
        contract_res=None,
        amount=None,
    )
    assert out["verdict_label"] != LABEL_K_NO
    assert "discount" in out["calc_method"].lower() or "折扣" in out["calc_method"]


def test_filler_smoke(tmp_path: Path):
    assert resolve_template_path().is_file()
    job = {
        "job_id": "t-01010-3",
        "goal_ids": ["gospd01010_3"],
        "plan": resolve_workflow_plan(["gospd01010_3"]),
        "classified": [
            {
                "file_name": "c.pdf",
                "doc_type": "contract",
                "raw_text": "销售合同 固定单价 不含折扣",
                "fields": {
                    "contractNo": "HT25-0099",
                    "buyerName": "客户A",
                    "totalAmount": 11300,
                },
            },
            {
                "file_name": "o.pdf",
                "doc_type": "order",
                "fields": {"orderNo": "SO25-0099", "totalAmount": 11300},
            },
        ],
        "contract_terms": {
            "status": "PASS",
            "extracted": {"dimension_statuses": {"交易对价": "CLEAR"}},
        },
        "amount_test": {"status": "PASS"},
    }
    path = fill_gospd01010_3_workbook(job, tmp_path / "o.xlsx", entity_name="测")
    ws = load_workbook(path).active
    assert ws["E5"].value == "GOSPD01010.3"
    assert ws["B20"].value == 1
    assert ws["H20"].value == "Applicable 适用"
    assert ws["K20"].value == LABEL_K_NO
    assert ws["M20"].value == LABEL_M_YES
    assert ws["F20"].value == 11300


def test_multi_export_with_01010_3(tmp_path: Path, monkeypatch):
    monkeypatch.setattr("src.workflow.pipeline.workbook_output_dir", lambda: tmp_path)
    job = {
        "job_id": "m3",
        "goal_ids": ["gospd01010_3", "gospd01010_2"],
        "plan": resolve_workflow_plan(["gospd01010_3", "gospd01010_2"]),
        "classified": [
            {
                "file_name": "c.pdf",
                "doc_type": "contract",
                "raw_text": "合同",
                "fields": {"contractNo": "HT1", "totalAmount": 100},
            }
        ],
        "contract_terms": {
            "status": "PASS",
            "extracted": {
                "dimension_statuses": {"交易对价": "CLEAR", "履约义务": "CLEAR"}
            },
        },
        "amount_test": {"status": "PASS"},
    }
    assert selected_workbook_formats(job) == ["gospd01010_3", "gospd01010_2"]
    paths = build_workbooks_for_job(job)
    assert len(paths) == 2
    names = " ".join(p.name.upper() for p in paths)
    assert "01010.3" in names
    assert "01010.2" in names
