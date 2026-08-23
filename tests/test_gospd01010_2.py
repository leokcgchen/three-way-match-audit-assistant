"""GOSPD01010.2 履约义务底稿：配方 + 断言 + filler 冒烟。"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from src.audit.gospd01010_2_assertions import (
    assert_other_files_for_price,
    assert_performance_obligation,
    build_gospd01010_2_assertions,
)
from src.reporting.gospd01010_2_filler import (
    fill_gospd01010_2_workbook,
    resolve_template_path,
)
from src.workflow.pipeline import build_workbooks_for_job, selected_workbook_formats
from src.workflow.recipes import list_workpaper_goals, resolve_workflow_plan


def test_goal_listed_and_plan_skips_amount_three_way():
    ids = {g["goal_id"] for g in list_workpaper_goals()}
    assert "gospd01010_2" in ids
    plan = resolve_workflow_plan(["gospd01010_2"])
    assert "contract_terms" in plan["required_steps"]
    assert "amount_test" not in plan["required_steps"]
    assert "three_way_cutoff" not in plan["required_steps"]
    assert plan["workbook_formats"] == ["gospd01010_2"]
    assert "GOSPD01010.2" in plan["workbook_sheets"]


def test_multi_select_includes_01010_2_without_bias():
    plan = resolve_workflow_plan(["gospd01010_2", "gospd01030"])
    assert plan["workbook_formats"] == ["gospd01010_2", "gospd01030"]
    assert "contract_terms" in plan["required_steps"]
    assert "three_way_cutoff" in plan["required_steps"]


def test_performance_obligation_pass_fail():
    ok = assert_performance_obligation(
        has_contract=True,
        contract_res={
            "status": "PASS",
            "extracted": {"dimension_statuses": {"履约义务": "CLEAR"}},
        },
    )
    assert ok["verdict_label"] == "YES 是"

    # 整单 WARNING（支付条款）但履约义务 CLEAR → 仍 YES（01010.2 专用）
    mixed = assert_performance_obligation(
        has_contract=True,
        contract_res={
            "status": "WARNING",
            "extracted": {
                "issue_codes": ["PAYMENT_TERMS_MISSING"],
                "dimension_statuses": {
                    "履约义务": "CLEAR",
                    "支付条款": "AMBIGUOUS",
                },
            },
        },
    )
    assert mixed["verdict_label"] == "YES 是"
    assert mixed["dimension_status"] == "CLEAR"

    bad = assert_performance_obligation(
        has_contract=True,
        contract_res={
            "status": "WARNING",
            "extracted": {
                "issue_codes": ["PERFORMANCE_OBLIGATION_BOUNDARY_UNCLEAR"],
                "dimension_statuses": {"履约义务": "AMBIGUOUS"},
            },
        },
    )
    assert bad["verdict_label"] == "NO 否"

    missing = assert_performance_obligation(has_contract=False, contract_res=None)
    assert missing["verdict_label"] == "NO 否"


def test_performance_obligation_reads_dimensioned_issues():
    """无 dimension_statuses 时，仍可从 issues.dimension=履约义务判定 NO。"""
    bad = assert_performance_obligation(
        has_contract=True,
        contract_res={
            "status": "WARNING",
            "clarity_report": {
                "test_result": {
                    "test_status": "WARNING",
                    "issues": [
                        {
                            "issue_code": "PERFORMANCE_OBLIGATION_BOUNDARY_UNCLEAR",
                            "dimension": "履约义务",
                            "description": "边界不清",
                        }
                    ],
                }
            },
        },
    )
    assert bad["verdict_label"] == "NO 否"
    assert "PERFORMANCE_OBLIGATION_BOUNDARY_UNCLEAR" in bad["issue_codes"]


def test_other_files_applicable_when_order_present():
    out = assert_other_files_for_price(
        docs_by_type={
            "order": {
                "doc_type": "order",
                "file_name": "o.pdf",
                "fields": {"orderNo": "SO25-0281", "totalAmount": 1000},
            }
        }
    )
    assert out["applicable_label"] == "Applicable 适用"
    assert "订单" in out["file_type"]
    assert "SO25-0281" in out["file_index"]


def test_filler_smoke(tmp_path: Path):
    assert resolve_template_path().is_file()
    job = {
        "job_id": "test-01010-2",
        "goal_ids": ["gospd01010_2"],
        "plan": resolve_workflow_plan(["gospd01010_2"]),
        "fields_confirmed": True,
        "classified": [
            {
                "file_name": "c.pdf",
                "doc_type": "contract",
                "raw_text": "销售合同 履约义务为交付货物 签收后付款",
                "fields": {
                    "contractNo": "HT25-0281",
                    "buyerName": "测试客户",
                    "paymentTerms": "签收后30日",
                    "controlTransferTerms": "签收转移",
                    "totalAmount": 52904.58,
                },
            },
            {
                "file_name": "o.pdf",
                "doc_type": "order",
                "fields": {
                    "orderNo": "SO25-0281",
                    "totalAmount": 52904.58,
                    "buyerName": "测试客户",
                },
            },
        ],
        "contract_terms": {"status": "PASS"},
    }
    out = fill_gospd01010_2_workbook(job, tmp_path / "out.xlsx", entity_name="单元测试主体")
    assert out.is_file()
    ws = load_workbook(out).active
    assert ws["E5"].value == "GOSPD01010.2"
    assert ws["B20"].value == 1
    assert "HT25" in str(ws["D20"].value or "")
    assert ws["H20"].value == "Applicable 适用"
    assert ws["K20"].value == "YES 是"
    assert ws["G20"].value == 52904.58


def test_export_two_formats_no_bias(tmp_path: Path, monkeypatch):
    monkeypatch.setattr(
        "src.workflow.pipeline.workbook_output_dir",
        lambda: tmp_path,
    )
    job = {
        "job_id": "multi-fmt",
        "goal_ids": ["gospd01010_2", "gospd01030"],
        "plan": resolve_workflow_plan(["gospd01010_2", "gospd01030"]),
        "classified": [
            {
                "file_name": "c.pdf",
                "doc_type": "contract",
                "raw_text": "合同",
                "fields": {"contractNo": "HT1", "totalAmount": 100},
            },
            {
                "file_name": "i.pdf",
                "doc_type": "invoice",
                "fields": {
                    "invoiceNo": "INV1",
                    "postingDate": "2026-01-10",
                    "totalAmount": 100,
                },
                "ledger_posting_date": "2026-01-10",
            },
            {
                "file_name": "r.pdf",
                "doc_type": "receipt",
                "fields": {"acceptanceDate": "2026-01-08", "totalAmount": 100},
            },
        ],
        "contract_terms": {"status": "PASS"},
        "three_way": {
            "overall_status": "PASS",
            "match_result": {"overall_status": "PASS"},
            "cutoff_result": {"测试状态": "PASS"},
        },
        "period_end": "2025-12-31",
    }
    assert selected_workbook_formats(job) == ["gospd01010_2", "gospd01030"]
    paths = build_workbooks_for_job(job)
    assert len(paths) == 2
    names = [p.name.upper() for p in paths]
    assert any("01010.2" in n or "01010_2" in n for n in names)
    assert any(n.startswith("GOSPD01030") for n in names)
