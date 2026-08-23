"""GOSPD01030 期后截止断言与填表。"""

from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook

from src.audit.gospd01030_assertions import (
    assert_ar_correct_period,
    assert_correct_accounting_period,
    build_gospd01030_assertions,
)
from src.reporting.gospd01010_filler import group_classified_by_chain
from src.reporting.gospd01030_filler import (
    LOG_SHEET,
    W_NO_FALLBACK,
    W_YES_FALLBACK,
    build_gospd01030_sample_rows,
    fill_gospd01030_workbook,
    resolve_template_path,
)
from src.workflow.recipes import resolve_workflow_plan


def _base_docs(
    *,
    acceptance_date: str = "2026-01-02",
    posting_date: str = "2026-01-03",
    receipt_amount: Optional[float] = None,
    include_receipt_date: bool = True,
) -> list[dict[str, Any]]:
    receipt_fields: dict[str, Any] = {
        "documentNo": "R-1",
        "orderNo": "SO25-0099",
        "quantity": 10,
    }
    if include_receipt_date:
        receipt_fields["acceptanceDate"] = acceptance_date
    if receipt_amount is not None:
        receipt_fields["totalAmount"] = receipt_amount
    return [
        {
            "file_name": "c.pdf",
            "doc_type": "contract",
            "fields": {
                "contractNo": "HT25-0099",
                "transportTerms": "签收确认",
                "buyerName": "测试客户",
            },
        },
        {
            "file_name": "o.pdf",
            "doc_type": "order",
            "fields": {
                "orderNo": "SO25-0099",
                "contractNo": "HT25-0099",
                "quantity": 10,
                "totalAmount": 1130,
            },
        },
        {
            "file_name": "r.pdf",
            "doc_type": "receipt",
            "fields": receipt_fields,
        },
        {
            "file_name": "i.pdf",
            "doc_type": "invoice",
            "fields": {
                "invoiceNo": "INV-1",
                "orderNo": "SO25-0099",
                "contractNo": "HT25-0099",
                "totalAmount": 1130,
                "quantity": 10,
                "buyerName": "测试客户",
            },
            "ledger_voucher": "记-099",
            "ledger_amount": 1130,
            "ledger_posting_date": posting_date,
        },
    ]


def _job(
    docs: list[dict[str, Any]],
    *,
    three_way: Optional[dict[str, Any]] = None,
    period_end: str = "2025-12-31",
    matching_confirmed: bool = True,
) -> dict[str, Any]:
    tw = three_way or {
        "overall_status": "PASS",
        "match_result": {"overall_status": "PASS"},
        "cutoff_result": {"测试状态": "PASS"},
    }
    return {
        "goal_ids": ["gospd01030"],
        "plan": {"goal_ids": ["gospd01030"], "required_steps": ["three_way_cutoff"]},
        "period_end": period_end,
        "classified": docs,
        "matching_confirmed": matching_confirmed,
        "three_way": tw,
        "gospd_sample_results": {
            "SO25-0099": {
                "matching_confirmed": matching_confirmed,
                "three_way": tw,
            }
        },
    }


def _w_dv_labels() -> tuple[str, str]:
    ws = load_workbook(resolve_template_path()).active
    yes_v, no_v = W_YES_FALLBACK, W_NO_FALLBACK
    for dv in ws.data_validations.dataValidation:
        if "W" not in str(dv.sqref or ""):
            continue
        raw = str(dv.formula1 or "").strip().strip('"')
        parts = [p.strip() for p in raw.split(",") if p.strip()]
        for p in parts:
            pu = p.upper()
            if "YES" in pu:
                yes_v = p
            elif "NO" in pu or "否" in p:
                no_v = p
        break
    return yes_v, no_v


def test_gospd01030_recipe_steps():
    plan = resolve_workflow_plan(["gospd01030"])
    assert plan["required_steps"] == [
        "upload_ocr",
        "field_confirm",
        "evidence_match",
        "relations_gate4",
        "three_way_cutoff",
        "conclusion_gate5",
        "workbook_export",
    ]
    assert "contract_terms" in plan["skipped_steps"]
    assert "amount_test" in plan["skipped_steps"]
    assert "GOSPD01030" in plan["workbook_sheets"]
    assert "gospd01030" in plan["workbook_formats"]
    assert "期后" in plan["note"]
    assert "裁剪序时账" in plan["note"]
    assert "Gate4" not in plan["note"]
    labels = [s["label"] for s in plan["step_labels"]]
    assert "上传凭证" in labels
    assert "上传与识别" not in labels


def test_period_post_period_wrong():
    r = assert_correct_accounting_period(
        posting_date="2026-01-03",
        control_date="2025-12-28",
        period_end=date(2025, 12, 31),
        cutoff_status="PASS",
    )
    assert r["verdict"] is False
    assert r["verdict_label"] == "No 否"


def test_period_post_period_ok():
    r = assert_correct_accounting_period(
        posting_date="2026-01-03",
        control_date="2026-01-02",
        period_end=date(2025, 12, 31),
        cutoff_status="PASS",
    )
    assert r["verdict"] is True
    assert r["verdict_label"] == "YES 是"


def test_period_end_missing_is_not_silent_default():
    """job 未给期末时不得静默默认某年 12-31；期间结论应为 NOT_TESTED。"""
    from src.audit.gospd01030_assertions import resolve_period_end

    job = _job(_base_docs())
    job.pop("period_end", None)
    assert resolve_period_end(job) is None
    r = assert_correct_accounting_period(
        posting_date="2026-01-03",
        control_date="2026-01-02",
        period_end=None,
        cutoff_status="PASS",
    )
    assert r["verdict"] is None
    assert r.get("evidence_status") == "NOT_TESTED"
    out = fill_gospd01030_workbook(
        job, Path("D:/抽凭—合同合规性审阅agent") / "_tmp_pe.xlsx"
    )
    try:
        wb = load_workbook(out)
        ws = wb.active
        assert ws["M5"].value in (None, "")
        log = wb[LOG_SHEET]
        texts = [
            str(log.cell(r, c).value or "")
            for r in range(1, log.max_row + 1)
            for c in range(1, 8)
        ]
        assert any("NOT_TESTED" in t for t in texts)
    finally:
        out.unlink(missing_ok=True)


def test_skip_unconfirmed_gate4_chains():
    """Gate4 未确认的链不得写入正式样本行。"""
    job = _job(_base_docs(), matching_confirmed=False)
    skipped: list = []
    rows = build_gospd01030_sample_rows(job, skipped_chains=skipped)
    assert rows == []
    assert skipped and skipped[0]["status"] == "NOT_TESTED"


def test_voucher_rejects_order_no_as_document():
    job = _job(_base_docs())
    job["classified"][3].pop("ledger_voucher", None)
    job["classified"][3]["fields"]["documentNo"] = "SO25-0099"
    job["classified"][3]["fields"]["voucherNo"] = ""
    rows = build_gospd01030_sample_rows(job)
    assert rows[0]["voucher"] == ""


def test_period_early_recognition():
    r = assert_correct_accounting_period(
        posting_date="2025-12-20",
        control_date="2026-01-05",
        period_end=date(2025, 12, 31),
        cutoff_status="PASS",
    )
    assert r["verdict"] is False
    assert r["verdict_label"] == "No 否"


def test_period_insufficient_evidence():
    r = assert_correct_accounting_period(
        posting_date="2026-01-03",
        control_date=None,
        period_end=date(2025, 12, 31),
        cutoff_status="PASS",
    )
    assert r["verdict"] is None
    assert r["verdict_label"] == ""
    assert r.get("evidence_status") == "INSUFFICIENT_EVIDENCE"
    assert any("控制权" in g for g in r["gaps"])


def test_fill_gospd01030(tmp_path: Path):
    assert resolve_template_path().is_file()
    w_yes, _w_no = _w_dv_labels()

    job = _job(_base_docs(receipt_amount=None))
    job["entity_name"] = "演示被审计单位"
    job["period_end"] = "2025-12-31"
    # 订单号不得误入凭证列
    job["classified"][3]["fields"]["documentNo"] = "SO25-0099"

    rows = build_gospd01030_sample_rows(job)
    assert rows
    assert rows[0]["order_no"] == "SO25-0099"
    assert rows[0]["period_ok"] == "YES 是"
    assert rows[0]["amt_delivery"] is None
    assert rows[0]["voucher"] == "记-099"
    assert rows[0]["formula_v"] == "YES 是"

    a = build_gospd01030_assertions(
        docs=job["classified"], job=job, three_way=job["three_way"]
    )
    assert a["period"]["verdict"] is True

    out = fill_gospd01030_workbook(job, tmp_path / "out030.xlsx", entity_name="演示被审计单位")
    assert out.is_file()

    wb = load_workbook(out)
    ws = wb.active
    assert ws["F5"].value == "GOSPD01030"
    assert ws["C5"].value == "演示被审计单位"
    assert ws["I5"].value == "人民币"
    assert ws["K5"].value == "Yuan 元"
    assert ws["M5"].value == date(2025, 12, 31) or (
        hasattr(ws["M5"].value, "date") and ws["M5"].value.date() == date(2025, 12, 31)
    )
    assert ws["B30"].value == 1
    assert ws["C30"].value == "记-099"
    assert ws["L30"].value == "SO25-0099"
    assert ws["M30"].value == "签收确认"
    # V/K/S/T 必须保留公式
    assert str(ws["V30"].value).startswith("=")
    assert "$M$5" in str(ws["V30"].value)
    assert str(ws["K30"].value).startswith("=")
    assert str(ws["S30"].value).startswith("=")
    assert str(ws["T30"].value).startswith("=")
    # R 无金额不填 0
    assert ws["R30"].value in (None, "")
    # W 写 DV 精确「是」
    assert ws["W30"].value == w_yes
    # E13 运输条款；E14 是否无需检查系统发票；E15 是否检查销售订单
    assert ws["E13"].value == "签收确认"
    assert ws["E14"].value == "No 否"  # 样本有发票 → 需要检查
    assert ws["E15"].value == w_yes
    # 辅助日志页
    assert LOG_SHEET in wb.sheetnames
    log = wb[LOG_SHEET]
    log_text = "\n".join(
        str(log.cell(r, c).value or "")
        for r in range(1, log.max_row + 1)
        for c in range(1, 8)
    )
    assert "TEMPLATE_VALIDATION_CONFLICT" not in log_text
    assert "2025-12-31" in log_text or "期间截止日" in log_text
    assert "运输条款" in log_text or "E13" in log_text
    # 系统说明不得写进样本区 B30:B48
    assert ws["B48"].value in (None, "", 19) or (
        isinstance(ws["B48"].value, (int, float)) or str(ws["B48"].value).isdigit()
    )
    assert "自动回填" not in str(ws["B48"].value or "")
    assert any(
        "自动回填" in str(log.cell(r, 4).value or "")
        for r in range(1, log.max_row + 1)
    )
    # Sheet1 注释区下方可有说明，但不在样本网格内
    assert "自动回填" in str(ws["B53"].value or "")


def test_fill_w_uses_long_no_when_exception(tmp_path: Path):
    """期内发货+期后入账 → 独立 No；W 须写模板 DV 长文案「否」。"""
    _, w_no = _w_dv_labels()
    assert "Document" in w_no or "否" in w_no

    job = _job(
        _base_docs(
            acceptance_date="2025-12-28",
            posting_date="2026-01-03",
            receipt_amount=1130,
        )
    )
    rows = build_gospd01030_sample_rows(job)
    assert rows[0]["period_ok"] == "No 否"
    assert rows[0]["formula_v"] == "No 否"
    assert not rows[0]["formula_conflict"]
    assert rows[0]["all_ok"] == w_no

    out = fill_gospd01030_workbook(job, tmp_path / "w_no.xlsx")
    ws = load_workbook(out).active
    assert ws["W30"].value == w_no
    assert ws["W30"].value != "No 否"  # 禁止写普通短文案绕过 DV
    assert ws["X30"].value  # 有异常说明
    assert ws["R30"].value == 1130  # 有可比金额才写 R


def test_fill_formula_logic_conflict_blocks_w_yes(tmp_path: Path):
    """期内过账且控制权均在期内：独立 YES，V 公式(P>$M$5)=No → 冲突，W 不写「是」。"""
    w_yes, _ = _w_dv_labels()
    job = _job(
        _base_docs(
            acceptance_date="2025-12-15",
            posting_date="2025-12-20",
            receipt_amount=1130,
        )
    )
    rows = build_gospd01030_sample_rows(job)
    assert rows[0]["period_ok"] == "YES 是"
    assert rows[0]["formula_v"] == "No 否"
    assert "FORMULA_LOGIC_CONFLICT" in str(rows[0]["formula_conflict"])
    assert rows[0]["all_ok"] == ""

    out = fill_gospd01030_workbook(job, tmp_path / "conflict.xlsx")
    wb = load_workbook(out)
    ws = wb.active
    assert str(ws["V30"].value).startswith("=")  # 仍保留公式
    assert ws["W30"].value in (None, "")
    assert ws["W30"].value != w_yes
    x = str(ws["X30"].value or "")
    assert x
    assert "FORMULA_LOGIC_CONFLICT" not in x  # X 列改为自然语言
    assert "复核" in x or "不一致" in x or "公式" in x

    log = wb[LOG_SHEET]
    statuses = [
        str(log.cell(r, 6).value or "") for r in range(2, log.max_row + 1)
    ]
    assert "FORMULA_LOGIC_CONFLICT" in statuses


def test_fill_insufficient_evidence_leaves_w_blank(tmp_path: Path):
    """缺签收日 → 期间结论空；W 不得写「是」。"""
    w_yes, _ = _w_dv_labels()
    job = _job(_base_docs(include_receipt_date=False, receipt_amount=None))
    rows = build_gospd01030_sample_rows(job)
    assert rows[0]["period_ok"] == ""
    assert rows[0]["all_ok"] != w_yes

    out = fill_gospd01030_workbook(job, tmp_path / "insuff.xlsx")
    ws = load_workbook(out).active
    assert str(ws["V30"].value).startswith("=")
    assert ws["W30"].value != w_yes
    # 证据缺口应进入 X 或至少不宣称无异常
    assert ws["W30"].value in (None, "",) or "否" in str(ws["W30"].value)


def test_ar_period_follows_revenue_fail():
    """步骤3：收入期间否 → 应收亦否。"""
    r = assert_ar_correct_period(
        posting_date="2026-01-03",
        control_date="2025-12-28",
        period_end=date(2025, 12, 31),
        cutoff_status="PASS",
        revenue_period_verdict=False,
    )
    assert r["verdict"] is False
    assert any("应收" in g or "收入" in g for g in r["gaps"])


def test_assertions_cutoff_fail_does_not_flag_three_way():
    docs = _base_docs(acceptance_date="2025-12-20", posting_date="2026-01-05")
    job = _job(
        docs,
        three_way={
            "overall_status": "FAIL",
            "three_way_status": "PASS",
            "cutoff_status": "FAIL",
            "match_result": {"overall_status": "PASS"},
            "cutoff_result": {"测试状态": "FAIL"},
        },
        period_end="2025-12-31",
    )
    a = build_gospd01030_assertions(
        docs=docs, job=job, three_way=job["three_way"]
    )
    assert a["match_status"] != "FAIL"
    assert "三单匹配未通过" not in (a.get("gaps") or [])
    assert any("截止性未通过" in str(g) for g in (a.get("gaps") or []))


def test_ar_period_in_build_assertions():
    job = _job(_base_docs())
    a = build_gospd01030_assertions(
        docs=job["classified"], job=job, three_way=job["three_way"]
    )
    assert a["period"]["verdict"] is True
    assert a["ar_period"]["verdict"] is True
    assert a["ar_period_label"] == "YES 是"
    rows = build_gospd01030_sample_rows(job)
    assert rows[0]["ar_period_ok"] == "YES 是"


def test_strict_chain_rejects_weak_unique_attach():
    """01030：仅有弱编号的签收不得因「唯一桶」猜测并入 SO 链。"""
    classified = [
        {
            "file_name": "o.pdf",
            "doc_type": "order",
            "fields": {"orderNo": "SO25-0099", "contractNo": "HT25-0099"},
        },
        {
            "file_name": "r.pdf",
            "doc_type": "receipt",
            "fields": {"documentNo": "R-ONLY", "acceptanceDate": "2026-01-02"},
        },
    ]
    loose = group_classified_by_chain(classified)
    loose_map = {cid: docs for cid, docs in loose}
    assert "SO25-0099" in loose_map
    assert any(d["file_name"] == "r.pdf" for d in loose_map["SO25-0099"])

    strict = group_classified_by_chain(
        classified,
        allow_weak_unique_attach=False,
        allow_unique_so_ht_merge=False,
    )
    strict_map = {cid: docs for cid, docs in strict}
    assert "SO25-0099" in strict_map
    assert all(d["file_name"] != "r.pdf" for d in strict_map["SO25-0099"])
    assert "未识别业务号" in strict_map
    assert any(d["file_name"] == "r.pdf" for d in strict_map["未识别业务号"])


def test_strict_chain_no_blind_so_ht_merge():
    """01030：无互相引用时，唯一 SO+唯一 HT 不得盲合并。"""
    classified = [
        {
            "file_name": "o.pdf",
            "doc_type": "order",
            "fields": {"orderNo": "SO25-AAAA"},
        },
        {
            "file_name": "c.pdf",
            "doc_type": "contract",
            "fields": {"contractNo": "HT25-BBBB"},
        },
    ]
    loose = dict(group_classified_by_chain(classified))
    assert len(loose) == 1  # 便利模式合并

    strict = dict(
        group_classified_by_chain(
            classified,
            allow_weak_unique_attach=False,
            allow_unique_so_ht_merge=False,
        )
    )
    assert "SO25-AAAA" in strict
    assert "HT25-BBBB" in strict
    assert len(strict) == 2


def test_dual_official_export_no_bias(tmp_path: Path, monkeypatch):
    """双选官方模板：按勾选顺序各出一份，无 01030 优先覆盖。"""
    from src.workflow.pipeline import build_workbooks_for_job, selected_workbook_formats

    monkeypatch.setenv("WORKBOOK_OUTPUT_DIR", str(tmp_path))
    import src.workflow.pipeline as pipe

    monkeypatch.setattr(pipe, "workbook_output_dir", lambda: tmp_path)

    assert selected_workbook_formats(
        {"goal_ids": ["gospd01030", "gospd01010"]}
    ) == ["gospd01030", "gospd01010"]
    assert selected_workbook_formats(
        {"goal_ids": ["gospd01010", "gospd01030"]}
    ) == ["gospd01010", "gospd01030"]

    job = {
        "job_id": "dualtest01",
        "goal_ids": ["gospd01010", "gospd01030"],
        "plan": {
            "goal_ids": ["gospd01010", "gospd01030"],
            "required_steps": [
                "contract_terms",
                "amount_test",
                "three_way_cutoff",
            ],
        },
        "period_end": "2025-12-31",
        "matching_confirmed": True,
        "classified": [
            {
                "file_name": "c.pdf",
                "doc_type": "contract",
                "raw_text": "销售合同 双方盖章 商业实质 购销 签收后30日付款",
                "fields": {
                    "contractNo": "HT25-0099",
                    "paymentTerms": "签收后30日",
                    "transportTerms": "签收确认",
                    "buyerName": "测试客户",
                },
            },
            {
                "file_name": "o.pdf",
                "doc_type": "order",
                "fields": {
                    "orderNo": "SO25-0099",
                    "contractNo": "HT25-0099",
                    "quantity": 10,
                    "totalAmount": 1130,
                },
            },
            {
                "file_name": "r.pdf",
                "doc_type": "receipt",
                "fields": {
                    "documentNo": "R-1",
                    "orderNo": "SO25-0099",
                    "acceptanceDate": "2026-01-02",
                    "quantity": 10,
                    "totalAmount": 1130,
                },
            },
            {
                "file_name": "i.pdf",
                "doc_type": "invoice",
                "fields": {
                    "invoiceNo": "INV-1",
                    "orderNo": "SO25-0099",
                    "contractNo": "HT25-0099",
                    "totalAmount": 1130,
                    "quantity": 10,
                    "buyerName": "测试客户",
                },
                "ledger_voucher": "记-099",
                "ledger_amount": 1130,
                "ledger_posting_date": "2026-01-03",
            },
            {
                "file_name": "p.pdf",
                "doc_type": "payment",
                "fields": {"totalAmount": 1130, "orderNo": "SO25-0099"},
            },
        ],
        "gospd_sample_results": {
            "SO25-0099": {"matching_confirmed": True},
        },
        "contract_terms": {"status": "PASS"},
        "amount_test": {"status": "PASS"},
        "three_way": {
            "overall_status": "PASS",
            "match_result": {"overall_status": "PASS"},
            "cutoff_result": {"测试状态": "PASS"},
        },
    }
    paths = build_workbooks_for_job(job)
    assert len(paths) == 2
    names = [p.name for p in paths]
    assert names[0].startswith("GOSPD01010_")
    assert names[1].startswith("GOSPD01030_")
    assert all(p.is_file() for p in paths)


def test_e13_leaves_blank_when_samples_disagree(tmp_path: Path):
    """评量 E13：多样本运输条款不一致时留空，不以首行冒充。"""
    docs_a = _base_docs()
    docs_b = _base_docs()
    for d in docs_b:
        fields = d.get("fields") or {}
        for k in ("orderNo", "contractNo"):
            if k in fields and "0099" in str(fields[k]):
                fields[k] = str(fields[k]).replace("0099", "0088")
        if d.get("doc_type") == "contract":
            fields["transportTerms"] = "外销-FOB离岸价格"
            d["raw_text"] = "贸易术语 FOB 上海。风险在装船时转移。"
        d["fields"] = fields
    job = {
        "goal_ids": ["gospd01030"],
        "plan": {"goal_ids": ["gospd01030"], "required_steps": ["three_way_cutoff"]},
        "period_end": "2025-12-31",
        "classified": docs_a + docs_b,
        "matching_confirmed": True,
        "gospd_sample_results": {
            "SO25-0099": {
                "matching_confirmed": True,
                "three_way": {
                    "overall_status": "PASS",
                    "match_result": {"overall_status": "PASS"},
                    "cutoff_result": {"测试状态": "PASS"},
                },
            },
            "SO25-0088": {
                "matching_confirmed": True,
                "three_way": {
                    "overall_status": "PASS",
                    "match_result": {"overall_status": "PASS"},
                    "cutoff_result": {"测试状态": "PASS"},
                },
            },
        },
        "three_way": {
            "overall_status": "PASS",
            "match_result": {"overall_status": "PASS"},
            "cutoff_result": {"测试状态": "PASS"},
        },
    }
    out = fill_gospd01030_workbook(job, tmp_path / "e13_mix.xlsx")
    ws = load_workbook(out).active
    assert ws["E13"].value in (None, "")
    assert ws["M30"].value  # 行级仍有条款
    assert ws["M31"].value


def test_formula_conflict_names_missing_m5():
    """独立期间已有结论但未配 period_end 时，冲突文案须点名 M5。"""
    job = _job(
        _base_docs(acceptance_date="2025-12-28", posting_date="2026-01-03"),
        three_way={
            "overall_status": "FAIL",
            "three_way_status": "PASS",
            "cutoff_status": "FAIL",
            "match_result": {"overall_status": "PASS"},
            "cutoff_result": {"测试状态": "FAIL"},
        },
    )
    job.pop("period_end", None)
    rows = build_gospd01030_sample_rows(job, period_end=None)
    assert rows[0]["period_ok"] == "No 否"
    assert rows[0]["formula_v"] is None
    assert "M5" in str(rows[0]["formula_conflict"]) or "period_end" in str(
        rows[0]["formula_conflict"]
    )


def test_export_fob_without_on_board_leaves_p_blank():
    """外销 FOB 无装船日时，P 不得用仓库签收日冒充。"""
    docs = [
        {
            "file_name": "c.pdf",
            "doc_type": "contract",
            "raw_text": "销售合同 贸易术语 FOB 上海。买方订舱。风险在装船时转移。",
            "fields": {
                "contractNo": "HT25-0099",
                "transportTerms": "FOB Shanghai",
                "buyerName": "Overseas Buyer",
            },
        },
        {
            "file_name": "o.pdf",
            "doc_type": "order",
            "fields": {
                "orderNo": "SO25-0099",
                "contractNo": "HT25-0099",
                "quantity": 10,
                "totalAmount": 1130,
            },
        },
        {
            "file_name": "r.pdf",
            "doc_type": "receipt",
            "fields": {
                "documentNo": "WH-1",
                "orderNo": "SO25-0099",
                "acceptanceDate": "2025-12-20",
                "quantity": 10,
            },
        },
        {
            "file_name": "i.pdf",
            "doc_type": "invoice",
            "fields": {
                "invoiceNo": "CI-1",
                "orderNo": "SO25-0099",
                "contractNo": "HT25-0099",
                "totalAmount": 1130,
                "quantity": 10,
                "buyerName": "Overseas Buyer",
            },
            "ledger_amount": 1130,
            "ledger_posting_date": "2025-12-21",
        },
    ]
    job = _job(docs)
    rows = build_gospd01030_sample_rows(job)
    assert "FOB" in str(rows[0].get("transport") or "")
    assert rows[0]["receipt_date"] is None
    assert "禁止用签收日冒充" in str(rows[0].get("exception") or "")


def test_clear_unused_preset_sample_rows(tmp_path: Path):
    job = _job(_base_docs())
    out = fill_gospd01030_workbook(job, tmp_path / "clear_preset.xlsx")
    ws = load_workbook(out).active
    # 仅写入 1 行样本；预置第 7 样本行不得残留「签收确认」
    assert ws["M36"].value in (None, "")
    assert ws["N36"].value in (None, "")
    assert ws["W36"].value in (None, "")
    assert str(ws["V36"].value or "").startswith("=")
