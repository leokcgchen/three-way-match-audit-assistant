"""GOSPD01010 模板填表冒烟。"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from src.reporting.gospd01010_filler import (
    build_gospd_sample_rows,
    fill_gospd01010_workbook,
    resolve_template_path,
)


def test_template_exists():
    assert resolve_template_path().is_file()


def test_fill_sample_row(tmp_path: Path):
    job = {
        "goal_ids": ["gospd01010"],
        "plan": {"goal_ids": ["gospd01010"]},
        "classified": [
            {
                "file_name": "c.pdf",
                "doc_type": "contract",
                "raw_text": "销售合同 双方盖章 商业实质 购销",
                "fields": {
                    "contractNo": "HT25-0281",
                    "paymentTerms": "签收后30日付款",
                    "controlTransferTerms": "签收后转移控制权",
                    "transportTerms": "签收确认",
                    "buyerName": "测试客户",
                },
            },
            {
                "file_name": "o.pdf",
                "doc_type": "order",
                "fields": {
                    "orderNo": "SO25-0281",
                    "quantity": 10,
                    "totalAmount": 1130,
                },
            },
            {
                "file_name": "r.pdf",
                "doc_type": "receipt",
                "fields": {
                    "documentNo": "R-001",
                    "acceptanceDate": "2025-01-08",
                    "quantity": 10,
                },
            },
            {
                "file_name": "i.pdf",
                "doc_type": "invoice",
                "fields": {
                    "invoiceNo": "INV-1",
                    "orderNo": "SO25-0281",
                    "totalAmount": 1130,
                    "quantity": 10,
                    "buyerName": "测试客户",
                },
                "ledger_voucher": "记-001",
                "ledger_amount": 1130,
            },
            {
                "file_name": "p.pdf",
                "doc_type": "payment",
                "fields": {"totalAmount": 1130},
            },
        ],
        "contract_terms": {"status": "PASS", "human_readable_summary": "条款清晰"},
        "amount_test": {"status": "PASS", "human_readable_summary": "金额一致"},
        "three_way": {
            "overall_status": "PASS",
            "match_result": {"overall_status": "PASS"},
            "cutoff_result": {"测试状态": "PASS"},
        },
    }
    rows = build_gospd_sample_rows(job)
    assert rows[0]["order_no"] == "SO25-0281"
    assert rows[0]["step1"].startswith("Yes")
    assert rows[0]["step21"].startswith("Yes")
    assert rows[0]["step22"].startswith("Yes")

    out = fill_gospd01010_workbook(job, tmp_path / "out.xlsx", entity_name="演示公司")
    assert out.is_file()
    wb = load_workbook(out)
    ws = wb.active
    assert ws["B22"].value == 1
    assert ws["I22"].value == "SO25-0281"
    assert str(ws["R22"].value).startswith("Yes")
    assert ws["P22"].value == "=E22-N22"


def test_multi_business_same_workbook(tmp_path: Path):
    """同 job 两笔业务 → 同一份底稿两行。"""
    from src.reporting.gospd01010_filler import group_classified_by_chain

    classified = [
        {
            "file_name": "SO25-0281_HT25-0281_01_销售合同.pdf",
            "doc_type": "contract",
            "fields": {"contractNo": "HT25-0281", "buyerName": "客户甲"},
        },
        {
            "file_name": "SO25-0281_HT25-0281_02_销售订单.pdf",
            "doc_type": "order",
            "fields": {"orderNo": "SO25-0281", "totalAmount": 1000, "quantity": 1},
        },
        {
            "file_name": "SO25-0281_HT25-0281_05_增值税发票.pdf",
            "doc_type": "invoice",
            "fields": {"orderNo": "SO25-0281", "totalAmount": 1000, "buyerName": "客户甲"},
        },
        {
            "file_name": "SO25-0099_HT25-0099_01_销售合同.pdf",
            "doc_type": "contract",
            "fields": {"contractNo": "HT25-0099", "buyerName": "客户乙"},
        },
        {
            "file_name": "SO25-0099_HT25-0099_02_销售订单.pdf",
            "doc_type": "order",
            "fields": {"orderNo": "SO25-0099", "totalAmount": 2000, "quantity": 2},
        },
        {
            "file_name": "SO25-0099_HT25-0099_05_增值税发票.pdf",
            "doc_type": "invoice",
            "fields": {"orderNo": "SO25-0099", "totalAmount": 2000, "buyerName": "客户乙"},
        },
    ]
    chains = group_classified_by_chain(classified)
    assert len(chains) == 2
    ids = {c[0] for c in chains}
    assert "SO25-0281" in ids
    assert "SO25-0099" in ids

    job = {
        "goal_ids": ["gospd01010"],
        "plan": {"goal_ids": ["gospd01010"]},
        "classified": classified,
        # 仅归因到第一笔的 job 级测试
        "amount_test": {
            "status": "PASS",
            "accuracy_report": {"amount_test": {"sales_order_no": "SO25-0281"}},
        },
        "contract_terms": {"status": "PASS"},
        "three_way": {
            "overall_status": "PASS",
            "match_request": {"order": {"order_no": "SO25-0281"}},
            "match_result": {"overall_status": "PASS"},
            "cutoff_result": {"测试状态": "PASS"},
        },
    }
    rows = build_gospd_sample_rows(job)
    assert len(rows) == 2
    by_order = {r["order_no"]: r for r in rows}
    assert by_order["SO25-0281"]["step21"].startswith("Yes")
    assert "尚未单独跑测" in (by_order["SO25-0099"]["exception"] or "")

    out = fill_gospd01010_workbook(job, tmp_path / "multi.xlsx")
    wb = load_workbook(out)
    ws = wb.active
    assert ws["B22"].value == 1
    assert ws["B23"].value == 2
    orders = {ws["I22"].value, ws["I23"].value}
    assert orders == {"SO25-0281", "SO25-0099"}


def test_truncated_so_merges_into_longer():
    from src.reporting.gospd01010_filler import group_classified_by_chain

    classified = [
        {"file_name": "ghost.pdf", "doc_type": "receipt", "fields": {"orderNo": "SO25-002"}},
        {"file_name": "inv.pdf", "doc_type": "invoice", "fields": {"orderNo": "SO25-0021"}},
    ]
    chains = group_classified_by_chain(classified)
    ids = {c[0] for c in chains}
    assert ids == {"SO25-0021"}
    assert len(chains[0][1]) == 2
