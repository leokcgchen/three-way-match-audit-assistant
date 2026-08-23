"""受控补缺 × 底稿导向：最小闭环验收（无真实 LLM）。

闭环：
  触发入库(PROPOSED) → Gate5/导出被拦
  → 人工 VERIFIED/REJECTED → 定向复跑(可 mock)
  → Gate5 通过 → 官方/通用底稿生成且旁注口径正确
"""

from __future__ import annotations

from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from src.api.main import app
from src.audit.gap_fill_orchestrator import ingest_verified_claims
from src.audit.gap_fill_replay import apply_advisory_decision
from src.audit.gospd_assertions import build_gospd_assertions
from src.audit.workpaper_notes import pending_advisory_for_job
from src.models.advisory_candidates import new_advisory_candidate
from src.reporting.audit_workbook_xlsx import (
    build_audit_workbook_payload,
    generate_audit_workbook_xlsx,
)
from src.reporting.gospd01010_filler import fill_gospd01010_workbook, resolve_template_path
from src.workflow.job_store import JOB_STORE
from src.workflow.pipeline import build_workbooks_for_job


SOURCE = "销售订单号：SO25-0281；数量：10；签收确认。"


def _mock_docs() -> list[dict]:
    return [
        {
            "file_name": "c.pdf",
            "doc_type": "contract",
            "raw_text": "销售合同 双方盖章 商业实质 购销 签收后付款 " + SOURCE,
            "fields": {
                "contractNo": "HT25-0281",
                "paymentTerms": "签收后30日付款",
                "controlTransferTerms": "签收后转移控制权",
                "transportTerms": "签收确认",
                "buyerName": "测试客户",
            },
        },
        {
            "file_name": "o.pdf",
            "doc_type": "order",
            "raw_text": SOURCE,
            "fields": {
                "orderNo": "SO25-0281",
                "quantity": 10,
                "totalAmount": 1130,
            },
        },
        {
            "file_name": "r.pdf",
            "doc_type": "receipt",
            "raw_text": SOURCE,
            "fields": {
                "documentNo": "R-001",
                "acceptanceDate": "2025-01-08",
                "quantity": 10,
            },
        },
        {
            "file_name": "i.pdf",
            "doc_type": "invoice",
            "raw_text": SOURCE,
            "fields": {
                "invoiceNo": "INV-1",
                "orderNo": "SO25-0281",
                "totalAmount": 1130,
                "quantity": 10,
                "buyerName": "测试客户",
            },
            "ledger_voucher": "记-001",
            "ledger_amount": 1130,
        },
        {
            "file_name": "p.pdf",
            "doc_type": "payment",
            "fields": {"totalAmount": 1130},
        },
    ]


def _seed_ready_job(*, with_pending: bool = True) -> dict:
    job = JOB_STORE.create(title="advisory-loop")
    jid = job["job_id"]
    JOB_STORE.set_goals(jid, ["gospd01010"])
    cand = None
    if with_pending:
        cand = new_advisory_candidate(
            task_type="AMOUNT_GAP_FILL",
            kind="fact",
            business_id="SO25-0281",
            payload={
                "field_name": "quantity",
                "normalized_candidate": 10,
                "file_name": "r.pdf",
                "excerpt": "销售订单号：SO25-0281",
                "confidence": 0.9,
            },
            evidence={
                "excerpt": "销售订单号：SO25-0281",
                "source_doc": "r.pdf",
            },
            verify={"passed": True, "reason": "ok"},
            invalidates=["amount", "gate5"],
            fingerprint="loop-qty",
            status="PROPOSED",
        )
    patch = {
        "classified": _mock_docs(),
        "fields_confirmed": True,
        "fields_confirm_sig": "sig",
        "matching_confirmed": True,
        "matching_confirm_sig": "msig",
        "evidence": {"status": "PASS"},
        "relations": [],
        "duplicates": {"ran": True, "findings": []},
        "contract_terms": {
            "status": "PASS",
            "extracted": {"issue_sources": {"rule": [], "llm": []}},
        },
        "amount_test": {
            "status": "PASS",
            "accuracy_report": {
                "amount_test": {"test_status": "PASS"},
                "source_values": {
                    "quantity_source": "receipt",
                    "price_source": "order",
                },
            },
        },
        "three_way": {
            "overall_status": "PASS",
            "match_result": {"overall_status": "PASS"},
            "cutoff_result": {"测试状态": "PASS"},
        },
        "advisory_candidates": [cand] if cand else [],
    }
    sample = {
        "contract_terms": patch["contract_terms"],
        "amount_test": patch["amount_test"],
        "three_way": patch["three_way"],
        "evidence": patch["evidence"],
        "matching_confirmed": True,
        "matching_confirm_sig": "msig",
    }
    patch["gospd_sample_results"] = {"SO25-0281": sample}
    patch["active_chain_id"] = "SO25-0281"
    return JOB_STORE.update(jid, **patch)


def test_trigger_ingest_then_gate_blocks():
    """触发入库 → Gate5 被拦。"""
    out = ingest_verified_claims(
        [],
        task_type="MATCHING_DISAMBIGUATION",
        claims=[
            {
                "file_name": "o.pdf",
                "disposition": "ADOPT",
                "excerpt": "销售订单号：SO25-0281",
                "confidence": 0.92,
                "business_id": "SO25-0281",
            }
        ],
        full_text=SOURCE,
        trigger_reasons=["MATCHING_AMBIGUITY"],
        business_id="SO25-0281",
        require_excerpt=True,
    )
    assert out["proposed"]
    assert out["proposed"][0]["status"] == "PROPOSED"

    job = _seed_ready_job(with_pending=False)
    jid = job["job_id"]
    JOB_STORE.update(jid, advisory_candidates=out["store"])
    assert pending_advisory_for_job(JOB_STORE.get(jid))
    with pytest.raises(ValueError, match="顾问候选"):
        JOB_STORE.confirm_conclusion(jid)


def test_decide_reject_then_confirm_and_fill_gospd(tmp_path: Path, monkeypatch):
    """拒绝候选 → Gate5 → 官方底稿可生成；异常说明无「AI认为」。"""
    if not resolve_template_path().is_file():
        pytest.skip("无 GOSPD01010 模板")

    job = _seed_ready_job(with_pending=True)
    jid = job["job_id"]
    cid = job["advisory_candidates"][0]["candidate_id"]

    monkeypatch.setattr(
        "src.audit.gap_fill_replay.run_amount",
        lambda docs: {"status": "PASS", "replayed": True},
    )
    decided = apply_advisory_decision(
        jid, cid, "REJECTED", reason="不适用", auto_replay=True
    )
    assert decided["after"]["status"] == "REJECTED"
    assert pending_advisory_for_job(JOB_STORE.get(jid)) == []

    confirmed = JOB_STORE.confirm_conclusion(jid)
    assert confirmed["conclusion_confirmed"] is True

    out = fill_gospd01010_workbook(
        JOB_STORE.get(jid),
        tmp_path / "gospd_loop.xlsx",
        entity_name="闭环验收",
    )
    assert out.is_file()
    wb = load_workbook(out)
    ws = wb.active
    assert str(ws["R22"].value or "").startswith("Yes") or ws["R22"].value
    exc = str(ws["V22"].value or "")
    assert "AI认为有风险" not in exc
    assert "审计结论为" not in exc


def test_decide_verify_amount_replays_then_official_workbook(
    tmp_path: Path, monkeypatch
):
    """接受金额候选 → 尝试复跑 amount → 官方 GOSPD01010 可导出。"""
    job = _seed_ready_job(with_pending=True)
    jid = job["job_id"]
    JOB_STORE.set_goals(jid, ["gospd01010"])
    store = [
        new_advisory_candidate(
            task_type="AMOUNT_GAP_FILL",
            business_id="SO25-0281",
            payload={
                "field_name": "quantity",
                "normalized_candidate": 10,
                "file_name": "r.pdf",
                "excerpt": "销售订单号：SO25-0281",
                "confidence": 0.9,
            },
            evidence={"excerpt": "销售订单号：SO25-0281", "source_doc": "r.pdf"},
            verify={"passed": True, "reason": "ok"},
            invalidates=["amount", "gate5"],
            fingerprint="loop-qty-2",
        )
    ]
    JOB_STORE.update(
        jid,
        classified=_mock_docs(),
        fields_confirmed=True,
        matching_confirmed=True,
        evidence={"status": "PASS"},
        contract_terms={"status": "PASS"},
        amount_test={"status": "PASS"},
        three_way={
            "overall_status": "PASS",
            "match_result": {"overall_status": "PASS"},
            "cutoff_result": {"测试状态": "PASS"},
        },
        advisory_candidates=store,
        conclusion_confirmed=False,
    )

    cid = store[0]["candidate_id"]

    def _fake_amount(docs):
        return {"status": "WARNING", "replayed": True}

    monkeypatch.setattr("src.audit.gap_fill_replay.run_amount", _fake_amount)
    out = apply_advisory_decision(jid, cid, "VERIFIED", auto_replay=True)
    assert out["after"]["status"] == "VERIFIED"
    assert ("amount" in (out.get("replayed") or [])) or any(
        "amount" in s for s in (out.get("skipped") or [])
    )

    fresh = JOB_STORE.get(jid)
    JOB_STORE.update(
        jid,
        fields_confirmed=True,
        matching_confirmed=True,
        evidence=fresh.get("evidence") or {"status": "PASS"},
        contract_terms=fresh.get("contract_terms") or {"status": "PASS"},
        amount_test=fresh.get("amount_test") or {"status": "PASS"},
        three_way=fresh.get("three_way")
        or {
            "overall_status": "PASS",
            "match_result": {"overall_status": "PASS"},
            "cutoff_result": {"测试状态": "PASS"},
        },
    )
    confirmed = JOB_STORE.confirm_conclusion(jid)
    assert confirmed["conclusion_confirmed"] is True

    paths = build_workbooks_for_job(JOB_STORE.get(jid))
    assert paths
    assert "GOSPD01010" in paths[0].name.upper()
    wb = load_workbook(paths[0])
    assert wb.active is not None


def test_http_export_not_blocked_by_non_field_advisory():
    """HTTP：非字段类顾问候选不挡导出（口径 A）。"""
    client = TestClient(app)
    job = _seed_ready_job(with_pending=True)
    jid = job["job_id"]

    JOB_STORE.update(jid, conclusion_confirmed=True, conclusion_confirm_sig="x")
    r = client.post(f"/api/v1/workflow/jobs/{jid}/workbook/export")
    detail = str((r.json() or {}).get("detail") or "")
    assert not (r.status_code == 400 and ("顾问" in detail or "AI 候选" in detail))



def test_assertions_notes_scoped_to_chain():
    """旁注按业务链归属，不串到另一笔。"""
    docs_a = _mock_docs()
    job = {
        "goal_ids": ["gospd01010"],
        "classified": docs_a,
        "contract_terms": {"status": "PASS"},
        "amount_test": {"status": "PASS"},
        "three_way": {
            "overall_status": "PASS",
            "match_result": {"overall_status": "PASS"},
            "cutoff_result": {"测试状态": "PASS"},
        },
        "advisory_candidates": [
            new_advisory_candidate(
                task_type="CONTRACT_CLARITY_REVIEW",
                business_id="SO25-0099",
                payload={"issue_code": "REBATE_TERM_AMBIGUOUS"},
                evidence={"excerpt": "x"},
                fingerprint="other-chain",
                status="PROPOSED",
            )
        ],
    }
    out = build_gospd_assertions(
        docs=docs_a, job=job, chain_id="SO25-0281", apply_job_tests=True
    )
    assert "SO25-0099" not in (out.get("system_observation") or "")
    assert "REBATE_TERM_AMBIGUOUS" not in (out.get("system_observation") or "")


def test_generic_payload_includes_advisory_sheet(tmp_path: Path):
    cand = new_advisory_candidate(
        task_type="FIELD_GAP_FILL",
        business_id="SO25-0281",
        payload={"field_name": "paymentTerms", "normalized_candidate": "签收后30日"},
        evidence={"excerpt": "销售订单号：SO25-0281"},
        status="VERIFIED",
        fingerprint="sheet-1",
    )
    payload = build_audit_workbook_payload(
        amount={"status": "PASS", "human_readable_summary": "ok"},
        advisory_candidates=[cand],
        conclusion_confirmed=True,
    )
    assert payload["advisory"]["ran"] is True
    path = generate_audit_workbook_xlsx(payload, tmp_path / "adv.xlsx")
    wb = load_workbook(path)
    assert "顾问候选与旁注" in wb.sheetnames
