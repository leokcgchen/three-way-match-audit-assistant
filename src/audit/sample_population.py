"""裁剪序时账两用：工作台立样本笔 + 后续测试用入账日/金额。

本系统不做抽样设计。上传的是已按抽样裁过的序时账：
- 工作台只看业务号是否对上、凭证齐不齐
- 同一份里的入账日/金额留给测试，不挡字段确认
- 上传页不再另传序时账
"""

from __future__ import annotations

import re
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Optional

_ORDER_COL_BEST = ("销售订单号", "订单编号", "销售订单", "订单号", "sono", "salesorder", "order_no")
_PRIMARY_KEY_STRONG = (
    "业务编号",
    "业务id",
    "业务索引",
    "业务流水",
    "样本编号",
    "样本索引",
    "审计索引",
    "业务主键",
)
_PRIMARY_KEY_SECONDARY = (
    "销售订单",
    "采购订单",
    "订单编号",
    "订单号",
    "交易号",
    "交易编号",
    "参考号",
    "reference",
    "case ref",
    "business key",
    "record id",
)
_PRIMARY_KEY_BLOCKED = (
    "凭证",
    "发票",
    "行号",
    "序号",
    "日期",
    "时间",
    "金额",
    "客户",
    "供应商",
    "名称",
    "数量",
    "单位",
    "币种",
    "date",
    "amount",
    "customer",
    "vendor",
    "name",
    "quantity",
    "currency",
    "voucher",
    "invoice",
)
_DATE_COL = ("过账日期", "入账日期", "记账日期", "凭证日期", "账载日期", "过账日", "posting")
_AMT_COL_CREDIT = ("贷方金额", "贷方发生额", "贷方")
_AMT_COL_DEBIT = ("借方金额", "借方发生额", "借方")
_AMT_COL_ANY = ("价税合计", "含税金额", "入账金额", "账载金额", "收入金额", "金额")
_CUST_COL = ("客户名称", "往来单位名称", "客户")
_INVOICE_COL = ("发票号码", "发票编号", "发票号", "invoice no", "invoice number", "invoice_no")
_VOUCHER_COL = ("凭证号", "凭证编号", "记账凭证号", "voucher no", "voucher number")
_MATERIAL_COL = ("物料名称", "商品名称", "产品名称", "规格型号", "物料", "material", "item description")
_QUANTITY_COL = ("数量", "开票数量", "验收数量", "签收数量", "qty", "quantity")
_UNIT_COL = ("单位", "计量单位", "unit")
_CURRENCY_COL = ("币种", "货币", "currency")


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _norm_header(text: Any) -> str:
    return re.sub(r"\s+", "", str(text or "")).casefold()


def _header_hits(header: str, hints: Iterable[str]) -> bool:
    n = _norm_header(header)
    if not n:
        return False
    return any(_norm_header(h) in n or n in _norm_header(h) for h in hints if h)


def normalize_biz_ids(raw: Iterable[Any]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for x in raw or []:
        if not _is_usable_primary_key_value(x):
            continue
        nid = _normalize_primary_key_value(x)
        if not nid or nid in seen:
            continue
        seen.add(nid)
        out.append(nid)
    return out


def _pick_col(headers: list[str], hints: Iterable[str], *, skip: set[str] | None = None) -> str:
    skip_n = {_norm_header(x) for x in (skip or set())}
    for hint in hints:
        for h in headers:
            if _norm_header(h) in skip_n:
                continue
            if _header_hits(h, (hint,)):
                return h
    return ""


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        try:
            return value.strftime("%Y-%m-%d")
        except Exception:
            return str(value).strip()
    text = str(value).strip()
    if text.lower() in {"none", "nan", "-"}:
        return ""
    return text


def _normalize_primary_key_value(value: Any) -> str:
    from src.legacy_ocr.ledger_parser import normalize_biz_id

    if isinstance(value, bool):
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return normalize_biz_id(value)


def _is_usable_primary_key_value(value: Any) -> bool:
    """抽样清单主键只校验“可作为索引”，不套用 OCR 的特定编号格式。"""
    if value is None or isinstance(value, (bool, date, datetime)):
        return False
    text = _normalize_primary_key_value(value)
    if not text or len(text) > 80 or "\n" in text or "\r" in text:
        return False
    if text.casefold() in {"none", "nan", "null", "-"}:
        return False
    return any(ch.isalnum() for ch in text)


def _primary_key_header_tier(header: str) -> int:
    if _header_hits(header, _PRIMARY_KEY_STRONG):
        return 3
    if _header_hits(header, _PRIMARY_KEY_SECONDARY):
        return 2
    if _header_hits(header, _PRIMARY_KEY_BLOCKED):
        return -1
    return 0


def _rank_primary_key_candidates(
    headers: list[str], data_rows: list[list[Any]]
) -> list[dict[str, Any]]:
    active_rows = [row for row in data_rows if any(_cell_str(value) for value in row)]
    if not active_rows:
        return []
    ranked: list[dict[str, Any]] = []
    for index, header in enumerate(headers):
        header = str(header or "").strip()
        if not header:
            continue
        tier = _primary_key_header_tier(header)
        raw_values = [row[index] for row in active_rows if index < len(row) and _cell_str(row[index])]
        usable = [value for value in raw_values if _is_usable_primary_key_value(value)]
        normalized = [_normalize_primary_key_value(value) for value in usable]
        fill_ratio = len(usable) / len(active_rows)
        unique_ratio = len(set(normalized)) / len(normalized) if normalized else 0.0
        pure_numeric = bool(normalized) and all(re.fullmatch(r"\d+(?:\.0+)?", value) for value in normalized)
        average_length = sum(len(value) for value in normalized) / len(normalized) if normalized else 0.0

        method = ""
        accepted = False
        if (
            tier == 3
            and usable
            and fill_ratio >= 0.5
            and unique_ratio > 0
        ) or (
            tier == 2
            and usable
            and fill_ratio >= 0.5
            and unique_ratio >= 0.9
        ):
            method = "keyword"
            accepted = True
        elif (
            tier == 0
            and len(usable) >= 2
            and fill_ratio >= 0.8
            and unique_ratio >= 0.98
            and not pure_numeric
            and average_length <= 40
        ):
            method = "uniqueness_fallback"
            accepted = True

        score = tier * 100 + unique_ratio * 20 + fill_ratio * 10
        if method == "uniqueness_fallback":
            score += 40
        confidence = 0.0
        if method == "keyword":
            confidence = min(0.99, (0.94 if tier == 3 else 0.88) + fill_ratio * 0.03 + unique_ratio * 0.02)
        elif method == "uniqueness_fallback":
            confidence = min(0.85, 0.7 + fill_ratio * 0.07 + unique_ratio * 0.08)
        ranked.append(
            {
                "column": header,
                "index": index,
                "method": method,
                "confidence": round(confidence, 3),
                "score": round(score, 3),
                "fill_ratio": round(fill_ratio, 3),
                "unique_ratio": round(unique_ratio, 3),
                "accepted": accepted,
                "reason": (
                    "关键词与唯一性校验通过"
                    if method == "keyword"
                    else "高非空率、高唯一率标识符列"
                    if method == "uniqueness_fallback"
                    else "受限字段"
                    if tier < 0
                    else "唯一性或非空率不足"
                ),
            }
        )
    return sorted(ranked, key=lambda item: (bool(item["accepted"]), item["score"]), reverse=True)


def _select_primary_key(headers: list[str], data_rows: list[list[Any]]) -> dict[str, Any] | None:
    candidates = _rank_primary_key_candidates(headers, data_rows)
    return candidates[0] if candidates and candidates[0]["accepted"] else None


def _looks_like_header_row(cells: list[str]) -> bool:
    hints = (
        *_PRIMARY_KEY_STRONG,
        *_PRIMARY_KEY_SECONDARY,
        *_PRIMARY_KEY_BLOCKED,
        *_DATE_COL,
        *_AMT_COL_CREDIT,
        *_AMT_COL_DEBIT,
        *_AMT_COL_ANY,
        *_CUST_COL,
        *_INVOICE_COL,
        *_VOUCHER_COL,
        *_MATERIAL_COL,
        *_QUANTITY_COL,
        *_UNIT_COL,
        *_CURRENCY_COL,
    )
    return any(_header_hits(cell, hints) for cell in cells if cell)


def _parse_sheet_rows(
    headers: list[str],
    data_rows: list[list[Any]],
    *,
    sheet: str,
    primary_key_col: str,
) -> list[dict[str, Any]]:
    from src.legacy_ocr.ledger_parser import normalize_ledger_date

    if not primary_key_col:
        return []
    order_col = _pick_col(headers, _ORDER_COL_BEST, skip={primary_key_col})
    date_col = _pick_col(headers, _DATE_COL)
    amt_col = _pick_col(headers, _AMT_COL_CREDIT) or _pick_col(headers, _AMT_COL_DEBIT) or _pick_col(
        headers, _AMT_COL_ANY, skip={primary_key_col}
    )
    cust_col = _pick_col(headers, _CUST_COL)
    invoice_col = _pick_col(headers, _INVOICE_COL, skip={primary_key_col})
    voucher_col = _pick_col(headers, _VOUCHER_COL, skip={primary_key_col})
    material_col = _pick_col(headers, _MATERIAL_COL)
    quantity_col = _pick_col(headers, _QUANTITY_COL)
    unit_col = _pick_col(headers, _UNIT_COL)
    currency_col = _pick_col(headers, _CURRENCY_COL)
    idx = {h: i for i, h in enumerate(headers)}

    def take(row: list[Any], col: str) -> Any:
        if not col or col not in idx:
            return None
        i = idx[col]
        return row[i] if i < len(row) else None

    out: list[dict[str, Any]] = []
    for row in data_rows:
        raw_id = take(row, primary_key_col)
        if not _is_usable_primary_key_value(raw_id):
            continue
        nid = _normalize_primary_key_value(raw_id)
        order_number = ""
        if order_col:
            raw_order = take(row, order_col)
            if _is_usable_primary_key_value(raw_order):
                order_number = _normalize_primary_key_value(raw_order)
        book_date = ""
        if date_col:
            book_date = normalize_ledger_date(take(row, date_col)) or _cell_str(take(row, date_col))
        amt_raw = take(row, amt_col) if amt_col else None
        book_amount = None
        if amt_raw is not None and str(amt_raw).strip() not in {"", "None", "nan"}:
            try:
                book_amount = float(str(amt_raw).replace(",", "").replace("，", ""))
            except ValueError:
                book_amount = None
        quantity = None
        quantity_raw = take(row, quantity_col) if quantity_col else None
        if quantity_raw is not None and str(quantity_raw).strip() not in {"", "None", "nan"}:
            try:
                quantity = float(str(quantity_raw).replace(",", "").replace("，", ""))
            except ValueError:
                quantity = None
        invoice_number = _cell_str(take(row, invoice_col)) if invoice_col else ""
        voucher_number = _cell_str(take(row, voucher_col)) if voucher_col else ""
        material_name = _cell_str(take(row, material_col)) if material_col else ""
        unit = _cell_str(take(row, unit_col)) if unit_col else ""
        currency = _cell_str(take(row, currency_col)).upper() if currency_col else ""
        out.append(
            {
                "business_id": nid,
                "order_numbers": [order_number] if order_number else [],
                "invoice_numbers": [invoice_number] if invoice_number else [],
                "voucher_numbers": [voucher_number] if voucher_number else [],
                "material_names": [material_name] if material_name else [],
                "quantities": [quantity] if quantity is not None else [],
                "units": [unit] if unit else [],
                "currencies": [currency] if currency else [],
                "book_date": book_date,
                "book_amount": book_amount,
                "customer": _cell_str(take(row, cust_col)) if cust_col else "",
                "sheet": sheet,
            }
        )
    return out


def parse_sample_workbook(path: str | Path) -> dict[str, Any]:
    """从序时账/抽样表抽出业务主键；业务编号优先，未知 ERP 表头按唯一性安全推断。"""
    from openpyxl import load_workbook

    src = Path(path)
    if not src.is_file():
        raise ValueError("抽样文件不存在")
    wb = load_workbook(src, data_only=True, read_only=True)
    all_rows: list[dict[str, Any]] = []
    sheets_used: list[str] = []
    selections: list[dict[str, Any]] = []
    scanned_headers: list[list[str]] = []
    rejected_candidates: list[dict[str, Any]] = []
    try:
        for ws in wb.worksheets:
            raw: list[list[Any]] = []
            for row in ws.iter_rows(values_only=True):
                raw.append(list(row))
            header_idx = -1
            headers: list[str] = []
            selection: dict[str, Any] | None = None
            best_header_score = float("-inf")
            for i, row in enumerate(raw[:12]):
                cells = [str(c).strip() if c is not None else "" for c in row]
                if sum(1 for value in row if isinstance(value, str) and value.strip()) < 2:
                    continue
                if not _looks_like_header_row(cells):
                    continue
                scanned_headers.append([cell for cell in cells if cell])
                candidate = _select_primary_key(cells, raw[i + 1 :])
                ranked = _rank_primary_key_candidates(cells, raw[i + 1 :])
                rejected_candidates.extend(
                    {**item, "sheet": str(ws.title), "header_row": i + 1} for item in ranked[:5]
                )
                if candidate:
                    header_score = float(candidate["score"]) + min(sum(bool(cell) for cell in cells), 8) * 2
                    if header_score <= best_header_score:
                        continue
                    header_idx = i
                    headers = cells
                    selection = candidate
                    best_header_score = header_score
            if header_idx < 0 or not headers or not selection:
                continue
            parsed = _parse_sheet_rows(
                headers,
                raw[header_idx + 1 :],
                sheet=str(ws.title),
                primary_key_col=str(selection["column"]),
            )
            if parsed:
                sheets_used.append(str(ws.title))
                all_rows.extend(parsed)
                selections.append({**selection, "sheet": str(ws.title), "header_row": header_idx + 1})
    finally:
        wb.close()
    if not all_rows:
        header_preview = "；".join("、".join(row[:8]) for row in scanned_headers[:3]) or "（未识别到表头）"
        accepted = next((item for item in rejected_candidates if item.get("accepted")), None)
        if accepted:
            raise ValueError(
                f"已识别候选主键列「{accepted['column']}」，但没有可用业务索引值。扫描表头：{header_preview}"
            )
        raise ValueError(f"未能自动确定唯一业务索引列。已扫描表头：{header_preview}")

    from src.legacy_ocr.ledger_parser import compact_biz_id

    merged: list[dict[str, Any]] = []
    by_business: dict[str, dict[str, Any]] = {}
    for row in all_rows:
        key = compact_biz_id(row["business_id"])
        existing = by_business.get(key)
        if existing is None:
            existing = dict(row)
            for field in (
                "order_numbers",
                "invoice_numbers",
                "voucher_numbers",
                "material_names",
                "quantities",
                "units",
                "currencies",
            ):
                existing[field] = list(row.get(field) or [])
            by_business[key] = existing
            merged.append(existing)
            continue
        for field in (
            "order_numbers",
            "invoice_numbers",
            "voucher_numbers",
            "material_names",
            "quantities",
            "units",
            "currencies",
        ):
            for value in row.get(field) or []:
                if value not in existing[field]:
                    existing[field].append(value)

    order_owners: dict[str, dict[str, Any]] = {}
    for row in merged:
        for order_number in row.get("order_numbers") or []:
            key = compact_biz_id(order_number)
            owner = order_owners.setdefault(
                key,
                {"value": order_number, "business_ids": []},
            )
            business_id = str(row.get("business_id") or "")
            if business_id and business_id not in owner["business_ids"]:
                owner["business_ids"].append(business_id)
    ambiguous_aliases = [
        {
            "type": "order_number",
            "value": owner["value"],
            "business_ids": owner["business_ids"],
        }
        for owner in order_owners.values()
        if len(owner["business_ids"]) > 1
    ]
    ledger_rows = [
        {
            "business_id": r["business_id"],
            "book_date": r.get("book_date") or "",
            "book_amount": r.get("book_amount"),
            "customer": r.get("customer") or "",
            "sheet": r.get("sheet") or "",
        }
        for r in all_rows
    ]
    primary = selections[0]
    return {
        "rows": merged,
        "business_ids": [r["business_id"] for r in merged],
        "ambiguous_aliases": ambiguous_aliases,
        "sheets": sheets_used,
        "ledger_rows": ledger_rows,
        "ledger_columns": ["business_id", "book_date", "book_amount", "customer", "sheet"],
        "ledger_mapping": {
            "biz_id": "business_id",
            "posting_date": "book_date",
            "amount": "book_amount",
        },
        "ledger_auto_ok": any(bool(r.get("book_date")) for r in ledger_rows),
        "ledger_standard_map": {
            "业务编号": "business_id",
            "入账日期": "book_date",
            "金额": "book_amount",
        },
        "primary_key_column": primary["column"],
        "primary_key_method": primary["method"],
        "primary_key_confidence": primary["confidence"],
        "primary_key_candidates": primary.get("candidates")
        or [
            {
                key: item.get(key)
                for key in ("column", "method", "confidence", "fill_ratio", "unique_ratio", "accepted", "reason")
            }
            for item in rejected_candidates[:10]
        ],
    }


def ledger_patch_from_parsed(parsed: dict[str, Any], *, path: str = "") -> dict[str, Any]:
    """把裁剪序时账解析结果写成现有测试用的 ledger_* 字段。"""
    return {
        "ledger_path": path or None,
        "ledger_rows": list(parsed.get("ledger_rows") or []),
        "ledger_columns": list(parsed.get("ledger_columns") or []),
        "ledger_mapping": dict(parsed.get("ledger_mapping") or {}),
        "ledger_auto_ok": bool(parsed.get("ledger_auto_ok")),
        "ledger_standard_map": dict(parsed.get("ledger_standard_map") or {}),
    }


def build_sample_population(
    *,
    business_ids: Iterable[Any],
    source: str = "external_import",
    note: str = "",
    rows: list[dict[str, Any]] | None = None,
    sheets: list[str] | None = None,
    primary_key_column: str = "",
    primary_key_method: str = "",
    primary_key_confidence: float = 0.0,
    primary_key_candidates: list[dict[str, Any]] | None = None,
    ambiguous_aliases: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    ids = normalize_biz_ids(business_ids)
    by_id = {str(r.get("business_id") or ""): r for r in (rows or []) if r.get("business_id")}
    kept_rows = [by_id[i] for i in ids if i in by_id]
    return {
        "business_ids": ids,
        "count": len(ids),
        "source": str(source or "external_import"),
        "note": str(note or ""),
        "imported_at": _utc_now(),
        "cannot_claim": "不替代抽样设计；不证明总体完整性",
        "rows": kept_rows,
        "sheets": list(sheets or []),
        "primary_key_column": str(primary_key_column or ""),
        "primary_key_method": str(primary_key_method or ""),
        "primary_key_confidence": float(primary_key_confidence or 0.0),
        "primary_key_candidates": list(primary_key_candidates or []),
        "ambiguous_aliases": list(ambiguous_aliases or []),
    }


def chain_in_population(chain_id: str, population: Optional[dict[str, Any]]) -> Optional[bool]:
    """None=未导入总体；True/False=是否在清单内。"""
    if not isinstance(population, dict) or not population.get("business_ids"):
        return None
    from src.legacy_ocr.ledger_parser import compact_biz_id, normalize_biz_id

    nid = normalize_biz_id(chain_id)
    pool = {normalize_biz_id(x) for x in (population.get("business_ids") or [])}
    if nid in pool:
        return True
    c = compact_biz_id(nid)
    return any(compact_biz_id(x) == c for x in pool if x)


def desk_sample_ids(job: dict[str, Any]) -> list[str]:
    """工作台样本笔：一旦导入清单，只能由清单立笔。

    OCR 识别结果是凭证证据，不是抽样总体，绝不能据此扩充业务。
    未导入清单的旧任务仍保留按识别结果立笔的兼容行为。
    """
    from src.legacy_ocr.ledger_parser import compact_biz_id
    from src.workflow.chain_workspace import list_business_chains

    pop = job.get("sample_population") if isinstance(job.get("sample_population"), dict) else {}
    ids = normalize_biz_ids(pop.get("business_ids") or [])
    if isinstance(job.get("sample_population"), dict):
        return ids
    seen = {compact_biz_id(x) for x in ids}
    compact_ids = list(seen)
    for row in list_business_chains(list(job.get("classified") or [])):
        cid = str(row.get("chain_id") or "").strip()
        if not cid or cid == "未识别业务号":
            continue
        key = compact_biz_id(cid)
        if key in seen:
            continue
        if any(p.startswith(key) and len(p) > len(key) for p in compact_ids):
            continue
        seen.add(key)
        ids.append(cid)
    return ids
