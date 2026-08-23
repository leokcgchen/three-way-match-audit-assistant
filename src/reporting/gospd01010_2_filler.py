"""按官方 GOSPD01010.2 模板填表：履约义务区分抽凭。

列映射（Sheet1，表头行 17，数据起始行 20，预置样本 1–20）：
B 样本编号 | C 会计分录编号 | D 合同索引号 | E 客户名称 | F 主要合同条款
G 交易价格 | H 检查其他相关文件确定交易价格(若适用)
I 其他相关文件类型 | J 其他相关文件索引号
K 是否已适当确定合同中可明确区分的履约义务？
（H/K 仅填模板下拉枚举；无独立异常说明列，旁注写入备注行）
"""

from __future__ import annotations

import shutil
from copy import copy
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.audit.gospd01010_2_assertions import build_gospd01010_2_assertions
from src.reporting.gospd01010_filler import group_classified_by_chain

ROOT = Path(__file__).resolve().parents[2]
TEMPLATE_CANDIDATES = (
    ROOT / "templates" / "GOSPD01010.2.xlsx",
    ROOT / "templates" / "GOSPD01010_2.xlsx",
)

DATA_START_ROW = 20
MAX_SAMPLE_ROWS = 20

COL = {
    "sample_no": 2,  # B
    "voucher": 3,  # C
    "contract_idx": 4,  # D
    "customer": 5,  # E
    "contract_terms": 6,  # F
    "txn_price": 7,  # G
    "other_applicable": 8,  # H
    "other_file_type": 9,  # I
    "other_file_index": 10,  # J
    "po_ok": 11,  # K
}


def resolve_template_path() -> Path:
    for p in TEMPLATE_CANDIDATES:
        if p.is_file():
            return p
    raise FileNotFoundError(
        "未找到 GOSPD01010.2 模板，请将文件放到 templates/GOSPD01010.2.xlsx"
    )


def _f(doc: Optional[dict[str, Any]], *keys: str) -> Any:
    if not doc:
        return None
    from src.models.field_values import rule_readable_fields

    fields = rule_readable_fields(doc)
    for k in keys:
        if fields.get(k) not in (None, ""):
            return fields.get(k)
        if doc.get(k) not in (None, ""):
            return doc.get(k)
    return None


def _num(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(
            str(v)
            .replace(",", "")
            .replace("，", "")
            .replace("¥", "")
            .replace("元", "")
            .replace("CNY", "")
            .strip()
        )
    except ValueError:
        return None


def _by_type(docs: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    out: dict[str, dict[str, Any]] = {}
    for d in docs or []:
        t = str(d.get("doc_type") or "")
        if t and t not in out:
            out[t] = d
    return out


def _contract_terms_text(contract: Optional[dict], order: Optional[dict]) -> str:
    bits: list[str] = []
    for key, label in (
        ("paymentTerms", "付款"),
        ("controlTransferTerms", "控制权"),
        ("performanceObligations", "履约义务"),
        ("settlementTerms", "结算"),
    ):
        v = _f(contract, key) or _f(order, key)
        if v:
            bits.append(f"{label}:{v}")
    return "；".join(bits) or str(_f(contract, "remarks") or "")


def _txn_price(
    contract: Optional[dict],
    order: Optional[dict],
    invoice: Optional[dict],
) -> Optional[float]:
    for doc in (contract, order, invoice):
        n = _num(_f(doc, "totalAmount", "amount", "grossAmount"))
        if n is not None and n > 0:
            return n
    return None


def _job_test_biz_keys(job: dict[str, Any]) -> set[str]:
    keys: set[str] = set()
    for src in (
        job.get("contract_terms"),
        job.get("amount_test"),
        job.get("three_way"),
        job.get("evidence"),
    ):
        if isinstance(src, dict):
            for k in ("business_id", "sales_order_no", "order_no", "chain_id"):
                v = str(src.get(k) or "").strip()
                if v:
                    keys.add(v.upper())
    return keys


def _chain_related(chain_id: str, docs: list, test_keys: set[str]) -> bool:
    if not test_keys:
        return True
    blob = " ".join(
        [chain_id]
        + [str(d.get("file_name") or "") for d in docs]
        + [str((d.get("fields") or {}).get("orderNo") or "") for d in docs]
    ).upper()
    return any(k in blob for k in test_keys)


def _row_from_chain(
    *,
    sample_no: int,
    chain_id: str,
    docs: list[dict[str, Any]],
    job: dict[str, Any],
    apply_job_tests: bool,
) -> dict[str, Any]:
    by = _by_type(docs)
    contract = by.get("contract")
    order = by.get("order")
    invoice = by.get("invoice")

    assertions = build_gospd01010_2_assertions(
        docs=docs,
        job=job,
        chain_id=chain_id,
        apply_job_tests=apply_job_tests,
    )
    other = assertions.get("other_files") or {}

    voucher = (
        (invoice or {}).get("ledger_voucher")
        or _f(invoice, "voucherNo", "documentNo")
        or ""
    )
    contract_no = (
        _f(contract, "contractNo", "documentNo")
        or _f(order, "contractNo")
        or (chain_id if "HT" in str(chain_id).upper() else "")
    )
    customer = (
        _f(invoice, "buyerName", "customerName")
        or _f(order, "buyerName", "customerName")
        or _f(contract, "buyerName", "customerName")
        or ""
    )

    # 异常旁注：本模板无 V 列，压缩进条款备注尾部（明确非结论）
    terms = _contract_terms_text(contract, order)
    appendix = str(assertions.get("exception_appendix") or assertions.get("exception") or "").strip()
    if appendix and "系统观察" in appendix:
        terms = f"{terms}\n（旁注见底稿备注行）" if terms else "（旁注见底稿备注行）"

    return {
        "sample_no": sample_no,
        "chain_id": chain_id,
        "voucher": voucher,
        "contract_idx": contract_no,
        "customer": customer,
        "contract_terms": terms,
        "txn_price": _txn_price(contract, order, invoice),
        "other_applicable": assertions.get("other_applicable")
        or other.get("applicable_label")
        or "Not applicable 不适用",
        "other_file_type": assertions.get("other_file_type") or other.get("file_type") or "",
        "other_file_index": assertions.get("other_file_index")
        or other.get("file_index")
        or "",
        "po_ok": assertions.get("po_label") or "",
        "exception": assertions.get("exception") or "",
        "system_observation": assertions.get("system_observation") or "",
        "pending_judgment": assertions.get("pending_judgment") or "",
        "assertions": assertions,
    }


def build_gospd01010_2_sample_rows(job: dict[str, Any]) -> list[dict[str, Any]]:
    classified = list(job.get("classified") or [])
    chains = group_classified_by_chain(classified)
    if not chains:
        return []
    test_keys = _job_test_biz_keys(job)
    samples = (
        job.get("gospd_sample_results")
        if isinstance(job.get("gospd_sample_results"), dict)
        else {}
    )
    single = len(chains) == 1
    rows: list[dict[str, Any]] = []
    for i, (chain_id, docs) in enumerate(chains, start=1):
        sample = samples.get(chain_id) or {}
        has_per = bool(sample.get("contract_terms"))
        apply = has_per or single or _chain_related(chain_id, docs, test_keys)
        rows.append(
            _row_from_chain(
                sample_no=i,
                chain_id=chain_id,
                docs=docs,
                job=job,
                apply_job_tests=apply,
            )
        )
    return rows


def _copy_row_style(ws: Worksheet, src_row: int, dst_row: int) -> None:
    for c in range(1, 12):
        sc = ws.cell(src_row, c)
        dc = ws.cell(dst_row, c)
        if sc.has_style:
            dc._style = copy(sc._style)
        dc.number_format = sc.number_format


def fill_gospd01010_2_workbook(
    job: dict[str, Any],
    output_path: Path,
    *,
    entity_name: str = "",
    currency: str = "人民币",
    unit: str = "Yuan 元",
) -> Path:
    template = resolve_template_path()
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template, output_path)
    try:
        output_path.chmod(0o666)
    except OSError:
        pass

    wb = load_workbook(output_path)
    ws = wb[wb.sheetnames[0]]

    # 表头元信息：B5 标签旁 C5 单位名；E5 程序索引；G5 币种；I5 单位
    ws.cell(5, 3, entity_name or ws.cell(5, 3).value or "")
    ws.cell(5, 5, "GOSPD01010.2")
    ws.cell(5, 7, currency)
    ws.cell(5, 9, unit)

    rows = build_gospd01010_2_sample_rows(job)
    for i, row in enumerate(rows):
        r = DATA_START_ROW + i
        if i >= MAX_SAMPLE_ROWS:
            _copy_row_style(ws, DATA_START_ROW, r)
        elif r > DATA_START_ROW and ws.cell(r, COL["sample_no"]).value in (None, ""):
            _copy_row_style(ws, DATA_START_ROW, r)

        ws.cell(r, COL["sample_no"], row.get("sample_no") or (i + 1))
        ws.cell(r, COL["voucher"], row.get("voucher") or "")
        ws.cell(r, COL["contract_idx"], row.get("contract_idx") or "")
        ws.cell(r, COL["customer"], row.get("customer") or "")
        ws.cell(r, COL["contract_terms"], row.get("contract_terms") or "")
        if row.get("txn_price") is not None:
            ws.cell(r, COL["txn_price"], row["txn_price"])
        # 仅写模板枚举，避免破坏数据验证
        ws.cell(
            r,
            COL["other_applicable"],
            row.get("other_applicable") or "Not applicable 不适用",
        )
        ws.cell(r, COL["other_file_type"], row.get("other_file_type") or "")
        ws.cell(r, COL["other_file_index"], row.get("other_file_index") or "")
        ws.cell(r, COL["po_ok"], row.get("po_ok") or "")

    note_row = max(45, DATA_START_ROW + max(len(rows), MAX_SAMPLE_ROWS) + 3)
    goals = (job.get("plan") or {}).get("goal_ids") or job.get("goal_ids") or []
    notes_bits = [
        "— 本底稿由工作台按目标 gospd01010_2 自动回填；",
        f"已选目标={','.join(map(str, goals)) or 'gospd01010_2'}；",
        f"样本链={','.join(str(r.get('chain_id') or '') for r in rows)}。",
        "H/K 列仅使用模板下拉枚举；履约义务结论来自合同条款测试（履约维度）。",
    ]
    # 汇总旁注（模板无异常列）
    obs = [str(r.get("system_observation") or "") for r in rows if r.get("system_observation")]
    pend = [str(r.get("pending_judgment") or "") for r in rows if r.get("pending_judgment")]
    if obs:
        notes_bits.append("系统观察：" + " | ".join(o.replace("\n", " ")[:120] for o in obs[:3]))
    if pend:
        notes_bits.append("待判断：" + " | ".join(p.replace("\n", " ")[:120] for p in pend[:3]))
    ws.cell(note_row, 2, "".join(notes_bits))

    wb.save(output_path)
    return output_path
