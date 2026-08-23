"""按官方 GOSPD01030 改进模板填表：销售截止（期后）。

对齐：
- templates/GOSPD01030.xlsx
- docs/GOSPD01030_底稿填制指引与Prompt.md

硬约束：
- 表头：F5 程序号 / I5 币种 / K5 单位 / M5 期间截止日
- V 列保留公式，禁止写死 Yes/No
- R 列无可比金额时保持灰度空白，禁止填 0 / 禁止用账面金额回填
- E13 运输条款按下拉写入；E14「是否无需检查系统发票」按是否有发票写 YES/No；E15 是否检查销售订单
- W 列必须写数据验证允许的精确值
"""

from __future__ import annotations

import shutil
from copy import copy
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.audit.gospd01030_assertions import build_gospd01030_assertions
from src.reporting.gospd01010_filler import group_classified_by_chain

ROOT = Path(__file__).resolve().parents[2]
DATA_START_ROW = 30  # 29=示例1；正式样本自 30
SAMPLE_GRID_END_ROW = 48  # 模板预置样本区末行（含）；禁止把系统说明写入此范围内
TEMPLATE_COMMENT_END_ROW = 52  # B50=注释： B51-B52=模板纪律原文
MAX_SAMPLE_ROWS = 15
LOG_SHEET = "02_填制依据与运行日志"

# B样本 C凭证 D过账 E客户 F入账金额 G数量 I发票号 J发票金额
# K/S/T/V 公式列；L订单 M运输 N交货类型 O交货号 P签收日 Q交货数量 R交货金额
# W无异常 X异常说明
COL = {
    "sample_no": 2,
    "voucher": 3,
    "posting_date": 4,
    "customer": 5,
    "amt_book": 6,
    "qty_book": 7,
    "invoice_no": 9,
    "invoice_amt": 10,
    "diff_inv": 11,  # K = F-J
    "order_no": 12,
    "transport": 13,
    "delivery_type": 14,
    "delivery_no": 15,
    "receipt_date": 16,
    "qty_doc": 17,
    "amt_delivery": 18,
    "diff_amt": 19,  # S = F-R
    "diff_qty": 20,  # T = G-Q
    "period_ok": 22,  # V 公式 — 禁止覆盖
    "all_ok": 23,  # W
    "exception": 24,  # X
}

W_YES_FALLBACK = "YES 是"
W_NO_FALLBACK = (
    "No Document the details of exception identified and further testing steps. "
    "否，记录异常的详细信息和进一步测试"
)

GREY_FILL = PatternFill(fill_type="solid", fgColor="D9D9D9")
CONFLICT_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
NO_FILL = PatternFill(fill_type=None)


def resolve_template_path() -> Path:
    candidates = [
        ROOT / "templates" / "GOSPD01030.xlsx",
        Path(r"D:\抽凭—合同合规性审阅agent\templates\GOSPD01030.xlsx"),
    ]
    for p in candidates:
        if p.is_file():
            return p
    raise FileNotFoundError(
        "未找到 GOSPD01030 模板，请将文件放到 templates/GOSPD01030.xlsx"
    )


def _by_type(docs: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    out: dict[str, dict[str, Any]] = {}
    for d in docs or []:
        t = str(d.get("doc_type") or "")
        if t and t not in out:
            out[t] = d
    return out


def _f(doc: Optional[dict[str, Any]], *keys: str) -> Any:
    if not doc:
        return None
    try:
        from src.models.field_values import rule_readable_fields

        fields = rule_readable_fields(doc)
    except Exception:
        fields = dict(doc.get("fields") or {})
    for k in keys:
        if fields.get(k) not in (None, ""):
            return fields.get(k)
        if doc.get(k) not in (None, ""):
            return doc.get(k)
    return None


def _num(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(",", "").replace("，", "").replace("CNY", "").strip())
    except ValueError:
        return None


def _parse_date(raw: Any) -> Optional[date]:
    if raw is None or raw == "":
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    s = str(raw).strip()[:10].replace("/", "-").replace(".", "-")
    try:
        return date.fromisoformat(s)
    except ValueError:
        pass
    # DD/MM/YYYY or DD-MM-YYYY
    parts = str(raw).strip().replace("/", "-").split("-")
    if len(parts) == 3 and len(parts[2]) == 4:
        try:
            d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
            return date(y, m, d)
        except ValueError:
            return None
    return None


def _sample_three_way(job: dict[str, Any], chain_id: str) -> Optional[dict[str, Any]]:
    samples = (
        job.get("gospd_sample_results")
        if isinstance(job.get("gospd_sample_results"), dict)
        else {}
    )
    per = samples.get(chain_id) or {}
    if isinstance(per.get("three_way"), dict):
        return per["three_way"]
    if isinstance(job.get("three_way"), dict):
        return job["three_way"]
    return None


def _delivery_amount_only(
    receipt: Optional[dict[str, Any]],
    delivery: Optional[dict[str, Any]],
) -> Optional[float]:
    """R 列：仅当交货/签收单据确实载有可比金额（>0）时返回；禁止用账面回填。"""
    for doc in (receipt, delivery):
        amt = _num(_f(doc, "totalAmount", "amount", "grossAmount"))
        if amt is not None and abs(amt) > 1e-12:
            return amt
    return None


def _voucher_no(invoice: Optional[dict[str, Any]]) -> str:
    """会计凭证编号：只用序时账/凭证字段，禁止把订单号/单据号误写入 C 列。"""
    import re

    raw = (invoice or {}).get("ledger_voucher") or _f(
        invoice, "voucherNo", "voucherNumber", "accountingVoucherNo", "journalNo"
    )
    if raw in (None, ""):
        return ""
    text = str(raw).strip()
    # 订单号形态不得进凭证列
    if re.match(r"^(SO|KJSO|EXSO)\d", text, flags=re.I):
        return ""
    if re.match(r"^HT\d", text, flags=re.I):
        return ""
    return text


def _resolve_period_end(job: dict[str, Any], explicit: str = "") -> Optional[date]:
    """与断言层同一口径；未配置则返回 None（不静默默认某年年末）。"""
    from src.audit.gospd01030_assertions import resolve_period_end

    if explicit:
        parsed = _parse_date(explicit)
        if parsed:
            return parsed
    return resolve_period_end(job)


def _resolve_entity_name(
    job: dict[str, Any],
    *,
    explicit: str = "",
    classified: Optional[list[dict[str, Any]]] = None,
) -> str:
    if explicit and str(explicit).strip():
        return str(explicit).strip()
    for key in (
        "entity_name",
        "auditee_name",
        "auditee",
        "client_name",
        "company_name",
        "company",
    ):
        val = job.get(key)
        if val not in (None, ""):
            return str(val).strip()
    for item in classified or job.get("classified") or []:
        if str(item.get("doc_type") or "") != "contract":
            continue
        seller = _f(item, "sellerName", "vendorName", "supplierName", "companyName")
        if seller:
            return str(seller).strip()
    title = str(job.get("title") or "").strip()
    if title and title not in {"新任务", "untitled", "Untitled"}:
        return title
    return ""


def _e13_transport_options(ws: Worksheet) -> list[str]:
    opts: list[str] = []
    try:
        for dv in ws.data_validations.dataValidation:
            if "E13" not in str(dv.sqref or ""):
                continue
            raw = str(dv.formula1 or "").strip().strip('"')
            opts = [p.strip() for p in raw.split(",") if p.strip()]
            break
    except Exception:
        pass
    return opts or [
        "客户自提",
        "签收确认",
        "验收确认",
        "外销-FCA货交承运人",
        "外销-FOB离岸价格",
        "外销-CIF成本加保险费、运费",
        "外销-CIP运费、保险费付至指定目的地",
        "外销-DDP完税后交货",
    ]


def _normalize_transport(raw: Any, allowed: list[str]) -> str:
    """将运输条款映射到 E13 下拉允许值。

    禁止：空值默认「签收确认」；禁止仅因字符串含 FOB/CIF 就当实际履约终态。
    调用方应优先传入贸易模式桥给出的 E13 标准项。
    """
    text = str(raw or "").strip()
    if not text:
        return ""
    for opt in allowed:
        if text == opt or opt in text or text in opt:
            return opt
    if text.startswith("外销-") or text in {
        "客户自提",
        "签收确认",
        "验收确认",
    }:
        return text
    if "自提" in text:
        return next((o for o in allowed if "自提" in o), "")
    if "验收" in text:
        return next((o for o in allowed if "验收" in o), "")
    if "签收" in text or "送货" in text:
        return next((o for o in allowed if "签收" in o), "")
    # 裸 Incoterm 弱映射；正式路径应已由贸易模式桥给出标准项
    upper = text.upper()
    if "FOB" in upper:
        return next((o for o in allowed if "FOB" in o), "")
    if "CIF" in upper:
        return next((o for o in allowed if "CIF" in o), "")
    if "CIP" in upper:
        return next((o for o in allowed if "CIP" in o), "")
    if "FCA" in upper:
        return next((o for o in allowed if "FCA" in o), "")
    # DAP/DPU：目的地交货 → 以签收确认为底稿下拉项（与贸易模式桥 DAP→签收确认 一致）
    if "DAP" in upper or "DPU" in upper or "目的地交货" in text:
        return next((o for o in allowed if "签收" in o), "")
    if "DDP" in upper:
        return next((o for o in allowed if "DDP" in o), "")
    return ""


def _resolve_eval_e13(
    rows: list[dict[str, Any]], allowed: list[str]
) -> tuple[str, str]:
    """评量 E13：仅当多样本运输条款一致时填写；否则留空并以行级 M 列为准。"""
    norms: list[str] = []
    for row in rows:
        n = _normalize_transport(row.get("transport"), allowed)
        if n:
            norms.append(n)
    uniq = list(dict.fromkeys(norms))
    if not uniq:
        return "", "多样本均无可用运输条款，E13 留空"
    if len(uniq) == 1:
        return uniq[0], f"标准化为「{uniq[0]}」"
    return (
        "",
        f"多样本运输条款不一致：{' / '.join(uniq)}；E13 留空，以行级 M 列为准",
    )


def _soften_exception_vs_p(
    exception: str,
    *,
    p_filled: bool,
    p_source: str,
) -> str:
    """P 列与贸易模式旁注同源：有控制权日时去掉「缺控制权日」类自相矛盾句。"""
    raw = str(exception or "").strip()
    if not raw:
        return ""
    parts = [p.strip() for p in raw.replace("；", ";").split(";") if p.strip()]
    out: list[str] = []
    for part in parts:
        if p_filled and (
            "缺少控制权候选事件日期证据" in part
            or "控制权日期未获支持，P列暂空" in part
            or "本笔尚缺控制权日期" in part
        ):
            continue
        if (
            p_filled
            and p_source == "on_board"
            and "缺件：" in part
            and "事件日期证据" in part
        ):
            continue
        out.append(part)
    return "；".join(out)


def _formula_v_label(receipt_d: Optional[date], period_end: Optional[date]) -> Optional[str]:
    """与模板 V 公式一致：IF(P>$M$5,\"YES 是\",\"No 否\")。"""
    if receipt_d is None or period_end is None:
        return None
    return "YES 是" if receipt_d > period_end else "No 否"


def _formula_conflict_message(
    period_indep: str,
    formula_v: Optional[str],
    *,
    receipt_d: Optional[date],
    period_end: Optional[date],
) -> str:
    if period_indep and formula_v and period_indep != formula_v:
        return (
            f"FORMULA_LOGIC_CONFLICT: 独立期间判断={period_indep}，"
            f"V公式(P>$M$5)={formula_v}；保留公式，行结论交审计师复核"
        )
    if period_indep and formula_v is None:
        missing: list[str] = []
        if receipt_d is None:
            missing.append("控制权/签收日(P列)")
        if period_end is None:
            missing.append("期间截止日(M5/period_end)")
        detail = "、".join(missing) if missing else "关键日期"
        return (
            f"FORMULA_LOGIC_CONFLICT: 独立期间判断={period_indep}，"
            f"但{detail}缺失，无法与 V 公式对齐"
        )
    return ""


def _w_allowed_labels(ws: Worksheet) -> tuple[str, str]:
    """从模板数据验证读取 W 列允许值（语义：是 / 否）。"""
    yes_v, no_v = W_YES_FALLBACK, W_NO_FALLBACK
    try:
        for dv in ws.data_validations.dataValidation:
            if "W" not in str(dv.sqref or ""):
                continue
            raw = str(dv.formula1 or "").strip().strip('"')
            parts = [p.strip() for p in raw.split(",") if p.strip()]
            if not parts:
                continue
            for p in parts:
                pu = p.upper()
                if "YES" in pu or p.startswith("是"):
                    yes_v = p
                elif "NO" in pu or "否" in p:
                    no_v = p
            break
    except Exception:
        pass
    return yes_v, no_v


def _row_from_chain(
    *,
    sample_no: int,
    chain_id: str,
    docs: list[dict[str, Any]],
    job: dict[str, Any],
    w_yes: str,
    w_no: str,
    period_end: Optional[date],
) -> dict[str, Any]:
    by = _by_type(docs)
    contract = by.get("contract")
    order = by.get("order")
    delivery = by.get("delivery")
    receipt = by.get("receipt") or delivery
    invoice = by.get("invoice")

    qty_book = _num(
        (invoice or {}).get("ledger_quantity")
        or _f(invoice, "quantity")
        or _f(order, "quantity")
    )
    amt_book = _num(
        (invoice or {}).get("ledger_amount")
        or _f(invoice, "totalAmount", "amount")
        or _f(order, "totalAmount")
    )
    invoice_amt = _num(_f(invoice, "totalAmount", "amount"))
    qty_doc = _num(_f(receipt, "quantity") or _f(delivery, "quantity"))
    amt_delivery = _delivery_amount_only(receipt, delivery)

    delivery_type = (
        "客户签收验收单"
        if (receipt or {}).get("doc_type") == "receipt"
        else ("销售发货单" if delivery else "")
    )
    if receipt and "验收" in str(receipt.get("file_name") or ""):
        delivery_type = "产品验收单/签收单"

    three_way = _sample_three_way(job, chain_id)

    # 贸易模式判断：名义≠实际；E13/P 列由此投影，禁止合同 FOB 字符串直接当答案
    from src.workflow.trading_mode_bridge import (
        control_date_for_cutoff,
        e13_from_trading_mode,
        interpret_chain_trading_mode,
        prefers_on_board_cutoff,
    )

    tm = interpret_chain_trading_mode(
        docs,
        transaction_id=str(chain_id or f"sample-{sample_no}"),
        use_llm=None,
        persist=False,
    )
    tm_e13 = e13_from_trading_mode(tm)
    tm_ctrl_date, tm_date_meaning = control_date_for_cutoff(tm)
    prefers_on_board = prefers_on_board_cutoff(tm)

    receipt_raw = (
        _f(receipt, "acceptanceDate", "deliveryDate", "documentDate", "receiptDate")
        or _f(delivery, "deliveryDate", "documentDate")
        or ""
    )
    receipt_d = _parse_date(receipt_raw)
    p_source = "receipt" if receipt_d else "none"
    control_override: Optional[str] = None

    if prefers_on_board:
        # 外销：P 列只用装船/交承运人日；无则留空，禁止仓库签收日冒充
        if tm_ctrl_date:
            receipt_raw = tm_ctrl_date
            receipt_d = _parse_date(tm_ctrl_date)
            p_source = "on_board"
            control_override = tm_ctrl_date
        else:
            receipt_raw = ""
            receipt_d = None
            p_source = "none"
            control_override = ""
    elif tm_ctrl_date and not receipt_d:
        receipt_raw = tm_ctrl_date
        receipt_d = _parse_date(tm_ctrl_date)
        p_source = "on_board" if prefers_on_board else "receipt"
        control_override = tm_ctrl_date

    # 断言与 V 公式共用同一期间截止日 / 控制权日
    job_for_assert = dict(job)
    if period_end and not job_for_assert.get("period_end"):
        job_for_assert["period_end"] = period_end.isoformat()
    assertions = build_gospd01030_assertions(
        docs=docs,
        job=job_for_assert,
        three_way=three_way,
        chain_id=chain_id,
        control_date_override=control_override,
    )

    order_no = (
        _f(order, "orderNo", "documentNo", "salesOrderNo")
        or _f(invoice, "orderNo", "salesOrderNo")
        or (chain_id if str(chain_id).upper().startswith("SO") else "")
    )
    posting = (
        (invoice or {}).get("ledger_posting_date")
        or _f(invoice, "postingDate")
        or ""
    )

    period_verdict = assertions.get("period", {}).get("verdict")
    # 独立期间结论用语对齐 V 公式选项（仅日志/冲突比对，不写入 V）
    if period_verdict is True:
        period_indep = "YES 是"
    elif period_verdict is False:
        period_indep = "No 否"
    else:
        period_indep = ""

    formula_v = _formula_v_label(receipt_d, period_end)
    conflict = _formula_conflict_message(
        period_indep,
        formula_v,
        receipt_d=receipt_d,
        period_end=period_end,
    )

    all_ok = assertions.get("all_ok")
    if conflict:
        # 与公式冲突：不得写「是」；若独立判断已否定或断言有异常，写入 DV「否」便于底稿可读
        if period_indep == "No 否" or all_ok is False:
            all_ok_label = w_no
            all_ok = False
        else:
            all_ok_label = ""
            all_ok = None
    elif all_ok is True:
        all_ok_label = w_yes
    elif all_ok is False:
        all_ok_label = w_no
    else:
        all_ok_label = ""

    exception = str(assertions.get("exception") or "")
    if conflict:
        exception = f"{conflict}；{exception}".rstrip("；") if exception else conflict
    if all_ok is None and not exception:
        gaps = assertions.get("gaps") or []
        if gaps:
            exception = "；".join(str(g) for g in gaps)

    tm_cells = (tm or {}).get("gospd_cells") or {}
    tm_view = (tm or {}).get("workbook_view") or {}
    tm_x = str(tm_cells.get("X_exception") or "").strip()
    if tm_x:
        exception = f"{exception}；{tm_x}".strip("；") if exception else tm_x
    if tm_date_meaning and prefers_on_board and receipt_d:
        note = f"P列含义：{tm_date_meaning}"
        exception = f"{exception}；{note}".strip("；") if exception else note
    elif prefers_on_board and not receipt_d:
        note = "外销控制权日未获装船/交承运人证据，P列留空（禁止用签收日冒充）"
        exception = f"{exception}；{note}".strip("；") if exception else note

    exception = _soften_exception_vs_p(
        exception,
        p_filled=bool(receipt_d),
        p_source=p_source,
    )

    # E13：优先贸易模式桥标准项；禁止空值默认「签收确认」
    transport_raw = (
        tm_e13
        or assertions.get("transport")
        or _f(contract, "transportTerms")
        or _f(order, "transportTerms")
        or ""
    )
    if not str(transport_raw or "").strip():
        weak = "运输条款证据不足，M列留空（禁止默认签收确认）"
        exception = f"{exception}；{weak}".strip("；") if exception else weak

    from src.reporting.gospd01030_exception_nl import build_gospd01030_exception_nl

    exception_nl = build_gospd01030_exception_nl(
        assertions=assertions,
        three_way=three_way,
        qty_book=qty_book,
        qty_doc=qty_doc,
        amt_book=amt_book,
        invoice_amt=invoice_amt,
        posting_date=posting,
        receipt_date=receipt_d,
        period_end=period_end,
        formula_conflict=conflict,
        raw_exception=exception,
        transport=str(transport_raw or ""),
    )
    # X 列写人话；技术原文进日志字段
    exception_for_cell = exception_nl or exception

    return {
        "sample_no": sample_no,
        "chain_id": chain_id,
        "voucher": _voucher_no(invoice),
        "posting_date": _parse_date(posting) or str(posting or ""),
        "customer": _f(invoice, "buyerName", "customerName")
        or _f(order, "buyerName", "customerName")
        or _f(contract, "buyerName", "customerName")
        or "",
        "seller": _f(contract, "sellerName", "vendorName", "supplierName") or "",
        "amt_book": amt_book,
        "qty_book": qty_book,
        "invoice_no": _f(invoice, "invoiceNo") or "",
        "invoice_amt": invoice_amt,
        "order_no": order_no or "",
        "has_order": bool(order_no),
        "transport": transport_raw,
        "transport_raw": transport_raw,
        "trading_mode": tm_view,
        "trading_mode_cells": tm_cells,
        "delivery_type": (
            str(tm_cells.get("N_delivery_document_type") or "") or delivery_type
        ),
        "delivery_no": (
            str(tm_cells.get("O_delivery_document_no") or "")
            or _f(receipt, "documentNo", "deliveryNo", "acceptanceNo")
            or _f(delivery, "documentNo", "deliveryNo")
            or ""
        ),
        "receipt_date": receipt_d,
        "qty_doc": qty_doc,
        "amt_delivery": amt_delivery,
        "period_ok": period_indep,  # 独立判断，供日志；不写 V
        "ar_period_ok": (assertions.get("ar_period") or {}).get("verdict_label")
        or assertions.get("ar_period_label")
        or "",
        "formula_v": formula_v,
        "formula_conflict": conflict,
        "all_ok": all_ok_label,
        "exception": exception_for_cell,
        "exception_tech": exception,
        "assertions": assertions,
    }


def build_gospd01030_sample_rows(
    job: dict[str, Any],
    *,
    w_yes: str = W_YES_FALLBACK,
    w_no: str = W_NO_FALLBACK,
    period_end: Optional[date] = None,
    skipped_chains: Optional[list[dict[str, Any]]] = None,
) -> list[dict[str, Any]]:
    """只导出 Gate4 已确认的业务链；未确认链记入 skipped_chains（NOT_TESTED）。"""
    from src.workflow.chain_workspace import sample_matching_ok

    classified = list(job.get("classified") or [])
    # 01030 严格归链：禁止「唯一桶弱并入」与「唯一 SO+HT 盲合并」
    chains = group_classified_by_chain(
        classified,
        allow_weak_unique_attach=False,
        allow_unique_so_ht_merge=False,
    )
    if not chains:
        return []
    pe = period_end if period_end is not None else _resolve_period_end(job)
    rows: list[dict[str, Any]] = []
    skip_buf = skipped_chains if skipped_chains is not None else []
    sample_no = 0
    for chain_id, docs in chains:
        if chain_id == "未识别业务号":
            skip_buf.append(
                {
                    "chain_id": chain_id,
                    "reason": "未识别业务号，禁止猜测并链",
                    "status": "NOT_TESTED",
                }
            )
            continue
        if not sample_matching_ok(job, chain_id):
            skip_buf.append(
                {
                    "chain_id": chain_id,
                    "reason": "Gate4 勾稽未确认，不写入正式样本",
                    "status": "NOT_TESTED",
                }
            )
            continue
        sample_no += 1
        rows.append(
            _row_from_chain(
                sample_no=sample_no,
                chain_id=chain_id,
                docs=docs,
                job=job,
                w_yes=w_yes,
                w_no=w_no,
                period_end=pe,
            )
        )
    return rows


def _copy_row_style(ws: Worksheet, src_row: int, dst_row: int, max_col: int = 24) -> None:
    for c in range(1, max_col + 1):
        sc = ws.cell(src_row, c)
        dc = ws.cell(dst_row, c)
        if sc.has_style:
            dc.font = copy(sc.font)
            dc.border = copy(sc.border)
            dc.fill = copy(sc.fill)
            dc.number_format = sc.number_format
            dc.protection = copy(sc.protection)
            dc.alignment = copy(sc.alignment)
    # 公式列随行号延展
    ws.cell(dst_row, COL["diff_inv"], f"=F{dst_row}-J{dst_row}")
    ws.cell(dst_row, COL["diff_amt"], f"=F{dst_row}-R{dst_row}")
    ws.cell(dst_row, COL["diff_qty"], f"=G{dst_row}-Q{dst_row}")
    ws.cell(dst_row, COL["period_ok"], f'=IF(P{dst_row}>$M$5,"YES 是","No 否")')


def _clear_grey(cell) -> None:
    cell.fill = NO_FILL


def _clear_unused_preset_rows(ws: Worksheet, used_count: int) -> None:
    """清空模板预置但未写入的样本行，去掉「签收确认 / [键入文件类型] / W=是」噪音。"""
    clear_cols = [
        COL["voucher"],
        COL["posting_date"],
        COL["customer"],
        COL["amt_book"],
        COL["qty_book"],
        COL["invoice_no"],
        COL["invoice_amt"],
        COL["order_no"],
        COL["transport"],
        COL["delivery_type"],
        COL["delivery_no"],
        COL["receipt_date"],
        COL["qty_doc"],
        COL["amt_delivery"],
        COL["all_ok"],
        COL["exception"],
    ]
    for r in range(DATA_START_ROW + max(used_count, 0), SAMPLE_GRID_END_ROW + 1):
        for c in clear_cols:
            cell = ws.cell(r, c)
            cell.value = None
            if c in (COL["amt_delivery"], COL["exception"]):
                cell.fill = GREY_FILL
            else:
                _clear_grey(cell)
        # 保留 V 公式；差额公式随空行仍可延展
        ws.cell(r, COL["diff_inv"], f"=F{r}-J{r}")
        ws.cell(r, COL["diff_amt"], f"=F{r}-R{r}")
        ws.cell(r, COL["diff_qty"], f"=G{r}-Q{r}")
        ws.cell(r, COL["period_ok"], f'=IF(P{r}>$M$5,"YES 是","No 否")')


def _ensure_log_sheet(wb):
    if LOG_SHEET in wb.sheetnames:
        ws = wb[LOG_SHEET]
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        return ws
    ws = wb.create_sheet(LOG_SHEET)
    headers = [
        "行号",
        "样本/合同号",
        "字段",
        "写入值摘要",
        "依据来源",
        "状态码",
        "冲突/旁注",
        "时间",
    ]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(1, i, h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="E7E6E6")
        cell.border = THIN
    widths = [8, 18, 14, 36, 40, 28, 48, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws


def _log(
    log_ws,
    *,
    excel_row: int,
    sample_id: str,
    field: str,
    value_summary: str,
    source: str,
    status: str,
    note: str = "",
) -> None:
    r = log_ws.max_row + 1
    vals = [
        excel_row,
        sample_id,
        field,
        value_summary,
        source,
        status,
        note,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    ]
    for c, v in enumerate(vals, start=1):
        cell = log_ws.cell(r, c, v)
        cell.border = THIN
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if status in (
            "FORMULA_LOGIC_CONFLICT",
            "TEMPLATE_VALIDATION_CONFLICT",
            "INSUFFICIENT_EVIDENCE",
            "NOT_APPLICABLE",
            "NOT_TESTED",
        ):
            cell.fill = CONFLICT_FILL


def fill_gospd01030_workbook(
    job: dict[str, Any],
    output_path: Path,
    *,
    entity_name: str = "",
    currency: str = "人民币",
    unit: str = "Yuan 元",
    period_end: str = "",
) -> Path:
    template = resolve_template_path()
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template, output_path)
    try:
        output_path.chmod(0o666)
    except OSError:
        pass

    wb = load_workbook(output_path)
    ws = wb[wb.sheetnames[0]]
    w_yes, w_no = _w_allowed_labels(ws)
    transport_opts = _e13_transport_options(ws)

    # 表头（改进模板坐标）— 期间截止日与断言同源；未配置则留空并记 NOT_TESTED
    pe = _resolve_period_end(job, period_end)
    entity = _resolve_entity_name(job, explicit=entity_name)
    ws["C5"] = entity
    ws["F5"] = "GOSPD01030"
    ws["I5"] = currency or "人民币"
    ws["K5"] = unit or "Yuan 元"
    ws["M5"] = pe

    # 保证断言层也能读到同一期末（仅当已配置）
    if pe is not None and not job.get("period_end"):
        job = {**job, "period_end": pe.isoformat()}

    skipped: list[dict[str, Any]] = []
    rows = build_gospd01030_sample_rows(
        job, w_yes=w_yes, w_no=w_no, period_end=pe, skipped_chains=skipped
    )
    from src.workflow.workbook_row_edits import apply_edits_to_rows

    rows = apply_edits_to_rows(rows, job, fmt="gospd01030")
    if not entity:
        # 仍空时用合同卖方兜底（样本行已解析）
        for row in rows:
            if row.get("seller"):
                entity = str(row["seller"])
                ws["C5"] = entity
                break

    log_ws = _ensure_log_sheet(wb)
    if pe is None:
        _log(
            log_ws,
            excel_row=5,
            sample_id="HEADER",
            field="期间截止日M5",
            value_summary="(空)",
            source="job.period_end 未配置",
            status="NOT_TESTED",
            note="GOSPD01030 须配置报告期末日后再导出正式结论",
        )
    for sk in skipped:
        _log(
            log_ws,
            excel_row=0,
            sample_id=str(sk.get("chain_id") or ""),
            field="样本行",
            value_summary="跳过",
            source="build_gospd01030_sample_rows",
            status=str(sk.get("status") or "NOT_TESTED"),
            note=str(sk.get("reason") or ""),
        )
    # 评量询问（2026-08 改进版）：
    # B13 运输条款 ↔ E13 运输/交付模式下拉
    # B14 是否无需检查系统发票 ↔ E14 YES/No
    # B15 是否检查销售订单 ↔ E15 YES/No
    # 多样本不一致时 E13 留空，禁止用首行冒充实体统一条款
    transport_std, e13_note = _resolve_eval_e13(rows, transport_opts)
    ws["E13"] = transport_std or None
    _log(
        log_ws,
        excel_row=13,
        sample_id="EVAL",
        field="E13",
        value_summary=str(ws["E13"].value) if ws["E13"].value else "(空)",
        source="评量：遵循哪些运输条款 → E13 下拉允许值",
        status="OK" if transport_std else "NOT_APPLICABLE",
        note=e13_note,
    )

    has_any_invoice = any(bool(r.get("invoice_no") or r.get("invoice_amt")) for r in rows)
    # 「是否无需检查系统发票」：有发票样本则需要检查 → No 否；否则 YES 是
    e14 = "No 否" if has_any_invoice else w_yes
    ws["E14"] = e14
    _log(
        log_ws,
        excel_row=14,
        sample_id="EVAL",
        field="E14",
        value_summary=str(e14),
        source="评量：是否无需检查系统发票（有发票号/金额→否=需要检查）",
        status="OK",
    )

    has_any_order = any(r.get("has_order") for r in rows)
    ws["E15"] = w_yes if has_any_order else "No 否"
    _log(
        log_ws,
        excel_row=15,
        sample_id="EVAL",
        field="E15",
        value_summary=str(ws["E15"].value),
        source="评量：是否检查销售订单（有订单样本→是）",
        status="OK",
    )
    _log(
        log_ws,
        excel_row=5,
        sample_id="HEADER",
        field="M5_期间截止日",
        value_summary=str(pe),
        source="resolve_period_end(job)，覆盖模板默认值",
        status="OK",
    )
    _log(
        log_ws,
        excel_row=5,
        sample_id="HEADER",
        field="C5_被审计单位",
        value_summary=entity or "(空)",
        source="job.entity/合同卖方",
        status="OK" if entity else "INSUFFICIENT_EVIDENCE",
    )

    for i, row in enumerate(rows):
        r = DATA_START_ROW + i
        if i >= MAX_SAMPLE_ROWS or (
            r > DATA_START_ROW and ws.cell(r, COL["sample_no"]).value in (None, "")
        ):
            _copy_row_style(ws, DATA_START_ROW, r)

        sid = str(row.get("chain_id") or row.get("order_no") or i + 1)

        ws.cell(r, COL["sample_no"], row.get("sample_no") or (i + 1))
        ws.cell(r, COL["voucher"], row.get("voucher") or "")
        posting = row.get("posting_date")
        if isinstance(posting, date):
            ws.cell(r, COL["posting_date"], posting)
        else:
            ws.cell(r, COL["posting_date"], posting or "")
        ws.cell(r, COL["customer"], row.get("customer") or "")
        if row.get("amt_book") is not None:
            ws.cell(r, COL["amt_book"], row["amt_book"])
        if row.get("qty_book") is not None:
            ws.cell(r, COL["qty_book"], row["qty_book"])
        ws.cell(r, COL["invoice_no"], row.get("invoice_no") or "")
        if row.get("invoice_amt") is not None:
            ws.cell(r, COL["invoice_amt"], row["invoice_amt"])

        # L 订单：有订单则启用并写入；无则保持灰度空白
        order_no = row.get("order_no") or ""
        if order_no:
            cell_l = ws.cell(r, COL["order_no"], order_no)
            _clear_grey(cell_l)
        else:
            ws.cell(r, COL["order_no"]).value = None

        ws.cell(r, COL["transport"], _normalize_transport(row.get("transport"), transport_opts))
        ws.cell(r, COL["delivery_type"], row.get("delivery_type") or "")
        ws.cell(r, COL["delivery_no"], row.get("delivery_no") or "")
        rd = row.get("receipt_date")
        if isinstance(rd, date):
            ws.cell(r, COL["receipt_date"], rd)
        elif rd:
            ws.cell(r, COL["receipt_date"], rd)
        if row.get("qty_doc") is not None:
            ws.cell(r, COL["qty_doc"], row["qty_doc"])

        # R：仅有可比金额才写；否则空白+灰度，禁止 0
        amt_d = row.get("amt_delivery")
        cell_r = ws.cell(r, COL["amt_delivery"])
        if amt_d is not None:
            cell_r.value = round(float(amt_d), 2)
            _clear_grey(cell_r)
            _log(
                log_ws,
                excel_row=r,
                sample_id=sid,
                field="R_交货金额",
                value_summary=str(cell_r.value),
                source="交货/签收单据金额（>0）",
                status="OK",
            )
        else:
            cell_r.value = None
            cell_r.fill = GREY_FILL
            _log(
                log_ws,
                excel_row=r,
                sample_id=sid,
                field="R_交货金额",
                value_summary="(空/灰度)",
                source="交货单无可比金额；禁止填0、禁止用账面回填",
                status="NOT_APPLICABLE",
            )

        # K/S/T/V：只保留/延展公式
        ws.cell(r, COL["diff_inv"], f"=F{r}-J{r}")
        ws.cell(r, COL["diff_amt"], f"=F{r}-R{r}")
        ws.cell(r, COL["diff_qty"], f"=G{r}-Q{r}")
        v_formula = f'=IF(P{r}>$M$5,"YES 是","No 否")'
        ws.cell(r, COL["period_ok"], v_formula)
        _log(
            log_ws,
            excel_row=r,
            sample_id=sid,
            field="V_公式",
            value_summary=v_formula,
            source="保留模板公式，禁止写死结论",
            status="OK",
        )

        conflict = str(row.get("formula_conflict") or "")
        _log(
            log_ws,
            excel_row=r,
            sample_id=sid,
            field="V_vs_独立判断",
            value_summary=f"独立={row.get('period_ok') or '(空)'} / 公式口径={row.get('formula_v') or '(空)'}",
            source="assert_correct_accounting_period vs IF(P>$M$5)",
            status="FORMULA_LOGIC_CONFLICT" if conflict else "OK",
            note=conflict,
        )
        # 步骤3：应收账款期间（无独立列，写入日志；结论已并入 W/X）
        ar_label = str(row.get("ar_period_ok") or "")
        ar_st = ((row.get("assertions") or {}).get("ar_period") or {}).get(
            "evidence_status"
        ) or ("OK" if ar_label else "NOT_TESTED")
        _log(
            log_ws,
            excel_row=r,
            sample_id=sid,
            field="步骤3_应收账款期间",
            value_summary=ar_label or "(空)",
            source="assert_ar_correct_period（与收入期间同源过账日；非函证）",
            status=str(ar_st),
        )

        # W：仅写验证允许值；证据不足/冲突时留空（可被 Gate5 审计师覆写）
        all_ok = row.get("all_ok") or ""
        if all_ok:
            ws.cell(r, COL["all_ok"], all_ok)
            if row.get("auditor_edited"):
                _log(
                    log_ws,
                    excel_row=r,
                    sample_id=sid,
                    field="W_综合结论",
                    value_summary=str(all_ok)[:120],
                    source="Gate5 审计师覆写",
                    status="OK",
                )
        else:
            ws.cell(r, COL["all_ok"]).value = None
            _log(
                log_ws,
                excel_row=r,
                sample_id=sid,
                field="W_综合结论",
                value_summary="(空)",
                source="冲突/未测完时不写「是」"
                + ("；含审计师覆写清空" if row.get("auditor_edited") else ""),
                status="NOT_TESTED" if not conflict else "FORMULA_LOGIC_CONFLICT",
            )

        # X：有事项才启用（可被 Gate5 覆写）；默认写自然语言，技术原文进日志
        exc = str(row.get("exception") or "").strip()
        cell_x = ws.cell(r, COL["exception"])
        if exc:
            cell_x.value = exc
            _clear_grey(cell_x)
            tech = str(row.get("exception_tech") or "").strip()
            _log(
                log_ws,
                excel_row=r,
                sample_id=sid,
                field="X_异常说明",
                value_summary=exc[:160],
                source="自然语言解释（规则结论）"
                + ("；含审计师覆写" if row.get("auditor_edited") else ""),
                status="OK",
                note=(tech[:240] if tech and tech != exc else ""),
            )
        else:
            cell_x.value = None
            cell_x.fill = GREY_FILL

    # 清空未使用的模板预置样本行，避免假「签收确认/是」
    _clear_unused_preset_rows(ws, len(rows))

    # 填制纪律/系统说明：不得写入样本明细区（B30:W48），放入日志页 + 模板「注释」区下方
    goals = (job.get("plan") or {}).get("goal_ids") or job.get("goal_ids") or []
    fill_note = (
        "— 本底稿由工作台按目标 gospd01030 自动回填；"
        f"已选目标={','.join(map(str, goals)) or 'gospd01030'}；"
        f"样本链={','.join(str(r.get('chain_id') or '') for r in rows)}。"
        "K/S/T/V 保留公式；R 无金额不填0；E13 仅多样本一致时写运输条款；"
        "E14 是否无需检查系统发票；E15 是否检查销售订单；"
        f"详见工作表「{LOG_SHEET}」。"
    )
    _log(
        log_ws,
        excel_row=0,
        sample_id="FILL_NOTE",
        field="填制说明",
        value_summary=fill_note,
        source="系统运行说明（不进样本表）",
        status="OK",
    )
    note_row = TEMPLATE_COMMENT_END_ROW + 1  # B53，位于模板注释之后
    if note_row <= SAMPLE_GRID_END_ROW:
        note_row = SAMPLE_GRID_END_ROW + 2
    # 若样本扩展超过预置区，说明写在末样本行之后、且避开预置网格
    last_sample = DATA_START_ROW + max(len(rows) - 1, 0)
    if last_sample >= note_row:
        note_row = last_sample + 2
    ws.cell(note_row, 2, fill_note)
    wb.save(output_path)
    return output_path
