"""按官方 GOSPD01010.3 模板填表：交易价格适当性抽凭。

列映射（表头行 17，数据起始行 20，预置 1–20）：
B 样本编号 | C 会计分录编号 | D 合同索引号 | E 客户名称 | F 金额 | G 合同编号
H 检查其他相关文件(若适用) | I 文件类型 | J 文件索引
K 合同交易价格是否需要计算？ | L 计算方式描述
M 是否已适当确定交易价格？ | N 异常说明
"""

from __future__ import annotations

import shutil
from copy import copy
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.audit.gospd01010_3_assertions import build_gospd01010_3_assertions
from src.reporting.gospd01010_filler import group_classified_by_chain

ROOT = Path(__file__).resolve().parents[2]
TEMPLATE_CANDIDATES = (
    ROOT / "templates" / "GOSPD01010.3.xlsx",
    ROOT / "templates" / "GOSPD01010_3.xlsx",
)

DATA_START_ROW = 20
MAX_SAMPLE_ROWS = 20

COL = {
    "sample_no": 2,  # B
    "voucher": 3,  # C
    "contract_idx": 4,  # D
    "customer": 5,  # E
    "amount": 6,  # F
    "contract_no": 7,  # G
    "other_applicable": 8,  # H
    "other_file_type": 9,  # I
    "other_file_index": 10,  # J
    "needs_calc": 11,  # K
    "calc_method": 12,  # L
    "price_ok": 13,  # M
    "exception": 14,  # N
}


def resolve_template_path() -> Path:
    for p in TEMPLATE_CANDIDATES:
        if p.is_file():
            return p
    raise FileNotFoundError(
        "未找到 GOSPD01010.3 模板，请将文件放到 templates/GOSPD01010.3.xlsx"
    )


def _f(doc: Optional[dict[str, Any]], *keys: str) -> Any:
    if not doc:
        return None
    from src.models.field_values import rule_readable_fields

    fields = rule_readable_fields(doc)
    for k in keys:
        if fields.get(k) not in (None, ""):
            return fields.get(k)
        if doc.get(k) not in (None, ""):
            return doc.get(k)
    return None


def _num(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(
            str(v)
            .replace(",", "")
            .replace("，", "")
            .replace("¥", "")
            .replace("元", "")
            .replace("CNY", "")
            .strip()
        )
    except ValueError:
        return None


def _by_type(docs: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    out: dict[str, dict[str, Any]] = {}
    for d in docs or []:
        t = str(d.get("doc_type") or "")
        if t and t not in out:
            out[t] = d
    return out


def _row_from_chain(
    *,
    sample_no: int,
    chain_id: str,
    docs: list[dict[str, Any]],
    job: dict[str, Any],
    apply_job_tests: bool,
) -> dict[str, Any]:
    by = _by_type(docs)
    contract = by.get("contract")
    order = by.get("order")
    invoice = by.get("invoice")

    assertions = build_gospd01010_3_assertions(
        docs=docs,
        job=job,
        chain_id=chain_id,
        apply_job_tests=apply_job_tests,
    )

    voucher = (
        (invoice or {}).get("ledger_voucher")
        or _f(invoice, "voucherNo", "documentNo")
        or ""
    )
    contract_no = (
        _f(contract, "contractNo", "documentNo")
        or _f(order, "contractNo")
        or (chain_id if "HT" in str(chain_id).upper() else "")
    )
    customer = (
        _f(invoice, "buyerName", "customerName")
        or _f(order, "buyerName", "customerName")
        or _f(contract, "buyerName", "customerName")
        or ""
    )
    amt = None
    for doc in (invoice, order, contract):
        n = _num(_f(doc, "totalAmount", "amount", "grossAmount"))
        if n is not None and n > 0:
            amt = n
            break
    if amt is None and isinstance(job.get("amount_test"), dict):
        ar = job["amount_test"].get("accuracy_report") or {}
        if isinstance(ar, dict):
            amt = _num((ar.get("ledger_values") or {}).get("ledger_debit_total"))

    exc = str(assertions.get("exception") or "").strip()
    appendix = str(assertions.get("exception_appendix") or "").strip()
    if appendix and appendix not in exc:
        from src.audit.workpaper_notes import merge_exception_text

        exc = merge_exception_text(exc, appendix)

    return {
        "sample_no": sample_no,
        "chain_id": chain_id,
        "voucher": voucher,
        "contract_idx": contract_no,
        "customer": customer,
        "amount": amt,
        "contract_no": contract_no,
        "other_applicable": assertions.get("other_applicable") or "Not applicable 不适用",
        "other_file_type": assertions.get("other_file_type") or "",
        "other_file_index": assertions.get("other_file_index") or "",
        "needs_calc": assertions.get("needs_calc_label") or "NO 否",
        "calc_method": assertions.get("calc_method") or "",
        "price_ok": assertions.get("price_ok_label") or "",
        "exception": exc,
        "system_observation": assertions.get("system_observation") or "",
        "pending_judgment": assertions.get("pending_judgment") or "",
        "assertions": assertions,
    }


def build_gospd01010_3_sample_rows(job: dict[str, Any]) -> list[dict[str, Any]]:
    classified = list(job.get("classified") or [])
    chains = group_classified_by_chain(classified)
    if not chains:
        return []
    samples = (
        job.get("gospd_sample_results")
        if isinstance(job.get("gospd_sample_results"), dict)
        else {}
    )
    single = len(chains) == 1
    rows: list[dict[str, Any]] = []
    for i, (chain_id, docs) in enumerate(chains, start=1):
        sample = samples.get(chain_id) or {}
        has_per = bool(sample.get("contract_terms") or sample.get("amount_test"))
        apply = has_per or single
        rows.append(
            _row_from_chain(
                sample_no=i,
                chain_id=chain_id,
                docs=docs,
                job=job,
                apply_job_tests=apply,
            )
        )
    return rows


def _copy_row_style(ws: Worksheet, src_row: int, dst_row: int) -> None:
    for c in range(1, 15):
        sc = ws.cell(src_row, c)
        dc = ws.cell(dst_row, c)
        if sc.has_style:
            dc._style = copy(sc._style)
        dc.number_format = sc.number_format


def fill_gospd01010_3_workbook(
    job: dict[str, Any],
    output_path: Path,
    *,
    entity_name: str = "",
    currency: str = "人民币",
    unit: str = "Yuan 元",
) -> Path:
    template = resolve_template_path()
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template, output_path)
    try:
        output_path.chmod(0o666)
    except OSError:
        pass

    wb = load_workbook(output_path)
    ws = wb[wb.sheetnames[0]]

    ws.cell(5, 3, entity_name or ws.cell(5, 3).value or "")
    ws.cell(5, 5, "GOSPD01010.3")
    ws.cell(5, 7, currency)
    ws.cell(5, 9, unit)

    rows = build_gospd01010_3_sample_rows(job)
    for i, row in enumerate(rows):
        r = DATA_START_ROW + i
        if i >= MAX_SAMPLE_ROWS or (
            r > DATA_START_ROW and ws.cell(r, COL["sample_no"]).value in (None, "")
        ):
            _copy_row_style(ws, DATA_START_ROW, r)

        ws.cell(r, COL["sample_no"], row.get("sample_no") or (i + 1))
        ws.cell(r, COL["voucher"], row.get("voucher") or "")
        ws.cell(r, COL["contract_idx"], row.get("contract_idx") or "")
        ws.cell(r, COL["customer"], row.get("customer") or "")
        if row.get("amount") is not None:
            ws.cell(r, COL["amount"], row["amount"])
        ws.cell(r, COL["contract_no"], row.get("contract_no") or "")
        ws.cell(r, COL["other_applicable"], row.get("other_applicable") or "")
        ws.cell(r, COL["other_file_type"], row.get("other_file_type") or "")
        ws.cell(r, COL["other_file_index"], row.get("other_file_index") or "")
        ws.cell(r, COL["needs_calc"], row.get("needs_calc") or "")
        ws.cell(r, COL["calc_method"], row.get("calc_method") or "")
        ws.cell(r, COL["price_ok"], row.get("price_ok") or "")
        ws.cell(r, COL["exception"], row.get("exception") or "")

    note_row = max(45, DATA_START_ROW + max(len(rows), MAX_SAMPLE_ROWS) + 3)
    goals = (job.get("plan") or {}).get("goal_ids") or job.get("goal_ids") or []
    ws.cell(
        note_row,
        2,
        "— 本底稿由工作台按目标 gospd01010_3 自动回填；"
        f"已选目标={','.join(map(str, goals)) or 'gospd01010_3'}；"
        f"样本链={','.join(str(r.get('chain_id') or '') for r in rows)}。"
        "H/K/M 仅用模板下拉枚举；M 列 YES 含前导空格以匹配模板。"
        "交易价格结论优先交易对价维度 + 金额测试。",
    )
    wb.save(output_path)
    return output_path
