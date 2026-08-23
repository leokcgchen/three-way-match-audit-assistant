"""按官方 GOSPD01010.1 模板填表：期内销售收入抽凭实质性程序。

列映射（Sheet1，数据起始行 22）：
B 样本编号 | C 会计分录编号 | D 客户名称 | E 入账数量(A) | F 入账金额(X)
G 合同索引号 | H 主要合同条款 | I 销售订单编号 | J 运输条款
K 交付单据类型 | L 交付单据编号 | M 签收日期 | N 交货数量(B) | O 销售金额(Y)
P/Q 差异公式已在模板 | R 步骤1合同可执行 | S 2.1金额准确 | T 2.2控制权转移
U 步骤均无异常 | V 异常说明
"""

from __future__ import annotations

import shutil
from copy import copy
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parents[2]
TEMPLATE_CANDIDATES = (
    ROOT / "templates" / "GOSPD01010.xlsx",
    ROOT / "templates" / "GOSPD01010_program.xlsx",
    ROOT / "templates" / "GOSPD01010_style_ref.xlsx",
)

# 模板数据区
DATA_START_ROW = 22
MAX_SAMPLE_ROWS = 16  # 模板预置 22–37

COL = {
    "sample_no": 2,  # B
    "voucher": 3,  # C
    "customer": 4,  # D
    "qty_book": 5,  # E
    "amt_book": 6,  # F
    "contract_idx": 7,  # G
    "contract_terms": 8,  # H
    "order_no": 9,  # I
    "transport": 10,  # J
    "delivery_type": 11,  # K
    "delivery_no": 12,  # L
    "receipt_date": 13,  # M
    "qty_doc": 14,  # N
    "amt_sales": 15,  # O
    "step1": 18,  # R
    "step21": 19,  # S
    "step22": 20,  # T
    "all_ok": 21,  # U
    "exception": 22,  # V
}


def resolve_template_path() -> Path:
    for p in TEMPLATE_CANDIDATES:
        if p.is_file():
            return p
    raise FileNotFoundError(
        "未找到 GOSPD01010 模板，请将文件放到 templates/GOSPD01010.xlsx"
    )


def _f(doc: Optional[dict[str, Any]], *keys: str) -> Any:
    if not doc:
        return None
    from src.models.field_values import rule_readable_fields

    fields = rule_readable_fields(doc)
    for k in keys:
        v = fields.get(k)
        if v is not None and str(v).strip() and str(v).strip().lower() not in {
            "none",
            "null",
            "nan",
            "-",
        }:
            return v
    return None


def _by_type(classified: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    out: dict[str, dict[str, Any]] = {}
    for item in classified or []:
        dt = str(item.get("doc_type") or "other")
        if dt == "other":
            continue
        prev = out.get(dt)
        if prev is None:
            out[dt] = item
            continue
        score = len([v for v in (item.get("fields") or {}).values() if v not in (None, "")])
        prev_score = len(
            [v for v in (prev.get("fields") or {}).values() if v not in (None, "")]
        )
        if score >= prev_score:
            out[dt] = item
    return out


def _num(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    try:
        return float(str(value).replace(",", "").replace("¥", "").replace("元", "").strip())
    except (TypeError, ValueError):
        return None


def _yn(ok: Optional[bool]) -> str:
    if ok is True:
        return "Yes 是"
    if ok is False:
        return "No 否"
    return ""


def _status_ok(status: Any) -> Optional[bool]:
    s = str(status or "").upper()
    if not s:
        return None
    if s in {"PASS", "OK", "通过"}:
        return True
    if s in {"FAIL", "ERROR", "未通过"}:
        return False
    if s in {"WARNING", "WARN", "需关注"}:
        return False
    return None


def _contract_step1_ok(contract: Optional[dict[str, Any]]) -> Optional[bool]:
    if not contract:
        return None
    status = contract.get("status")
    report = contract.get("clarity_report") or contract.get("report") or {}
    tr = report.get("test_result") or {}
    st = status or tr.get("test_status")
    ok = _status_ok(st)
    if ok is False and str(st).upper() == "WARNING":
        return True
    return ok


def _amount_step21_ok(amount: Optional[dict[str, Any]]) -> Optional[bool]:
    if not amount:
        return None
    status = amount.get("status")
    ar = amount.get("accuracy_report") or {}
    at = ar.get("amount_test") or {}
    return _status_ok(status or at.get("test_status"))


def _control_step22_ok(three_way: Optional[dict[str, Any]]) -> Optional[bool]:
    if not three_way:
        return None
    match = three_way.get("match_result") or {}
    if hasattr(match, "model_dump"):
        match = match.model_dump()
    cutoff = three_way.get("cutoff_result") or {}
    if hasattr(cutoff, "model_dump"):
        cutoff = cutoff.model_dump()
    m_ok = _status_ok(three_way.get("overall_status") or match.get("overall_status"))
    c_status = cutoff.get("测试状态") if isinstance(cutoff, dict) else None
    c_ok = _status_ok(c_status) if c_status else None
    if m_ok is False or c_ok is False:
        return False
    if m_ok is True or c_ok is True:
        return True
    return None


def _norm_biz(value: Any) -> str:
    from src.legacy_ocr.ledger_parser import normalize_biz_id

    return normalize_biz_id(value) if value is not None else ""


def _doc_all_keys(doc: dict[str, Any]) -> list[str]:
    from src.legacy_ocr.ledger_parser import (
        collect_document_biz_keys,
        extract_biz_ids_from_filename,
    )

    keys: list[str] = []
    for k in collect_document_biz_keys(dict(doc.get("fields") or {})):
        nk = _norm_biz(k)
        if nk and nk not in keys:
            keys.append(nk)
    for k in extract_biz_ids_from_filename(str(doc.get("file_name") or "")):
        nk = _norm_biz(k)
        if nk and nk not in keys:
            keys.append(nk)
    return keys


def _is_strong_biz(key: str) -> bool:
    """链主键只用订单/合同类编号，避免发票号、签收号拆成多笔。"""
    u = (key or "").upper()
    if not u:
        return False
    if u.startswith("SO") or u.startswith("PO"):
        return True
    if "HT" in u or u.startswith("CT") or u.startswith("CONTRACT"):
        return True
    return False


def _prefer_chain_key(keys: list[str]) -> str:
    """销售订单号优先（取更完整的编号，避免 SO25-002 盖住 SO25-0021），其次合同号。"""
    from src.legacy_ocr.ledger_parser import compact_biz_id

    strong = [k for k in keys if _is_strong_biz(k)]
    pool = strong or list(keys)
    if not pool:
        return ""
    so_keys = [k for k in pool if k.upper().startswith("SO") or k.upper().startswith("PO")]
    if so_keys:
        return max(so_keys, key=lambda k: (len(compact_biz_id(k)), k))
    for k in pool:
        u = k.upper()
        if "HT" in u or u.startswith("CT"):
            return k
    return pool[0]


def group_classified_by_chain(
    classified: list[dict[str, Any]],
    *,
    allow_weak_unique_attach: bool = True,
    allow_unique_so_ht_merge: bool = True,
) -> list[tuple[str, list[dict[str, Any]]]]:
    """同 job 多笔业务：按 SO/HT 等强业务号归链；同文件 SO+HT 合并。

    allow_weak_unique_attach:
        True（默认，01010 便利）— 无强号单据在「仅有一桶」时并入该桶；
        False（01030 严格）— 无明确引用则进「未识别业务号」，禁止猜测串单。
    allow_unique_so_ht_merge:
        True — 全任务唯一 SO + 唯一 HT 时自动 union；
        False — 仅当单据字段已互相引用时合并（order.contractNo 等）。
    """
    parent: dict[str, str] = {}

    def find(x: str) -> str:
        parent.setdefault(x, x)
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a: str, b: str) -> None:
        from src.legacy_ocr.ledger_parser import compact_biz_id

        ra, rb = find(a), find(b)
        if ra == rb:
            return
        ca, cb = compact_biz_id(ra), compact_biz_id(rb)
        if ca and cb and ca != cb:
            if cb.startswith(ca) and len(cb) > len(ca):
                parent[ra] = rb
                return
            if ca.startswith(cb) and len(ca) > len(cb):
                parent[rb] = ra
                return
        if ra.upper().startswith("SO") and not rb.upper().startswith("SO"):
            parent[rb] = ra
        elif rb.upper().startswith("SO") and not ra.upper().startswith("SO"):
            parent[ra] = rb
        else:
            parent[rb] = ra

    doc_keys: list[tuple[dict[str, Any], list[str], list[str]]] = []
    for doc in classified or []:
        all_keys = _doc_all_keys(doc)
        fields = dict(doc.get("fields") or {})
        # 字段级串联：订单号 ↔ 合同号（跨文件也能归同一笔）
        if doc.get("doc_type") in {"order", "invoice"}:
            so_raw = (
                fields.get("orderNo")
                or fields.get("salesOrderNo")
                or fields.get("documentNo")
            )
        else:
            so_raw = fields.get("orderNo") or fields.get("salesOrderNo")
        so = _norm_biz(so_raw)
        if doc.get("doc_type") == "contract":
            ht_raw = fields.get("contractNo") or fields.get("documentNo")
        else:
            ht_raw = fields.get("contractNo")
        ht = _norm_biz(ht_raw)
        if so and _is_strong_biz(so) and so not in all_keys:
            all_keys.append(so)
        if ht and _is_strong_biz(ht) and ht not in all_keys:
            all_keys.append(ht)
        strong = [k for k in all_keys if _is_strong_biz(k)]
        if so and ht and _is_strong_biz(so) and _is_strong_biz(ht):
            union(so, ht)
        doc_keys.append((doc, all_keys, strong))
        for i, a in enumerate(strong):
            for b in strong[i + 1 :]:
                union(a, b)

    # 仅合并同族截断号（SO25-002 ← SO25-0021），禁止跨 SO/HT 全局 union
    from src.legacy_ocr.ledger_parser import compact_biz_id as _compact_biz

    all_strong: list[str] = []
    for _, _, strong in doc_keys:
        for k in strong:
            if k not in all_strong:
                all_strong.append(k)
    for i, a in enumerate(all_strong):
        for b in all_strong[i + 1 :]:
            ca, cb = _compact_biz(a), _compact_biz(b)
            if not ca or not cb or ca == cb:
                continue
            if ca[:2] != cb[:2]:
                continue
            if cb.startswith(ca) or ca.startswith(cb):
                union(a, b)

    # 再扫一遍：合同只有 HT、订单只有 SO 时，用订单.contractNo 或「唯一 SO+唯一 HT」合并
    so_list: list[str] = []
    ht_list: list[str] = []
    for doc in classified or []:
        fields = dict(doc.get("fields") or {})
        if doc.get("doc_type") == "order":
            so = _norm_biz(fields.get("orderNo") or fields.get("documentNo"))
            ht = _norm_biz(fields.get("contractNo"))
            if so and _is_strong_biz(so):
                so_list.append(so)
            if so and ht and _is_strong_biz(ht):
                union(so, ht)
        if doc.get("doc_type") == "contract":
            ht = _norm_biz(fields.get("contractNo") or fields.get("documentNo"))
            if ht and _is_strong_biz(ht):
                ht_list.append(ht)
        if doc.get("doc_type") == "invoice":
            so = _norm_biz(fields.get("orderNo") or fields.get("salesOrderNo"))
            ht = _norm_biz(fields.get("contractNo"))
            if so and ht and _is_strong_biz(so) and _is_strong_biz(ht):
                union(so, ht)
    so_uniq = sorted({x for x in so_list if x})
    ht_uniq = sorted({x for x in ht_list if x})
    if allow_unique_so_ht_merge and len(so_uniq) == 1 and len(ht_uniq) == 1:
        union(so_uniq[0], ht_uniq[0])

    root_display: dict[str, str] = {}
    for _, _, strong in doc_keys:
        for k in strong:
            r = find(k)
            cand = _prefer_chain_key([k, r, root_display.get(r, r)])
            prev = root_display.get(r, r)
            root_display[r] = _prefer_chain_key([prev, cand, r])

    buckets: dict[str, list[dict[str, Any]]] = {}
    weak_docs: list[tuple[dict[str, Any], list[str]]] = []
    for doc, all_keys, strong in doc_keys:
        if not strong:
            weak_docs.append((doc, all_keys))
            continue
        root = find(strong[0])
        cid = root_display.get(root, root)
        buckets.setdefault(cid, []).append(doc)

    # 弱编号单据：字段里若引用某链 SO/HT 则并入；否则唯一链时（可关）并入
    for doc, all_keys in weak_docs:
        attached = False
        for cid in list(buckets.keys()):
            nk = _norm_biz(cid)
            # 发票/签收上的 orderNo 等已在 all_keys；也可能仅有弱号
            fields = dict(doc.get("fields") or {})
            refs = set(all_keys)
            for fk in ("orderNo", "contractNo", "salesOrderNo", "documentNo"):
                refs.add(_norm_biz(fields.get(fk)))
            if nk and nk in refs:
                buckets[cid].append(doc)
                attached = True
                break
            # 压缩匹配：SO250281 vs SO25-0281
            from src.legacy_ocr.ledger_parser import compact_biz_id

            c_nk = compact_biz_id(nk)
            if c_nk and any(compact_biz_id(x) == c_nk for x in refs if x):
                buckets[cid].append(doc)
                attached = True
                break
        if attached:
            continue
        if allow_weak_unique_attach and len(buckets) == 1:
            buckets[next(iter(buckets))].append(doc)
        elif not buckets:
            buckets["未识别业务号"] = [doc]
        else:
            buckets.setdefault("未识别业务号", []).append(doc)

    def _sort_key(item: tuple[str, list]) -> tuple:
        cid = item[0]
        u = cid.upper()
        pri = 0 if u.startswith("SO") else 1 if "HT" in u else 2
        return (pri, cid)

    return sorted(buckets.items(), key=_sort_key)


def _job_test_biz_keys(job: dict[str, Any]) -> set[str]:
    """从当前 job 级测试结果里抽出业务号，用于多行时归因。"""
    keys: set[str] = set()
    amount = job.get("amount_test") if isinstance(job.get("amount_test"), dict) else {}
    ar = (amount or {}).get("accuracy_report") or {}
    at = ar.get("amount_test") or {}
    for raw in (
        at.get("sales_order_no"),
        at.get("business_id"),
        (ar.get("source_fields") or {}).get("sales_order_no"),
        amount.get("sales_order_no"),
    ):
        nk = _norm_biz(raw)
        if nk:
            keys.add(nk)

    three_way = job.get("three_way") if isinstance(job.get("three_way"), dict) else {}
    req = (three_way or {}).get("match_request") or {}
    if hasattr(req, "model_dump"):
        req = req.model_dump()
    order = req.get("order") or {}
    for raw in (order.get("order_no"), order.get("contract_no")):
        nk = _norm_biz(raw)
        if nk:
            keys.add(nk)

    contract = (
        job.get("contract_terms") if isinstance(job.get("contract_terms"), dict) else {}
    )
    report = (contract or {}).get("clarity_report") or (contract or {}).get("report") or {}
    extracted = report.get("extracted") or {}
    for raw in (extracted.get("contract_id"), report.get("contract_id")):
        nk = _norm_biz(raw)
        if nk:
            keys.add(nk)

    evidence = job.get("evidence") if isinstance(job.get("evidence"), dict) else {}
    for raw in evidence.get("anchor_keys") or []:
        nk = _norm_biz(raw)
        if nk:
            keys.add(nk)
    return keys


def _chain_related(chain_id: str, docs: list[dict[str, Any]], test_keys: set[str]) -> bool:
    if not test_keys:
        return False
    pool = {_norm_biz(chain_id)}
    for d in docs:
        pool.update(_doc_all_keys(d))
    return bool(pool & test_keys)


def _row_from_chain(
    *,
    sample_no: int,
    chain_id: str,
    docs: list[dict[str, Any]],
    job: dict[str, Any],
    apply_job_tests: bool,
) -> dict[str, Any]:
    from src.audit.gospd_assertions import build_gospd_assertions

    by = _by_type(docs)
    contract = by.get("contract")
    order = by.get("order")
    delivery = by.get("delivery")
    receipt = by.get("receipt") or delivery
    invoice = by.get("invoice")

    qty_book = _num(
        (invoice or {}).get("ledger_quantity")
        or _f(invoice, "quantity")
        or _f(order, "quantity")
    )
    amt_book = _num(
        (invoice or {}).get("ledger_amount")
        or _f(invoice, "totalAmount", "amount")
        or _f(order, "totalAmount")
    )
    qty_doc = _num(_f(receipt, "quantity") or _f(delivery, "quantity") or qty_book)
    amt_sales = _num(
        _f(invoice, "totalAmount", "amount") or _f(order, "totalAmount") or amt_book
    )

    transport = (
        _f(contract, "transportTerms") or _f(order, "transportTerms") or "签收确认"
    )
    terms_bits = []
    for key, label in (
        ("paymentTerms", "付款"),
        ("controlTransferTerms", "控制权"),
        ("settlementTerms", "结算"),
    ):
        v = _f(contract, key) or _f(order, key)
        if v:
            terms_bits.append(f"{label}:{v}")
    contract_terms_text = "；".join(terms_bits) or _f(contract, "remarks") or ""

    delivery_type = (
        "客户签收验收单"
        if (receipt or {}).get("doc_type") == "receipt"
        else ("销售发货单" if delivery else "")
    )
    if receipt and "验收" in str(receipt.get("file_name") or ""):
        delivery_type = "产品验收单/签收单"

    assertions = build_gospd_assertions(
        docs=docs,
        job=job,
        chain_id=chain_id,
        apply_job_tests=apply_job_tests,
    )

    order_no = (
        _f(order, "orderNo", "documentNo", "salesOrderNo")
        or _f(invoice, "orderNo", "salesOrderNo")
        or (chain_id if str(chain_id).upper().startswith("SO") else "")
    )
    contract_no = (
        _f(contract, "contractNo", "documentNo")
        or _f(order, "contractNo")
        or (chain_id if "HT" in str(chain_id).upper() else "")
    )

    return {
        "sample_no": sample_no,
        "chain_id": chain_id,
        "voucher": (invoice or {}).get("ledger_voucher")
        or _f(invoice, "voucherNo", "documentNo")
        or "",
        "customer": _f(invoice, "buyerName", "customerName")
        or _f(order, "buyerName", "customerName")
        or _f(contract, "buyerName", "customerName")
        or "",
        "qty_book": qty_book,
        "amt_book": amt_book,
        "contract_idx": contract_no,
        "contract_terms": contract_terms_text,
        "order_no": order_no,
        "transport": transport,
        "delivery_type": delivery_type,
        "delivery_no": _f(receipt, "documentNo", "deliveryNo", "acceptanceNo")
        or _f(delivery, "documentNo", "deliveryNo")
        or "",
        "receipt_date": str(
            _f(receipt, "acceptanceDate", "deliveryDate", "documentDate", "receiptDate")
            or _f(delivery, "deliveryDate", "documentDate")
            or ""
        ),
        "qty_doc": qty_doc,
        "amt_sales": amt_sales,
        "step1": assertions["step1"].get("verdict_label") or "",
        "step21": assertions["step21"].get("verdict_label") or "",
        "step22": assertions["step22"].get("verdict_label") or "",
        "all_ok": assertions.get("all_ok_label") or "",
        "exception": assertions.get("exception") or "",
        "system_observation": assertions.get("system_observation") or "",
        "pending_judgment": assertions.get("pending_judgment") or "",
        "assertions": assertions,
    }


def build_gospd_sample_rows(job: dict[str, Any]) -> list[dict[str, Any]]:
    """同一 job 多笔业务 → 多行样本（写入同一份 GOSPD01010）。"""
    classified = list(job.get("classified") or [])
    chains = group_classified_by_chain(classified)
    if not chains:
        return []

    test_keys = _job_test_biz_keys(job)
    samples = job.get("gospd_sample_results") if isinstance(job.get("gospd_sample_results"), dict) else {}
    single = len(chains) == 1
    rows: list[dict[str, Any]] = []
    for i, (chain_id, docs) in enumerate(chains, start=1):
        sample = samples.get(chain_id) or {}
        has_per = bool(
            sample.get("contract_terms")
            or sample.get("amount_test")
            or sample.get("three_way")
        )
        # 有分笔结果 → 必用；单链或能归因时才套用 job 级旧结果
        apply = has_per or single or _chain_related(chain_id, docs, test_keys)
        rows.append(
            _row_from_chain(
                sample_no=i,
                chain_id=chain_id,
                docs=docs,
                job=job,
                apply_job_tests=apply,
            )
        )
    return rows


def _copy_row_style(ws: Worksheet, src_row: int, dst_row: int, max_col: int = 22) -> None:
    for c in range(1, max_col + 1):
        sc = ws.cell(src_row, c)
        dc = ws.cell(dst_row, c)
        if sc.has_style:
            dc.font = copy(sc.font)
            dc.border = copy(sc.border)
            dc.fill = copy(sc.fill)
            dc.number_format = sc.number_format
            dc.protection = copy(sc.protection)
            dc.alignment = copy(sc.alignment)
    # 差异公式
    ws.cell(dst_row, 16, f"=E{dst_row}-N{dst_row}")
    ws.cell(dst_row, 17, f"=F{dst_row}-O{dst_row}")


def fill_gospd01010_workbook(
    job: dict[str, Any],
    output_path: Path,
    *,
    entity_name: str = "",
    currency: str = "人民币",
    unit: str = "Yuan 元",
) -> Path:
    """复制官方模板并填入样本行，返回输出路径。"""
    template = resolve_template_path()
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template, output_path)
    # 解除只读拷贝属性
    try:
        output_path.chmod(0o666)
    except OSError:
        pass

    wb = load_workbook(output_path)
    ws = wb[wb.sheetnames[0]]

    # 元信息
    if entity_name:
        ws.cell(5, 3, entity_name)  # C5 被审计单位名称旁取值格因合并可能在 D；样表 B5 标签 C/D 值
    # 样表：B5 被审计单位名称 | D5? 实际 R5: B标签 D程序索引号 E=GOSPD...
    # 从探测：R5 B=被审计单位名称 | D=程序索引号 | E=GOSPD01010.1 | F=币种 | G=人民币 | H=单位 | I=Yuan
    # 被审计单位名称值格可能被合并到别处；写 C5 保险
    ws.cell(5, 3, entity_name or ws.cell(5, 3).value or "")
    ws.cell(5, 5, "GOSPD01010.1")
    ws.cell(5, 7, currency)
    ws.cell(5, 9, unit)

    # 评量询问
    ws.cell(12, 4, "No 否")  # 是否来自提供服务
    transport_default = "签收确认"
    rows = build_gospd_sample_rows(job)
    if rows and rows[0].get("transport"):
        transport_default = str(rows[0]["transport"])
    ws.cell(13, 4, transport_default)

    # 超出模板预置行时复制样式续行（同一份底稿追加多笔）
    for i, row in enumerate(rows):
        r = DATA_START_ROW + i
        if i >= MAX_SAMPLE_ROWS or (
            r > DATA_START_ROW and ws.cell(r, COL["sample_no"]).value in (None, "")
        ):
            _copy_row_style(ws, DATA_START_ROW, r)
        ws.cell(r, COL["sample_no"], row.get("sample_no") or (i + 1))
        ws.cell(r, COL["voucher"], row.get("voucher") or "")
        ws.cell(r, COL["customer"], row.get("customer") or "")
        if row.get("qty_book") is not None:
            ws.cell(r, COL["qty_book"], row["qty_book"])
        if row.get("amt_book") is not None:
            ws.cell(r, COL["amt_book"], row["amt_book"])
        ws.cell(r, COL["contract_idx"], row.get("contract_idx") or "")
        ws.cell(r, COL["contract_terms"], row.get("contract_terms") or "")
        ws.cell(r, COL["order_no"], row.get("order_no") or "")
        ws.cell(r, COL["transport"], row.get("transport") or "")
        ws.cell(r, COL["delivery_type"], row.get("delivery_type") or "")
        ws.cell(r, COL["delivery_no"], row.get("delivery_no") or "")
        ws.cell(r, COL["receipt_date"], row.get("receipt_date") or "")
        if row.get("qty_doc") is not None:
            ws.cell(r, COL["qty_doc"], row["qty_doc"])
        if row.get("amt_sales") is not None:
            ws.cell(r, COL["amt_sales"], row["amt_sales"])
        ws.cell(r, COL["step1"], row.get("step1") or "")
        ws.cell(r, COL["step21"], row.get("step21") or "")
        ws.cell(r, COL["step22"], row.get("step22") or "")
        ws.cell(r, COL["all_ok"], row.get("all_ok") or "")
        ws.cell(r, COL["exception"], row.get("exception") or "")
        ws.cell(r, 16, f"=E{r}-N{r}")
        ws.cell(r, 17, f"=F{r}-O{r}")

    note_row = max(45, DATA_START_ROW + len(rows) + 3)
    goals = (job.get("plan") or {}).get("goal_ids") or job.get("goal_ids") or []
    chain_ids = [str(r.get("chain_id") or "") for r in rows]
    ws.cell(
        note_row,
        2,
        "— 本底稿由工作台按目标 gospd01010 自动回填；"
        f"样本 {len(rows)} 笔（{', '.join(chain_ids)}）；"
        f"已选目标={','.join(map(str, goals)) or 'gospd01010'}；"
        "同 job 多业务写入同一份表；差异列 P/Q 为模板公式。",
    )

    wb.save(output_path)
    return output_path
