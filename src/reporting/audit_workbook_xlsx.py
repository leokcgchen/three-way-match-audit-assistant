"""按系统测试结论生成 GOSPD 风格 xlsx 底稿。

风格对齐 templates/GOSPD01010_style_ref.xlsx：
- 标题/小节/标签：等线 + 深蓝底 #000099 + 白字
- 输入/取值格：浅蓝底（对应样表 theme4+tint≈0.6）
- 分组标题与表头：等线 11，无底色、无黄/浅蓝表头
- 几乎不加边框（与样表一致）
字段与行数仍以本系统测试结论为准。
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# 与样表一致
NAVY = "000099"
LIGHT_INPUT = "BDD7EE"  # 近似样表 theme=4 tint≈0.6

NAVY_FILL = PatternFill("solid", fgColor=NAVY)
INPUT_FILL = PatternFill("solid", fgColor=LIGHT_INPUT)

FONT_TITLE = Font(name="等线", size=22, bold=True, color="FFFFFF")
FONT_SUB = Font(name="等线", size=11, bold=True, color="FFFFFF")
FONT_NAVY_LABEL = Font(name="等线", size=11, bold=True, color="FFFFFF")
FONT_SECTION = Font(name="等线", size=11, bold=True, color="FFFFFF")
FONT_BODY = Font(name="等线", size=11, color="000000")
FONT_BODY_BOLD = Font(name="等线", size=11, bold=True, color="000000")
FONT_INPUT = Font(name="等线", size=11, color="000000")
FONT_CONCLUSION = Font(name="等线", size=12, bold=True, color="000000")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _s(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if abs(value - round(value)) < 1e-9:
            return str(int(round(value)))
        return f"{value:.4g}"
    return str(value)


def _navy_bar(ws, row: int, col: int, text: str, *, merge_to: int, font=FONT_SECTION) -> None:
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_to)
    cell = ws.cell(row, col, text)
    cell.font = font
    cell.fill = NAVY_FILL
    cell.alignment = ALIGN_LEFT


def _label_value(
    ws,
    row: int,
    label_col: int,
    label: str,
    value_col: int,
    value: Any,
) -> None:
    lc = ws.cell(row, label_col, label)
    lc.font = FONT_NAVY_LABEL
    lc.fill = NAVY_FILL
    lc.alignment = ALIGN_CENTER
    vc = ws.cell(row, value_col, _s(value))
    vc.font = FONT_INPUT
    vc.fill = INPUT_FILL
    vc.alignment = ALIGN_CENTER


def _write_title_block(
    ws,
    *,
    title: str,
    subtitle: str,
    meta: Optional[Dict[str, str]] = None,
) -> int:
    """GOSPD 式标题区：大标题 + 副标题（深蓝条）+ 标签/浅蓝取值。"""
    ws.row_dimensions[1].height = 1.5
    ws.row_dimensions[2].height = 48
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=10)
    c2 = ws.cell(2, 2, title)
    c2.font = FONT_TITLE
    c2.fill = NAVY_FILL
    c2.alignment = ALIGN_LEFT

    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=10)
    c3 = ws.cell(3, 2, subtitle)
    c3.font = FONT_SUB
    c3.fill = NAVY_FILL
    c3.alignment = ALIGN_LEFT

    ws.row_dimensions[5].height = 28
    meta = meta or {}
    items = list(meta.items())
    # 样表：标签 | 值 | 标签 | 值 …
    col = 2
    for k, v in items[:4]:
        _label_value(ws, 5, col, k, col + 1, v)
        col += 2
    return 7


def _status_cn(status: Any) -> str:
    s = str(status or "").upper()
    return {
        "PASS": "通过（PASS）",
        "WARNING": "需关注（WARNING）",
        "FAIL": "未通过（FAIL）",
        "SKIPPED": "未执行（SKIPPED）",
    }.get(s, str(status or "-"))


def _write_readable_block(ws, row: int, title: str, paragraphs: Sequence[str]) -> int:
    _navy_bar(ws, row, 2, title, merge_to=8)
    row += 1
    for p in paragraphs:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        cell = ws.cell(row, 2, p)
        cell.font = FONT_CONCLUSION if row and title.endswith("结论") else FONT_BODY
        cell.alignment = ALIGN_LEFT
        ws.row_dimensions[row].height = max(28, 16 + 14 * (str(p).count("\n") + 1))
        row += 1
    return row + 1


def _autosize(ws, min_w: float = 14.0, max_w: float = 42.0) -> None:
    widths: Dict[int, float] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            text = str(cell.value)
            line_len = max((len(x) for x in text.splitlines()), default=0)
            widths[cell.column] = max(
                widths.get(cell.column, min_w), min(max_w, line_len * 1.05 + 2)
            )
    for col, w in widths.items():
        # 描述/结论列加宽
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = max(w, 18 if col >= 4 else w)
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 18


def _write_group_headers(
    ws,
    row: int,
    groups: Sequence[tuple[str, int, int]],
) -> int:
    """分组行（样表 C19/G19/I19…）：无底色、居中。"""
    for text, start, end in groups:
        if end > start:
            ws.merge_cells(start_row=row, start_column=start, end_row=row, end_column=end)
        cell = ws.cell(row, start, text)
        cell.font = FONT_BODY_BOLD
        cell.alignment = ALIGN_CENTER
    return row + 1


def _write_table(
    ws,
    start_row: int,
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
    *,
    section_title: str = "",
    group_headers: Optional[Sequence[tuple[str, int, int]]] = None,
) -> int:
    row = start_row
    if section_title:
        _navy_bar(ws, row, 2, section_title, merge_to=min(2 + max(len(headers) - 1, 1), 12))
        row += 1

    if group_headers:
        row = _write_group_headers(ws, row, group_headers)

    # 列头：无底色（对齐样表 row20）
    ws.row_dimensions[row].height = 36
    for i, h in enumerate(headers, start=2):
        cell = ws.cell(row, i, h)
        cell.font = FONT_BODY
        cell.alignment = ALIGN_CENTER
    row += 1

    for data in rows:
        for i, val in enumerate(data, start=2):
            cell = ws.cell(row, i, _s(val))
            cell.font = FONT_BODY
            cell.alignment = ALIGN_LEFT
        row += 1
    return row + 1


def _write_notes(ws, row: int, lines: Sequence[str]) -> None:
    _navy_bar(ws, row, 2, "注释：", merge_to=6, font=FONT_SECTION)
    for i, line in enumerate(lines):
        cell = ws.cell(row + 1 + i, 2, line)
        cell.font = FONT_BODY
        cell.alignment = ALIGN_LEFT


def _sheet_summary(wb: Workbook, payload: Dict[str, Any]) -> None:
    ws = wb.active
    ws.title = "汇总"
    generated = payload.get("generated_at") or datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    next_row = _write_title_block(
        ws,
        title="抽凭—合同合规性审阅底稿（系统测试结论）",
        subtitle="（风格对齐 GOSPD01010；内容以本系统已执行测试为准。由菜单手动生成，供审计人员阅读复核。）",
        meta={
            "生成时间": generated,
            "程序索引号": "AUDIT-AGENT",
            "币种": "人民币",
            "单位": "Yuan 元",
        },
    )

    # 人工可读总览
    lines = []
    for key, label in (
        ("evidence", "证据匹配"),
        ("amount", "金额准确性"),
        ("contract", "合同条款清晰性"),
        ("three_way", "三单匹配+截止性"),
    ):
        block = payload.get(key) or {}
        if not block.get("ran"):
            lines.append(f"· {label}：本次未运行。")
            continue
        summary = (block.get("summary") or "").strip() or "（无文字摘要）"
        lines.append(f"· {label}：{_status_cn(block.get('status'))}。{summary}")
    next_row = _write_readable_block(ws, next_row, "一、人工可读结论（请先看本段）", lines)

    _navy_bar(ws, next_row, 2, "二、测试执行一览表", merge_to=6)
    next_row += 1
    headers = ["测试模块", "是否已跑", "状态（可读）", "摘要说明"]
    rows = []
    for key, label in (
        ("evidence", "证据匹配"),
        ("amount", "金额准确性"),
        ("contract", "合同条款清晰性"),
        ("three_way", "三单匹配+截止性"),
    ):
        block = payload.get(key) or {}
        ran = bool(block.get("ran"))
        rows.append(
            [
                label,
                "是" if ran else "否",
                _status_cn(block.get("status")) if ran else "-",
                block.get("summary") or "",
            ]
        )
    next_row = _write_table(ws, next_row, headers, rows)
    _write_notes(
        ws,
        next_row,
        [
            "— 仅汇总本次会话中已点击运行的测试；未运行模块不编造结论。",
            "— 差异金额/天数/问题码以规则引擎终态为准；AI 解读（如有）仅作旁路说明。",
            "— 阅读建议：先看「人工可读结论」，再按需打开各明细表核对证据。",
        ],
    )
    _autosize(ws)


def _sheet_evidence(wb: Workbook, evidence: Dict[str, Any]) -> None:
    ws = wb.create_sheet("证据匹配")
    next_row = _write_title_block(
        ws,
        title="证据匹配底稿",
        subtitle="串联合同/订单/发货/签收/发票/回款与序时账编号。",
        meta={
            "测试状态": _s(evidence.get("status")),
            "业务编号": _s(
                evidence.get("business_id") or evidence.get("ledger_matched_biz_id")
            ),
        },
    )
    _navy_bar(ws, next_row, 2, "评量询问 / 证据链", merge_to=6)
    next_row += 1
    headers = ["角色", "文件", "业务编号", "已串联", "说明"]
    rows = []
    for node in evidence.get("nodes") or []:
        if node.get("role") == "other":
            continue
        rows.append(
            [
                node.get("role"),
                node.get("file_name"),
                ", ".join(node.get("biz_keys") or []),
                "是" if node.get("linked") else "否",
                node.get("note") or node.get("reason") or "",
            ]
        )
    if not rows:
        rows = [["-", "-", "-", "-", evidence.get("human_readable_summary") or "无节点"]]
    next_row = _write_table(
        ws,
        next_row,
        headers,
        rows,
        group_headers=[("基于相关文件", 2, 6)],
    )
    _write_notes(
        ws,
        next_row,
        [evidence.get("human_readable_summary") or "— 见上表节点串联结果。"],
    )
    _autosize(ws)


def _sheet_amount(wb: Workbook, amount: Dict[str, Any]) -> None:
    ws = wb.create_sheet("金额准确性")
    report = amount.get("accuracy_report") or amount.get("report") or {}
    detail = report.get("amount_test") or report.get("test_result") or {}
    src = report.get("source_values") or {}
    ledger = report.get("ledger_values") or {}
    recalc = report.get("recalculation") or {}
    fill = report.get("workpaper_fill") or {}
    status = (
        amount.get("status")
        or detail.get("test_status")
        or report.get("test_status")
        or ""
    )
    next_row = _write_title_block(
        ws,
        title="金额准确性测试底稿",
        subtitle="签收数量×未税单价×折扣×税率重算，与序时账比对；容差以系统规则为准。",
        meta={
            "测试状态": _status_cn(status),
            "业务编号": _s(report.get("business_id") or amount.get("business_id")),
            "异常类型": _s(detail.get("issue_type") or fill.get("异常类型")),
            "币种": "人民币",
        },
    )
    next_row = _write_readable_block(
        ws,
        next_row,
        "审计可读结论",
        [
            f"测试状态：{_status_cn(status)}。",
            str(
                detail.get("issue_description")
                or amount.get("human_readable_summary")
                or fill.get("审计结论")
                or "（无问题描述）"
            ),
            f"建议审计结论：{_s(fill.get('审计结论') or status)}。"
            + (
                f" 差异金额 {_s(detail.get('difference_amount'))} 元。"
                if detail.get("difference_amount") is not None
                else ""
            ),
        ],
    )
    _navy_bar(ws, next_row, 2, "金额测试明细（系统口径）", merge_to=8)
    next_row += 1
    headers = [
        "样本/字段",
        "采用值",
        "账面/对照",
        "差异",
        "方向",
        "问题描述",
        "审计结论",
    ]
    diff = detail.get("difference_amount")
    rows: List[List[Any]] = [
        [
            report.get("business_id") or amount.get("business_id") or "样本1",
            f"数量={src.get('quantity')}; 单价未税={src.get('unit_price_excl_tax')}; "
            f"折扣={src.get('discount_rate')}; 税率={src.get('vat_rate')}; "
            f"重算价税合计={recalc.get('gross_amount_incl_tax')}",
            ledger.get("ledger_debit_total")
            if ledger.get("ledger_debit_total") is not None
            else ledger.get("ledger_ar_debit"),
            diff,
            detail.get("direction"),
            detail.get("issue_description") or amount.get("human_readable_summary"),
            fill.get("审计结论") or status,
        ]
    ]
    for c in amount.get("checks") or []:
        rows.append(
            [
                f"{c.get('role')}:{c.get('file_name')}",
                c.get("recalculated_amount"),
                c.get("book_amount"),
                c.get("deviation_amount"),
                "",
                c.get("issue"),
                c.get("status"),
            ]
        )
    next_row = _write_table(
        ws,
        next_row,
        headers,
        rows,
        group_headers=[
            ("基于收入明细账", 2, 4),
            ("差异", 5, 6),
            ("测试结论", 7, 8),
        ],
    )

    _navy_bar(ws, next_row, 2, "计价要素与重算", merge_to=4)
    next_row += 1
    pricing_headers = ["要素", "值", "来源"]
    pricing_rows = [
        ["数量", src.get("quantity"), src.get("quantity_source")],
        ["未税单价", src.get("unit_price_excl_tax"), src.get("price_source")],
        ["折扣率", src.get("discount_rate"), "合同/订单"],
        ["税率", src.get("vat_rate"), "发票/合同"],
        ["重算不含税", recalc.get("net_amount_excl_tax"), recalc.get("formula")],
        ["重算税额", recalc.get("vat_amount"), ""],
        ["重算价税合计", recalc.get("gross_amount_incl_tax"), ""],
    ]
    next_row = _write_table(ws, next_row, pricing_headers, pricing_rows)
    _write_notes(
        ws,
        next_row,
        ["— 账面金额口径与重算公式以系统规则为准；不得用付款账期推算收入。"],
    )
    _autosize(ws)


def _sheet_contract(wb: Workbook, contract: Dict[str, Any]) -> None:
    ws = wb.create_sheet("合同条款")
    report = contract.get("clarity_report") or contract.get("report") or {}
    tr = report.get("test_result") or {}
    extracted = report.get("extracted") or contract.get("extracted") or {}
    status = contract.get("status") or tr.get("test_status") or ""
    next_row = _write_title_block(
        ws,
        title="合同条款清晰性测试底稿",
        subtitle="对价/支付/履约/运输控制权；歧义→WARNING，不因条款不清直接账务FAIL。",
        meta={
            "测试状态": _status_cn(status),
            "合同编号": _s(extracted.get("contract_id") or report.get("contract_id")),
            "客户名称": _s(report.get("customer_name")),
            "程序索引号": "CONTRACT-CLARITY",
        },
    )
    next_row = _write_readable_block(
        ws,
        next_row,
        "审计可读结论",
        [
            f"测试状态：{_status_cn(status)}。",
            str(
                tr.get("issue_description")
                or contract.get("human_readable_summary")
                or "未发现问题或未形成描述。"
            ),
            "说明：合同歧义通常建议人工复核（WARNING），不等于已认定账务错报。",
        ],
    )
    _navy_bar(ws, next_row, 2, "条款问题明细", merge_to=6)
    next_row += 1
    headers = ["问题码", "维度", "描述", "原文摘录", "是否需人工复核"]
    rows: List[List[Any]] = []
    for it in tr.get("issues") or []:
        rows.append(
            [
                it.get("issue_code"),
                it.get("dimension"),
                it.get("description"),
                it.get("excerpt"),
                "是" if tr.get("manual_review_required") else "否",
            ]
        )
    if not rows:
        rows.append(
            [
                tr.get("issue_code") or "NONE",
                tr.get("test_dimension") or "无",
                tr.get("issue_description")
                or contract.get("human_readable_summary")
                or "未发现问题",
                "",
                "是" if tr.get("manual_review_required") else "否",
            ]
        )
    next_row = _write_table(
        ws,
        next_row,
        headers,
        rows,
        group_headers=[("基于合同", 2, 6)],
    )
    _write_notes(
        ws,
        next_row,
        ["— 合同歧义通常建议 WARNING 与人工复核，不等于账务错报 FAIL。"],
    )
    _autosize(ws)


def _sheet_three_way(wb: Workbook, three_way: Dict[str, Any]) -> None:
    ws = wb.create_sheet("三单与截止")
    match = three_way.get("match_result") or {}
    if hasattr(match, "model_dump"):
        match = match.model_dump()
    cutoff = three_way.get("cutoff_result") or {}
    if hasattr(cutoff, "model_dump"):
        cutoff = cutoff.model_dump()
    req = three_way.get("match_request") or {}
    if hasattr(req, "model_dump"):
        req = req.model_dump()
    order = req.get("order") or {}
    receipt = req.get("warehouse_receipt") or {}
    invoice = req.get("invoice") or {}

    next_row = _write_title_block(
        ws,
        title="三单匹配与收入截止性底稿",
        subtitle="三单比对应与截止性（应确认日=控制权转移/验收日，不含付款账期）。",
        meta={
            "整体状态": _s(three_way.get("overall_status") or match.get("overall_status")),
            "三单状态": _s(match.get("overall_status")),
            "截止状态": _s(
                cutoff.get("测试状态")
                if cutoff
                else three_way.get("cutoff_skipped_reason") or "未执行/跳过"
            ),
            "程序索引号": "GOSPD01010.1",
        },
    )

    _navy_bar(ws, next_row, 2, "三单匹配", merge_to=10)
    next_row += 1
    tw_headers = [
        "订单号",
        "供应商",
        "订单金额(万元)",
        "入库金额(万元)",
        "发票金额(万元)",
        "签收/控制权日",
        "入账日",
        "三单决策",
        "三单状态",
        "风险说明",
    ]
    tw_rows = [
        [
            order.get("order_no"),
            order.get("supplier_name"),
            order.get("total_amount"),
            receipt.get("total_amount"),
            invoice.get("total_amount"),
            receipt.get("receipt_date"),
            invoice.get("posting_date"),
            match.get("decision") or match.get("overall_status"),
            match.get("overall_status"),
            "; ".join(match.get("risks") or [])
            or match.get("summary")
            or three_way.get("human_readable_summary"),
        ]
    ]
    next_row = _write_table(
        ws,
        next_row,
        tw_headers,
        tw_rows,
        group_headers=[
            ("基于合同/订单", 2, 4),
            ("基于相关文件", 5, 8),
            ("差异/结论", 9, 11),
        ],
    )

    _navy_bar(ws, next_row, 2, "截止性测试", merge_to=9)
    next_row += 1
    cut_headers = [
        "应确认日期",
        "入账日期",
        "偏差天数",
        "测试状态",
        "风险等级",
        "问题描述",
        "计算依据",
        "AI解读",
    ]
    if cutoff:
        interp = cutoff.get("LLM解读") or {}
        cut_rows = [
            [
                cutoff.get("应确认日期"),
                (cutoff.get("底稿回填") or {}).get("入账日期")
                or invoice.get("posting_date"),
                cutoff.get("偏差天数"),
                cutoff.get("测试状态"),
                cutoff.get("风险等级"),
                cutoff.get("问题描述"),
                cutoff.get("计算依据"),
                (interp.get("explanation") if isinstance(interp, dict) else "") or "",
            ]
        ]
    else:
        cut_rows = [
            [
                "",
                "",
                "",
                "SKIPPED",
                "",
                three_way.get("cutoff_skipped_reason") or "截止性未执行",
                "",
                "",
            ]
        ]
    next_row = _write_table(ws, next_row, cut_headers, cut_rows)
    _write_notes(
        ws,
        next_row,
        [
            "— 应确认日 = 控制权转移/验收完成日；付款账期不参与截止公式。",
            "— 样本量超过表行数时，请复制粘贴扩展行，勿破坏标题区格式。",
        ],
    )
    _autosize(ws)


def build_audit_workbook_payload(
    *,
    evidence: Optional[Dict[str, Any]] = None,
    amount: Optional[Dict[str, Any]] = None,
    contract: Optional[Dict[str, Any]] = None,
    three_way: Optional[Dict[str, Any]] = None,
    coverage: Optional[Dict[str, Any]] = None,
    relations: Optional[Any] = None,
    duplicates: Optional[Dict[str, Any]] = None,
    advisory_candidates: Optional[Any] = None,
    matching_confirmed: bool = False,
    conclusion_confirmed: bool = False,
) -> Dict[str, Any]:
    def _wrap(raw: Optional[Dict[str, Any]], status_keys: Iterable[str]) -> Dict[str, Any]:
        if not raw:
            return {"ran": False, "status": "", "summary": "", "data": {}}
        status = ""
        for k in status_keys:
            cur: Any = raw
            for part in k.split("."):
                if isinstance(cur, dict):
                    cur = cur.get(part)
                else:
                    cur = None
                    break
            if cur:
                status = str(cur)
                break
        summary = str(
            raw.get("human_readable_summary")
            or (raw.get("match_result") or {}).get("human_readable_summary")
            or ""
        )
        return {"ran": True, "status": status, "summary": summary, "data": raw}

    rel_list = list(relations or [])
    adv_list = [x for x in (advisory_candidates or []) if isinstance(x, dict)]
    return {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "evidence": _wrap(evidence, ("status",)),
        "amount": _wrap(amount, ("status", "accuracy_report.amount_test.test_status")),
        "contract": _wrap(contract, ("status", "clarity_report.test_result.test_status")),
        "three_way": _wrap(
            three_way,
            ("overall_status", "match_result.overall_status", "cutoff_result.测试状态"),
        ),
        "coverage_map": coverage or {},
        "relations": {"ran": bool(rel_list), "data": rel_list},
        "duplicates": {
            "ran": bool(duplicates and (duplicates.get("ran") or duplicates.get("findings"))),
            "data": duplicates or {},
        },
        "advisory": {"ran": bool(adv_list), "data": adv_list},
        "matching_confirmed": matching_confirmed,
        "conclusion_confirmed": conclusion_confirmed,
    }


def _sheet_relations(wb: Workbook, relations: list) -> None:
    ws = wb.create_sheet("单据关系候选")
    row = _write_title_block(
        ws,
        title="单据关系候选",
        subtitle="PROPOSED / VERIFIED / REJECTED；人工确认后留痕，不改规则终态。",
        meta={"条数": str(len(relations or []))},
    )
    row += 1
    headers = ["关系ID", "从", "到", "类型", "状态", "共享编号", "摘录", "操作人", "更新时间"]
    for col, h in enumerate(headers, start=2):
        cell = ws.cell(row, col, h)
        cell.font = FONT_BODY_BOLD
    row += 1
    for rel in relations or []:
        vals = [
            rel.get("relation_id"),
            rel.get("from_id"),
            rel.get("to_id"),
            rel.get("rel_type"),
            rel.get("status"),
            ", ".join(rel.get("shared_keys") or []),
            rel.get("excerpt"),
            rel.get("actor"),
            rel.get("updated_at"),
        ]
        for col, v in enumerate(vals, start=2):
            cell = ws.cell(row, col, _s(v))
            cell.font = FONT_BODY
            cell.alignment = ALIGN_LEFT
        row += 1
    _autosize(ws)


def _sheet_duplicates(wb: Workbook, duplicates: Dict[str, Any]) -> None:
    ws = wb.create_sheet("重复号多版本")
    findings = duplicates.get("findings") or []
    summary = duplicates.get("summary") or {}
    row = _write_title_block(
        ws,
        title="重复号 / 多版本检测",
        subtitle="仅提示，不自动改 PASS/FAIL。",
        meta={
            "发现数": str(summary.get("total") or len(findings)),
            "强信号": str(summary.get("fail_signals") or 0),
        },
    )
    row += 1
    headers = ["类型", "严重度", "编号", "说明", "文件", "备注"]
    for col, h in enumerate(headers, start=2):
        cell = ws.cell(row, col, h)
        cell.font = FONT_BODY_BOLD
    row += 1
    for f in findings:
        vals = [
            f.get("issue_type"),
            f.get("severity"),
            f.get("biz_id"),
            f.get("title"),
            "；".join(f.get("file_names") or []),
            f.get("note"),
        ]
        for col, v in enumerate(vals, start=2):
            cell = ws.cell(row, col, _s(v))
            cell.font = FONT_BODY
            cell.alignment = ALIGN_LEFT
        row += 1
    if not findings:
        ws.cell(row, 2, "未发现重复票号或同编号多文件。").font = FONT_BODY
    _autosize(ws)


def _sheet_coverage(wb: Workbook, coverage: Dict[str, Any]) -> None:
    ws = wb.create_sheet("规则覆盖地图")
    row = _write_title_block(
        ws,
        title="规则覆盖地图",
        subtitle="标明已检查 / 未检查 / 不适用（不改变测试终态）",
        meta={
            "版本": str(coverage.get("version") or "coverage-map-v0"),
            "生成": datetime.now().strftime("%Y-%m-%d %H:%M"),
        },
    )
    row += 1
    headers = ["维度", "状态", "CEAVOP", "结果状态", "说明"]
    for col, h in enumerate(headers, start=2):
        cell = ws.cell(row, col, h)
        cell.font = FONT_BODY_BOLD
    row += 1
    for dim in coverage.get("dimensions") or []:
        vals = [
            dim.get("label"),
            dim.get("status"),
            dim.get("ceavop"),
            dim.get("result_status") or "",
            dim.get("note") or "",
        ]
        for col, v in enumerate(vals, start=2):
            cell = ws.cell(row, col, _s(v))
            cell.font = FONT_BODY
            cell.alignment = ALIGN_LEFT
        row += 1
    row += 1
    _navy_bar(ws, row, 2, "PASS 不能证明", merge_to=6)
    row += 1
    for tip in coverage.get("what_pass_does_not_prove") or []:
        ws.cell(row, 2, f"· {_s(tip)}").font = FONT_BODY
        row += 1


def _sheet_advisory(wb: Workbook, candidates: list) -> None:
    ws = wb.create_sheet("顾问候选与旁注")
    row = _write_title_block(
        ws,
        title="顾问候选与底稿旁注溯源",
        subtitle=(
            "LLM/消歧主张仅 advisory；PROPOSED 不得视为审计结论。"
            "正式 Yes/No 仍以规则断言为准；本表供复核与留痕。"
        ),
        meta={"条数": str(len(candidates or []))},
    )
    row += 1
    headers = [
        "候选ID",
        "任务",
        "种类",
        "状态",
        "业务号",
        "摘要",
        "摘录",
        "回查",
        "操作人",
        "更新时间",
    ]
    for col, h in enumerate(headers, start=2):
        cell = ws.cell(row, col, h)
        cell.font = FONT_BODY_BOLD
    row += 1
    for cand in candidates or []:
        payload = cand.get("payload") if isinstance(cand.get("payload"), dict) else {}
        evidence = cand.get("evidence") if isinstance(cand.get("evidence"), dict) else {}
        verify = cand.get("verify") if isinstance(cand.get("verify"), dict) else {}
        summary = (
            payload.get("issue_code")
            or payload.get("field_name")
            or payload.get("disposition")
            or payload.get("normalized_candidate")
            or ""
        )
        vals = [
            cand.get("candidate_id"),
            cand.get("task_type"),
            cand.get("kind"),
            cand.get("status"),
            cand.get("business_id"),
            summary,
            evidence.get("excerpt") or payload.get("excerpt"),
            verify.get("reason") if verify.get("passed") is False else ("ok" if verify.get("passed") else ""),
            cand.get("actor"),
            cand.get("updated_at"),
        ]
        for col, v in enumerate(vals, start=2):
            cell = ws.cell(row, col, _s(v))
            cell.font = FONT_BODY
            cell.alignment = ALIGN_LEFT
        row += 1
    _autosize(ws)


def generate_audit_workbook_xlsx(
    payload: Dict[str, Any],
    output_path: Path,
) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    _sheet_summary(wb, payload)

    ev = payload.get("evidence") or {}
    if ev.get("ran"):
        _sheet_evidence(wb, ev.get("data") or {})

    am = payload.get("amount") or {}
    if am.get("ran"):
        _sheet_amount(wb, am.get("data") or {})

    ct = payload.get("contract") or {}
    if ct.get("ran"):
        _sheet_contract(wb, ct.get("data") or {})

    tw = payload.get("three_way") or {}
    if tw.get("ran"):
        _sheet_three_way(wb, tw.get("data") or {})

    rel = payload.get("relations") or {}
    if rel.get("ran"):
        _sheet_relations(wb, rel.get("data") or [])

    adv = payload.get("advisory") or {}
    if adv.get("ran"):
        _sheet_advisory(wb, adv.get("data") or [])

    dup = payload.get("duplicates") or {}
    if dup.get("ran"):
        _sheet_duplicates(wb, dup.get("data") or {})

    coverage = payload.get("coverage_map") or {}
    if coverage.get("dimensions"):
        _sheet_coverage(wb, coverage)

    wb.save(output_path)
    return output_path
