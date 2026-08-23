"""按官方 GOSPD01010.4 模板填表：交易价格分摊（SSP）抽凭。

列映射（表头行 16，示例 19–22，数据起始行 23，样本 1–19）：
B样本 C日期 D凭证 E客户 F已入账金额 G总销售金额 H折扣/可变对价
I履约义务 J单独售价 K相关文件类型 L相关文件编号 M合同索引 N主要条款
O折扣分摊标准 P其他文件是否适用 Q类型 R索引
S分摊基础 T重算收入 U差异(公式) V交易价格是否适当 W步骤无异常 X异常 Y注释
保留「底稿须知」工作表；U 列保留模板公式。
"""

from __future__ import annotations

import shutil
from copy import copy
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.audit.gospd01010_4_assertions import build_gospd01010_4_assertions
from src.reporting.gospd01010_filler import group_classified_by_chain

ROOT = Path(__file__).resolve().parents[2]
TEMPLATE_CANDIDATES = (
    ROOT / "templates" / "GOSPD01010.4.xlsx",
    ROOT / "templates" / "GOSPD01010_4.xlsx",
)

DATA_START_ROW = 23
MAX_SAMPLE_ROWS = 19

COL = {
    "sample_no": 2,  # B
    "date": 3,  # C
    "voucher": 4,  # D
    "customer": 5,  # E
    "amt_booked": 6,  # F
    "amt_sales": 7,  # G
    "discount": 8,  # H
    "po_name": 9,  # I
    "ssp": 10,  # J
    "ssp_file_type": 11,  # K
    "ssp_file_no": 12,  # L
    "contract_idx": 13,  # M
    "contract_terms": 14,  # N
    "criteria": 15,  # O
    "other_applicable": 16,  # P
    "other_file_type": 17,  # Q
    "other_file_index": 18,  # R
    "alloc_basis": 19,  # S
    "recalc_rev": 20,  # T
    "diff": 21,  # U
    "price_ok": 22,  # V
    "all_ok": 23,  # W
    "exception": 24,  # X
    "comment": 25,  # Y
}


def resolve_template_path() -> Path:
    for p in TEMPLATE_CANDIDATES:
        if p.is_file():
            return p
    raise FileNotFoundError(
        "未找到 GOSPD01010.4 模板，请将文件放到 templates/GOSPD01010.4.xlsx"
    )


def _f(doc: Optional[dict[str, Any]], *keys: str) -> Any:
    if not doc:
        return None
    from src.models.field_values import rule_readable_fields

    fields = rule_readable_fields(doc)
    for k in keys:
        if fields.get(k) not in (None, ""):
            return fields.get(k)
        if doc.get(k) not in (None, ""):
            return doc.get(k)
    return None


def _num(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(
            str(v)
            .replace(",", "")
            .replace("，", "")
            .replace("¥", "")
            .replace("元", "")
            .replace("CNY", "")
            .strip()
        )
    except ValueError:
        return None


def _by_type(docs: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    out: dict[str, dict[str, Any]] = {}
    for d in docs or []:
        t = str(d.get("doc_type") or "")
        if t and t not in out:
            out[t] = d
    return out


def _terms_text(contract: Optional[dict], order: Optional[dict]) -> str:
    bits: list[str] = []
    for key, label in (
        ("paymentTerms", "付款"),
        ("controlTransferTerms", "控制权"),
        ("performanceObligations", "履约义务"),
    ):
        v = _f(contract, key) or _f(order, key)
        if v:
            bits.append(f"{label}:{v}")
    return "；".join(bits) or str(_f(contract, "remarks") or "")


def _row_from_chain(
    *,
    sample_no: int,
    chain_id: str,
    docs: list[dict[str, Any]],
    job: dict[str, Any],
    apply_job_tests: bool,
) -> dict[str, Any]:
    by = _by_type(docs)
    contract = by.get("contract")
    order = by.get("order")
    invoice = by.get("invoice")

    assertions = build_gospd01010_4_assertions(
        docs=docs,
        job=job,
        chain_id=chain_id,
        apply_job_tests=apply_job_tests,
    )

    date = (
        _f(invoice, "postingDate", "documentDate")
        or _f(order, "documentDate")
        or (invoice or {}).get("ledger_posting_date")
        or ""
    )
    voucher = (
        (invoice or {}).get("ledger_voucher")
        or _f(invoice, "voucherNo", "documentNo")
        or ""
    )
    customer = (
        _f(invoice, "buyerName", "customerName")
        or _f(order, "buyerName", "customerName")
        or _f(contract, "buyerName", "customerName")
        or ""
    )
    contract_no = (
        _f(contract, "contractNo", "documentNo")
        or _f(order, "contractNo")
        or (chain_id if "HT" in str(chain_id).upper() else "")
    )

    amt_sales = None
    for doc in (contract, order, invoice):
        n = _num(_f(doc, "totalAmount", "amount", "grossAmount"))
        if n is not None and n > 0:
            amt_sales = n
            break
    amt_booked = amt_sales
    if isinstance(job.get("amount_test"), dict):
        ar = job["amount_test"].get("accuracy_report") or {}
        if isinstance(ar, dict):
            led = _num((ar.get("ledger_values") or {}).get("ledger_debit_total"))
            if led is not None:
                amt_booked = led

    discount = _num(_f(order, "discountAmount", "discount")) or _num(
        _f(contract, "discountAmount", "discount")
    )
    po_name = (
        _f(contract, "performanceObligations")
        or _f(order, "productName", "itemName")
        or "销售商品"
    )
    # 单一履约义务：SSP 取合同/订单总价
    ssp = amt_sales
    ssp_type = ""
    ssp_no = ""
    if order:
        ssp_type = "销售订单（单独售价参考）"
        ssp_no = str(_f(order, "orderNo", "documentNo") or order.get("file_name") or "")
    elif invoice:
        ssp_type = "销售发票（单独售价参考）"
        ssp_no = str(_f(invoice, "invoiceNo", "documentNo") or "")

    # 单一义务：分摊基础=1；重算收入≈总价+折扣（折扣为负时同模板）
    alloc = 1.0
    disc = float(discount or 0.0)
    sales = float(amt_sales or 0.0)
    recalc = sales * alloc + disc

    other = assertions.get("other_files") or {}
    exc = str(assertions.get("exception") or "").strip()
    appendix = str(assertions.get("exception_appendix") or "").strip()
    if appendix and appendix not in exc:
        from src.audit.workpaper_notes import merge_exception_text

        exc = merge_exception_text(exc, appendix)

    return {
        "sample_no": sample_no,
        "chain_id": chain_id,
        "date": str(date or ""),
        "voucher": voucher,
        "customer": customer,
        "amt_booked": amt_booked,
        "amt_sales": amt_sales,
        "discount": discount if discount is not None else None,
        "po_name": str(po_name)[:80],
        "ssp": ssp,
        "ssp_file_type": ssp_type,
        "ssp_file_no": ssp_no,
        "contract_idx": contract_no,
        "contract_terms": _terms_text(contract, order),
        "criteria": assertions.get("criteria_label") or "",
        "other_applicable": assertions.get("other_applicable") or "",
        "other_file_type": assertions.get("other_file_type")
        or other.get("file_type")
        or "",
        "other_file_index": assertions.get("other_file_index")
        or other.get("file_index")
        or "",
        "alloc_basis": alloc,
        "recalc_rev": round(recalc, 2) if amt_sales is not None else None,
        "price_ok": assertions.get("price_ok_label") or "",
        "all_ok": assertions.get("all_ok_label") or "",
        "exception": exc,
        "comment": assertions.get("comment") or "",
    }


def build_gospd01010_4_sample_rows(job: dict[str, Any]) -> list[dict[str, Any]]:
    classified = list(job.get("classified") or [])
    chains = group_classified_by_chain(classified)
    if not chains:
        return []
    samples = (
        job.get("gospd_sample_results")
        if isinstance(job.get("gospd_sample_results"), dict)
        else {}
    )
    single = len(chains) == 1
    rows: list[dict[str, Any]] = []
    for i, (chain_id, docs) in enumerate(chains, start=1):
        sample = samples.get(chain_id) or {}
        has_per = bool(sample.get("contract_terms") or sample.get("amount_test"))
        rows.append(
            _row_from_chain(
                sample_no=i,
                chain_id=chain_id,
                docs=docs,
                job=job,
                apply_job_tests=has_per or single,
            )
        )
    return rows


def _copy_row_style(ws: Worksheet, src_row: int, dst_row: int) -> None:
    for c in range(1, 26):
        sc = ws.cell(src_row, c)
        dc = ws.cell(dst_row, c)
        if sc.has_style:
            dc._style = copy(sc._style)
        dc.number_format = sc.number_format


def fill_gospd01010_4_workbook(
    job: dict[str, Any],
    output_path: Path,
    *,
    entity_name: str = "",
    currency: str = "人民币",
    unit: str = "Yuan 元",
) -> Path:
    template = resolve_template_path()
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template, output_path)
    try:
        output_path.chmod(0o666)
    except OSError:
        pass

    wb = load_workbook(output_path)
    ws = wb[wb.sheetnames[0]]

    ws.cell(5, 3, entity_name or ws.cell(5, 3).value or "")
    ws.cell(5, 5, "GOSPD01010.4")
    ws.cell(5, 7, currency)
    ws.cell(5, 9, unit)
    # 模板误写 01010.3 步骤标题时纠正
    if "01010.3" in str(ws.cell(7, 2).value or ""):
        ws.cell(7, 2, "GOSPD01010.4测试步骤（请展开）：")

    rows = build_gospd01010_4_sample_rows(job)
    for i, row in enumerate(rows):
        r = DATA_START_ROW + i
        if i >= MAX_SAMPLE_ROWS or (
            r > DATA_START_ROW and ws.cell(r, COL["sample_no"]).value in (None, "")
        ):
            _copy_row_style(ws, DATA_START_ROW, r)

        ws.cell(r, COL["sample_no"], row.get("sample_no") or (i + 1))
        ws.cell(r, COL["date"], row.get("date") or "")
        ws.cell(r, COL["voucher"], row.get("voucher") or "")
        ws.cell(r, COL["customer"], row.get("customer") or "")
        if row.get("amt_booked") is not None:
            ws.cell(r, COL["amt_booked"], row["amt_booked"])
        if row.get("amt_sales") is not None:
            ws.cell(r, COL["amt_sales"], row["amt_sales"])
        if row.get("discount") is not None:
            ws.cell(r, COL["discount"], row["discount"])
        ws.cell(r, COL["po_name"], row.get("po_name") or "")
        if row.get("ssp") is not None:
            ws.cell(r, COL["ssp"], row["ssp"])
        ws.cell(r, COL["ssp_file_type"], row.get("ssp_file_type") or "")
        ws.cell(r, COL["ssp_file_no"], row.get("ssp_file_no") or "")
        ws.cell(r, COL["contract_idx"], row.get("contract_idx") or "")
        ws.cell(r, COL["contract_terms"], row.get("contract_terms") or "")
        ws.cell(r, COL["criteria"], row.get("criteria") or "")
        ws.cell(r, COL["other_applicable"], row.get("other_applicable") or "")
        ws.cell(r, COL["other_file_type"], row.get("other_file_type") or "")
        ws.cell(r, COL["other_file_index"], row.get("other_file_index") or "")
        if row.get("alloc_basis") is not None:
            ws.cell(r, COL["alloc_basis"], row["alloc_basis"])
        if row.get("recalc_rev") is not None:
            ws.cell(r, COL["recalc_rev"], row["recalc_rev"])
        # 差异列保留/重写为模板公式
        ws.cell(r, COL["diff"], f"=F{r}-T{r}")
        ws.cell(r, COL["price_ok"], row.get("price_ok") or "")
        ws.cell(r, COL["all_ok"], row.get("all_ok") or "")
        ws.cell(r, COL["exception"], row.get("exception") or "")
        ws.cell(r, COL["comment"], row.get("comment") or "")

    note_row = max(48, DATA_START_ROW + max(len(rows), MAX_SAMPLE_ROWS) + 3)
    goals = (job.get("plan") or {}).get("goal_ids") or job.get("goal_ids") or []
    ws.cell(
        note_row,
        2,
        "— 本底稿由工作台按目标 gospd01010_4 自动回填；"
        f"已选目标={','.join(map(str, goals)) or 'gospd01010_4'}；"
        f"样本链={','.join(str(r.get('chain_id') or '') for r in rows)}。"
        "枚举对齐「底稿须知」；缺管理层分摊底稿时按单一履约义务处理，详见 Y 列注释。"
        "U 列差异保留公式 F-T。",
    )
    wb.save(output_path)
    return output_path
