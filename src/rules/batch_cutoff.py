"""批量截止性测试：按合同编号匹配序时账与签收单。"""

from __future__ import annotations

from typing import List, Optional

import pandas as pd

from src.rules.cutoff_checker import CutoffChecker
from src.utils.audit_utils import serialize_calculation_trail
from src.utils.date_extractor import extract_contract_id_from_row
from src.utils.date_extractor import extract_date_from_text as extract_date_text


def batch_cutoff_check(
    ledger_df: pd.DataFrame,
    receipt_df: pd.DataFrame,
    payment_days: int,
    match_key: str = "合同编号",
    extract_date_from_text: bool = False,
    receipt_date_column: str = "签收日期",
    extract_contract_from_text: bool = False,
    contract_text_columns: Optional[List[str]] = None,
) -> pd.DataFrame:
    """
    按 match_key 左连接序时账与签收单，批量执行截止性测试。

    要求输入列（标准化后）：
    - ledger_df: 至少包含 match_key, entry_date
    - receipt_df: 至少包含 match_key, 以及签收日期列

    可选：
    - extract_contract_from_text: 从摘要/备注等文本列提取合同编号后再匹配
    - contract_text_columns: 要扫描的文本列名列表
    """
    if match_key not in ledger_df.columns and not extract_contract_from_text:
        raise ValueError(f"序时账缺少匹配列: {match_key}")

    checker = CutoffChecker()
    ledger = ledger_df.copy()
    receipt = receipt_df.copy() if receipt_df is not None else pd.DataFrame()

    # 合同编号文本提取（匹配前）
    if extract_contract_from_text:
        ledger = _apply_contract_id_extraction(
            ledger, match_key, contract_text_columns or []
        )
        if not receipt.empty:
            # 签收单若也带有相同文本列，同步提取；否则保留原 match_key
            receipt = _apply_contract_id_extraction(
                receipt, match_key, contract_text_columns or [], optional=True
            )

    if match_key not in ledger.columns:
        raise ValueError(f"序时账缺少匹配列: {match_key}")

    ledger[match_key] = _normalize_match_key(ledger[match_key])
    # 提取失败的行保留，用空字符串占位，后面单独标记
    failed_contract_rows = ledger[
        ledger[match_key].eq("") | ledger[match_key].eq("nan") | ledger[match_key].isna()
    ].copy()
    ledger = ledger[ledger[match_key].ne("") & ledger[match_key].ne("nan")]

    if receipt.empty or match_key not in receipt.columns:
        receipt_std = pd.DataFrame(
            columns=[match_key, "receipt_date", "date_extract_status"]
        )
    else:
        receipt_std = receipt.copy()
        receipt_std[match_key] = _normalize_match_key(receipt_std[match_key])
        receipt_std = receipt_std[
            receipt_std[match_key].ne("") & receipt_std[match_key].ne("nan")
        ]
        receipt_std = _prepare_receipt_dates(
            receipt_std,
            extract_date_from_text=extract_date_from_text,
            receipt_date_column=receipt_date_column,
        )
        receipt_std = receipt_std.drop_duplicates(subset=[match_key], keep="first")

    if "entry_date" not in ledger.columns and not ledger.empty:
        raise ValueError("序时账缺少 entry_date 列（请先完成列映射标准化）")
    if ledger.empty and failed_contract_rows.empty:
        raise ValueError("序时账无可用记录")

    results = []
    if not ledger.empty:
        if "entry_date" not in ledger.columns:
            raise ValueError("序时账缺少 entry_date 列（请先完成列映射标准化）")
        merged = ledger.merge(
            receipt_std,
            on=match_key,
            how="left",
            suffixes=("", "_receipt"),
        )
        ledger_keys = set(ledger[match_key].astype(str).str.strip())
        receipt_keys = (
            set(receipt_std[match_key].astype(str).str.strip())
            if not receipt_std.empty and match_key in receipt_std.columns
            else set()
        )
        for _, row in merged.iterrows():
            results.append(
                _build_cutoff_record(
                    row=row,
                    match_key=match_key,
                    payment_days=payment_days,
                    checker=checker,
                    receipt_keys=receipt_keys,
                )
            )
    else:
        ledger_keys = set()

    # 合同编号提取失败的序时账行
    for _, row in failed_contract_rows.iterrows():
        record = row.to_dict()
        record.update(
            {
                match_key: None,
                "expected_revenue_date": None,
                "deviation_days": None,
                "cutoff_status": "NO_RECEIPT",
                "issue_description": "合同编号文本提取失败，无法匹配签收单",
                "calculation_basis": "extract_contract_from_text 未提取到合同编号",
                "date_extract_status": row.get("date_extract_status", "NOT_APPLICABLE"),
                "contract_id_extract_status": row.get(
                    "contract_id_extract_status", "FAIL"
                ),
                "提取后合同编号": None,
                "计算轨迹": serialize_calculation_trail(
                    [
                        {
                            "step": 1,
                            "action": "提取合同编号",
                            "input": None,
                            "output": None,
                            "error": "合同编号文本提取失败，无法匹配签收单",
                        }
                    ]
                ),
            }
        )
        results.append(record)

    # 签收单多余：无对应入账
    if not receipt_std.empty:
        for _, row in receipt_std.iterrows():
            key = str(row.get(match_key, "")).strip()
            if not key or key in ledger_keys:
                continue
            extra = row.to_dict()
            extract_status = extra.get("date_extract_status", "NOT_APPLICABLE")
            if pd.isna(extract_status) or extract_status is None:
                extract_status = "NOT_APPLICABLE"
            extra.update(
                {
                    "entry_date": None,
                    "expected_revenue_date": None,
                    "deviation_days": None,
                    "cutoff_status": "NO_LEDGER",
                    "issue_description": "签收单无对应入账记录",
                    "calculation_basis": f"{match_key}={key} 仅有签收单",
                    "date_extract_status": extract_status,
                    "contract_id_extract_status": extra.get(
                        "contract_id_extract_status", "NOT_APPLICABLE"
                    ),
                    "计算轨迹": serialize_calculation_trail(
                        [
                            {
                                "step": 1,
                                "action": "匹配序时账",
                                "input": key,
                                "output": None,
                                "error": "签收单无对应入账记录",
                            }
                        ]
                    ),
                }
            )
            results.append(extra)

    return pd.DataFrame(results)


def _apply_contract_id_extraction(
    df: pd.DataFrame,
    match_key: str,
    contract_text_columns: List[str],
    *,
    optional: bool = False,
) -> pd.DataFrame:
    """从文本列提取合同编号，写入 match_key / 提取后合同编号。"""
    out = df.copy()
    cols = [c for c in contract_text_columns if c in out.columns]
    if not cols:
        if optional:
            if "contract_id_extract_status" not in out.columns:
                out["contract_id_extract_status"] = "NOT_APPLICABLE"
            return out
        raise ValueError(
            "未找到可用于提取合同编号的文本列，请在界面选择摘要/备注等候选列"
        )

    extracted_ids = []
    statuses = []
    for _, row in out.iterrows():
        cid = extract_contract_id_from_row(row, cols)
        extracted_ids.append(cid)
        statuses.append("SUCCESS" if cid else "FAIL")

    out["提取后合同编号"] = extracted_ids
    out["contract_id_extract_status"] = statuses
    # 用提取结果覆盖匹配键；失败则为 None
    out[match_key] = extracted_ids
    return out


def _build_cutoff_record(
    *,
    row: pd.Series,
    match_key: str,
    payment_days: int,
    checker: CutoffChecker,
    receipt_keys: set,
) -> dict:
    key = str(row.get(match_key, "")).strip()
    entry_date = _as_date_str(row.get("entry_date"))
    receipt_date = _as_date_str(row.get("receipt_date"))
    extract_status = row.get("date_extract_status", "NOT_APPLICABLE")
    if pd.isna(extract_status) or extract_status is None:
        extract_status = "NOT_APPLICABLE"
    contract_status = row.get("contract_id_extract_status", "NOT_APPLICABLE")
    if pd.isna(contract_status) or contract_status is None:
        contract_status = "NOT_APPLICABLE"

    record = row.to_dict()
    record["date_extract_status"] = extract_status
    record["contract_id_extract_status"] = contract_status
    if not receipt_date:
        if key not in receipt_keys:
            issue = (
                f"无签收单，无法测试（合同编号「{key}」在签收单中不存在；"
                f"请检查两侧合同编号列是否映射为同一字段）"
            )
            basis = f"{match_key}={key} 未匹配到签收单"
        elif extract_status == "FAIL":
            issue = "签收日期文本提取失败，无法测试"
            basis = f"{match_key}={key} 已匹配签收单但日期提取失败"
        else:
            issue = "签收单无可用签收日期，无法测试"
            basis = f"{match_key}={key} 匹配到签收单但日期为空"
        record.update(
            {
                "expected_revenue_date": None,
                "deviation_days": None,
                "cutoff_status": "NO_RECEIPT",
                "issue_description": issue,
                "calculation_basis": basis,
                "计算轨迹": serialize_calculation_trail(
                    [
                        {
                            "step": 1,
                            "action": "匹配签收单",
                            "input": key,
                            "output": None,
                            "error": issue,
                        }
                    ]
                ),
            }
        )
    else:
        cutoff = checker.check(
            contract_payment_days=payment_days,
            receipt_date=receipt_date,
            entry_date=entry_date,
        )
        record.update(
            {
                "expected_revenue_date": cutoff.expected_revenue_date,
                "deviation_days": cutoff.deviation_days,
                "cutoff_status": cutoff.test_status,
                "issue_description": cutoff.issue_description,
                "calculation_basis": cutoff.calculation_basis,
                "计算轨迹": serialize_calculation_trail(cutoff.calculation_trail),
            }
        )
    return record



def _prepare_receipt_dates(
    receipt_std: pd.DataFrame,
    *,
    extract_date_from_text: bool,
    receipt_date_column: str,
) -> pd.DataFrame:
    """标准化 receipt_date，并按需从文本提取日期。"""
    # 兼容：映射后列名为 receipt_date，或仍为原始列名
    source_col = receipt_date_column
    if source_col not in receipt_std.columns:
        if "receipt_date" in receipt_std.columns:
            source_col = "receipt_date"
        else:
            raise ValueError(
                f"签收单缺少日期列: {receipt_date_column}（请先完成列映射）"
            )

    if extract_date_from_text:
        extracted_dates = []
        statuses = []
        for value in receipt_std[source_col]:
            text = None if pd.isna(value) else str(value)
            # 若单元格已是 YYYY-MM-DD，直接使用；否则从文本提取
            parsed = None
            if text:
                stripped = text.strip()
                if len(stripped) == 10 and stripped[4] == "-" and stripped[7] == "-":
                    try:
                        parsed = pd.to_datetime(stripped).date().isoformat()
                    except Exception:
                        parsed = extract_date_text(text)
                else:
                    parsed = extract_date_text(text)
            extracted_dates.append(parsed)
            statuses.append("SUCCESS" if parsed else "FAIL")
        receipt_std = receipt_std.copy()
        receipt_std["receipt_date"] = extracted_dates
        receipt_std["date_extract_status"] = statuses
        return receipt_std

    # 非提取模式：确保有 receipt_date；若已有 date_extract_status 则保留
    receipt_std = receipt_std.copy()
    if source_col != "receipt_date":
        receipt_std["receipt_date"] = receipt_std[source_col]
    if "receipt_date" not in receipt_std.columns:
        raise ValueError("签收单缺少 receipt_date 列（请先完成列映射标准化）")
    # 对仍是长文本的单元格做一次兜底提取，避免预览成功但执行未提取
    fixed_dates = []
    fixed_status = []
    has_prior_status = "date_extract_status" in receipt_std.columns
    for idx, value in enumerate(receipt_std["receipt_date"].tolist()):
        prior = (
            receipt_std["date_extract_status"].iloc[idx] if has_prior_status else None
        )
        as_date = _as_date_str(value)
        fixed_dates.append(as_date)
        if has_prior_status and prior in {"SUCCESS", "FAIL", "NOT_APPLICABLE"}:
            # 若先前标记成功但当前解析空，降为 FAIL
            if prior == "SUCCESS" and not as_date:
                fixed_status.append("FAIL")
            else:
                fixed_status.append(prior)
        elif as_date and isinstance(value, str) and len(str(value).strip()) > 10:
            fixed_status.append("SUCCESS")
        else:
            fixed_status.append("NOT_APPLICABLE")
    receipt_std["receipt_date"] = fixed_dates
    receipt_std["date_extract_status"] = fixed_status
    return receipt_std


def _as_date_str(value: Optional[object]) -> Optional[str]:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    # 已是标准日期字符串
    if isinstance(value, str):
        text = value.strip()
        if not text or text.lower() in {"nan", "nat", "none"}:
            return None
        if len(text) == 10 and text[4] == "-" and text[7] == "-":
            try:
                return pd.to_datetime(text).date().isoformat()
            except Exception:
                pass
        # 文本中夹杂日期：自动回退提取（防止勾选状态丢失时仍可用）
        extracted = extract_date_text(text)
        if extracted:
            return extracted
    try:
        return pd.to_datetime(value).date().isoformat()
    except Exception:
        text = str(value).strip()
        if not text or text.lower() in {"nan", "nat", "none"}:
            return None
        return extract_date_text(text)


def _normalize_match_key(series: pd.Series) -> pd.Series:
    """统一合同编号：去空白、去 .0 后缀、大写。"""
    out = series.astype(str).str.strip()
    out = out.str.replace(r"\.0$", "", regex=True)
    out = out.str.upper()
    return out.where(out.str.lower().ne("nan"), "")


def export_cutoff_excel(df: pd.DataFrame, output_path: str) -> str:
    """导出带状态表头着色的截止性测试结果 Excel。"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "截止性测试结果"

    headers = list(df.columns)
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True)
    status_fills = {
        "PASS": PatternFill("solid", fgColor="15803D"),
        "WARNING": PatternFill("solid", fgColor="CA8A04"),
        "FAIL": PatternFill("solid", fgColor="DC2626"),
        "NO_RECEIPT": PatternFill("solid", fgColor="64748B"),
        "NO_LEDGER": PatternFill("solid", fgColor="64748B"),
    }
    white = Font(color="FFFFFF", bold=True)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(1, col_idx, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    status_col = None
    if "cutoff_status" in headers:
        status_col = headers.index("cutoff_status") + 1

    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(r_idx, c_idx, _excel_value(value))
            if status_col and c_idx == status_col:
                status = str(value) if value is not None else ""
                fill = status_fills.get(status)
                if fill:
                    cell.fill = fill
                    cell.font = white

    for col_idx in range(1, len(headers) + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter]:
            max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
        ws.column_dimensions[letter].width = min(max(12, max_len + 2), 40)

    wb.save(output_path)
    return output_path


def _excel_value(value):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if pd.isna(value):
        return None
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:
            return str(value)
    return value
