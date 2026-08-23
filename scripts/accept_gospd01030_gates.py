"""GOSPD01030 质量门禁（in-process，不依赖已启动 API）。

用法（项目根）:
  set PYTHONPATH=.
  python scripts/accept_gospd01030_gates.py
"""

from __future__ import annotations

import sys
import tempfile
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))


def _fail(msg: str) -> None:
    raise AssertionError(msg)


def _base_docs(**kw):
    from tests.test_gospd01030 import _base_docs as bd

    return bd(**kw)


def _job(docs, **kw):
    from tests.test_gospd01030 import _job as j

    return j(docs, **kw)


def gate_g1_no_silent_period_end() -> None:
    from src.audit.gospd01030_assertions import (
        assert_correct_accounting_period,
        resolve_period_end,
    )

    job = _job(_base_docs())
    job.pop("period_end", None)
    if resolve_period_end(job) is not None:
        _fail("G1: 无 period_end 时不得解析出默认期末")
    r = assert_correct_accounting_period(
        posting_date="2026-01-03",
        control_date="2026-01-02",
        period_end=None,
        cutoff_status="PASS",
    )
    if r.get("evidence_status") != "NOT_TESTED" or r.get("verdict") is not None:
        _fail("G1: 缺期末须 NOT_TESTED 且 verdict=None")


def gate_g3_gate4_only() -> None:
    from src.reporting.gospd01030_filler import build_gospd01030_sample_rows

    skipped: list = []
    rows = build_gospd01030_sample_rows(
        _job(_base_docs(), matching_confirmed=False), skipped_chains=skipped
    )
    if rows:
        _fail("G3: Gate4 未确认不得有样本行")
    if not skipped or skipped[0].get("status") != "NOT_TESTED":
        _fail("G3: 未确认链须记 NOT_TESTED")


def gate_g4_strict_chain() -> None:
    from src.reporting.gospd01010_filler import group_classified_by_chain

    classified = [
        {
            "file_name": "o.pdf",
            "doc_type": "order",
            "fields": {"orderNo": "SO25-0099"},
        },
        {
            "file_name": "r.pdf",
            "doc_type": "receipt",
            "fields": {"documentNo": "R-ONLY"},
        },
    ]
    strict = dict(
        group_classified_by_chain(
            classified,
            allow_weak_unique_attach=False,
            allow_unique_so_ht_merge=False,
        )
    )
    so_docs = strict.get("SO25-0099") or []
    if any(d["file_name"] == "r.pdf" for d in so_docs):
        _fail("G4: 弱编号签收不得并入 SO")
    if "未识别业务号" not in strict:
        _fail("G4: 弱编号应进未识别业务号")


def gate_g5_g8_fill_workbook() -> None:
    from openpyxl import load_workbook

    from src.reporting.gospd01030_filler import fill_gospd01030_workbook
    from tests.test_gospd01030 import _w_dv_labels

    w_yes, w_no = _w_dv_labels()
    job = _job(
        _base_docs(
            acceptance_date="2025-12-28",
            posting_date="2026-01-03",
            receipt_amount=1130,
        )
    )
    with tempfile.TemporaryDirectory() as td:
        out = fill_gospd01030_workbook(job, Path(td) / "g.xlsx")
        ws = load_workbook(out).active
        v = str(ws["V30"].value or "")
        if not v.startswith("=") or "$M$5" not in v:
            _fail("G5: V30 须保留含 $M$5 的公式")
        if ws["W30"].value != w_no:
            _fail(f"G8: W 须写 DV 长否文案，实际={ws['W30'].value!r}")
        if ws["W30"].value == "No 否":
            _fail("G8: 禁止短 No 否 绕过 DV")
        if ws["R30"].value != 1130:
            _fail("可比金额场景 R 应写入")


def gate_g6_formula_conflict() -> None:
    from src.reporting.gospd01030_filler import build_gospd01030_sample_rows

    job = _job(
        _base_docs(
            acceptance_date="2025-12-15",
            posting_date="2025-12-20",
            receipt_amount=1130,
        )
    )
    row = build_gospd01030_sample_rows(job)[0]
    if "FORMULA_LOGIC_CONFLICT" not in str(row.get("formula_conflict") or ""):
        _fail("G6: 应检出公式冲突")
    if row.get("all_ok"):
        _fail("G6: 冲突时 W 不得写「是」")


def gate_g9_r_blank() -> None:
    from src.reporting.gospd01030_filler import build_gospd01030_sample_rows

    row = build_gospd01030_sample_rows(_job(_base_docs(receipt_amount=None)))[0]
    if row.get("amt_delivery") is not None:
        _fail("G9: 无可比金额时 amt_delivery 须为 None")


def gate_g10_ar_period() -> None:
    from src.audit.gospd01030_assertions import (
        assert_ar_correct_period,
        build_gospd01030_assertions,
    )
    from src.audit.program_matrix import get_program_matrix

    r = assert_ar_correct_period(
        posting_date="2026-01-03",
        control_date="2025-12-28",
        period_end=date(2025, 12, 31),
        cutoff_status="PASS",
        revenue_period_verdict=False,
    )
    if r.get("verdict") is not False:
        _fail("G10: 收入否 → 应收否")
    job = _job(_base_docs())
    a = build_gospd01030_assertions(
        docs=job["classified"], job=job, three_way=job["three_way"]
    )
    if not a.get("ar_period") or a["ar_period"].get("verdict") is not True:
        _fail("G10: 正常样本须产出 ar_period=True")
    m = get_program_matrix("gospd01030")
    ids = [x.get("id") for x in (m.get("assertions") or [])]
    if "ar_period" not in ids:
        _fail("G10: program_matrix 缺 ar_period")


def gate_g11_recipe() -> None:
    from src.workflow.recipes import resolve_workflow_plan

    plan = resolve_workflow_plan(["gospd01030"])
    if "three_way_cutoff" not in plan["required_steps"]:
        _fail("G11: 01030 须含 three_way_cutoff")
    for skip in ("contract_terms", "amount_test"):
        if skip not in plan["skipped_steps"]:
            _fail(f"G11: 01030 应跳过 {skip}")


def gate_g12_voucher() -> None:
    from src.reporting.gospd01030_filler import build_gospd01030_sample_rows

    job = _job(_base_docs())
    job["classified"][3].pop("ledger_voucher", None)
    job["classified"][3]["fields"]["documentNo"] = "SO25-0099"
    job["classified"][3]["fields"]["voucherNo"] = ""
    if build_gospd01030_sample_rows(job)[0].get("voucher"):
        _fail("G12: 订单号不得冒充凭证号")


def main() -> int:
    gates = [
        ("G1", gate_g1_no_silent_period_end),
        ("G3", gate_g3_gate4_only),
        ("G4", gate_g4_strict_chain),
        ("G5+G8", gate_g5_g8_fill_workbook),
        ("G6", gate_g6_formula_conflict),
        ("G9", gate_g9_r_blank),
        ("G10", gate_g10_ar_period),
        ("G11", gate_g11_recipe),
        ("G12", gate_g12_voucher),
    ]
    errors: list[str] = []
    for name, fn in gates:
        try:
            fn()
            print(f"OK {name}")
        except Exception as exc:  # noqa: BLE001
            print(f"FAIL {name}: {exc}")
            errors.append(f"{name}: {exc}")
    if errors:
        print("FAIL accept_gospd01030_gates")
        return 1
    print("OK accept_gospd01030_gates all green")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
