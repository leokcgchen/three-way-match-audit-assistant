"""GOSPD01030 真实样本端到端签收（SO25-0281 seed → 门禁 → 导出验表）。

默认用 FastAPI TestClient（始终加载当前代码）。
可选：ACCEPT_01030_HTTP=1 打已启动的 http://127.0.0.1:8000。

用法（项目根）:
  set PYTHONPATH=.
  python scripts/accept_gospd01030_e2e.py
"""

from __future__ import annotations

import json
import os
import subprocess
import sys
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path
from typing import Any, Optional

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

BASE = os.environ.get("ACCEPT_01030_BASE", "http://127.0.0.1:8000").rstrip("/")
USE_HTTP = os.environ.get("ACCEPT_01030_HTTP", "").strip() in {"1", "true", "yes"}
PERIOD_END = "2025-12-15"
LOG_SHEET = "02_填制依据与运行日志"

_client = None


def _get_client():
    global _client
    if _client is None:
        from fastapi.testclient import TestClient

        from src.api.main import app

        _client = TestClient(app)
    return _client


def call(method: str, path: str, data=None, *, timeout: int = 300) -> Any:
    if not USE_HTTP:
        client = _get_client()
        kw: dict[str, Any] = {}
        if data is not None:
            kw["json"] = data
        resp = client.request(method.upper(), path, **kw)
        if resp.status_code >= 400:
            raise RuntimeError(f"{method} {path} -> {resp.status_code}: {resp.text}")
        if not resp.content:
            return None
        ct = resp.headers.get("content-type") or ""
        if "application/json" in ct:
            return resp.json()
        return resp.content

    body = None if data is None else json.dumps(data).encode("utf-8")
    req = urllib.request.Request(BASE + path, data=body, method=method)
    if body is not None:
        req.add_header("Content-Type", "application/json")
    try:
        with urllib.request.urlopen(req, timeout=timeout) as r:
            raw = r.read()
            if not raw:
                return None
            ct = r.headers.get("Content-Type") or ""
            if "application/json" in ct:
                return json.loads(raw.decode("utf-8"))
            return raw
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"{method} {path} -> {exc.code}: {detail}") from exc


def ok(msg: str) -> None:
    print(f"  OK  {msg}")


def step(title: str) -> None:
    print(f"\n== {title} ==")


def _clear_blockers(jid: str) -> None:
    try:
        adv = call("GET", f"/api/v1/workflow/jobs/{jid}/advisory")
        for p in adv.get("pending") or []:
            cid = p.get("candidate_id")
            if cid:
                call(
                    "POST",
                    f"/api/v1/workflow/jobs/{jid}/advisory/{urllib.parse.quote(cid)}/decide",
                    {"status": "REJECTED", "reason": "e2e", "auto_replay": False},
                )
    except Exception as exc:  # noqa: BLE001
        print(f"  WARN advisory: {exc}")

    job = call("GET", f"/api/v1/workflow/jobs/{jid}")
    dup = job.get("duplicates") or {}
    if dup.get("blocks_downstream_hint"):
        try:
            call(
                "POST",
                f"/api/v1/workflow/jobs/{jid}/duplicates/acknowledge",
                {"reason": "e2e 知悉放行"},
            )
        except Exception as exc:  # noqa: BLE001
            print(f"  WARN duplicates: {exc}")

    for r in job.get("relations") or []:
        if r.get("status") == "PROPOSED":
            call(
                "POST",
                f"/api/v1/workflow/jobs/{jid}/relations/{urllib.parse.quote(r['relation_id'])}/decide",
                {"status": "VERIFIED", "reason": "e2e"},
            )

    try:
        trace = call("GET", f"/api/v1/workflow/jobs/{jid}/conclusion-trace")
    except Exception as exc:  # noqa: BLE001
        print(f"  WARN conclusion-trace: {exc}")
        return
    for f in trace.get("findings") or []:
        if not f.get("blocking") or f.get("acknowledged"):
            continue
        call(
            "POST",
            f"/api/v1/workflow/jobs/{jid}/hitl/finding/acknowledge",
            {
                "finding_id": f["finding_id"],
                "genuine": True,
                "reason": "e2e 确认为单据问题",
            },
        )


def _pick_so_chain(jid: str) -> str:
    chains = call("GET", f"/api/v1/workflow/jobs/{jid}/chains")
    chain_list = chains.get("chains") or []
    if not chain_list:
        raise RuntimeError("未识别到业务链")
    primary = next(
        (c for c in chain_list if str(c.get("chain_id") or "").upper().startswith("SO")),
        chain_list[0],
    )
    return str(primary["chain_id"])


def _ensure_cross_refs(jid: str, classified: list[dict]) -> list[dict]:
    for c in classified:
        dt = c.get("doc_type")
        fn = c.get("file_name") or ""
        fields = dict(c.get("fields") or {})
        patch: dict = {}
        if dt == "order":
            if not fields.get("orderNo"):
                patch["orderNo"] = "SO25-0281"
            if not fields.get("contractNo"):
                patch["contractNo"] = "HT25-0281"
            if not fields.get("buyerName"):
                patch["buyerName"] = "华东某整车制造有限公司"
        elif dt == "contract" and not fields.get("contractNo"):
            patch["contractNo"] = "HT25-0281"
        elif dt in {"receipt", "delivery"}:
            if not fields.get("orderNo"):
                patch["orderNo"] = "SO25-0281"
            if dt == "receipt" and not (
                fields.get("acceptanceDate") or fields.get("documentDate")
            ):
                patch["acceptanceDate"] = "2025-12-20"
            if not fields.get("buyerName"):
                patch["buyerName"] = "华东某整车制造有限公司"
        elif dt == "invoice":
            if not fields.get("orderNo"):
                patch["orderNo"] = "SO25-0281"
            if not fields.get("contractNo"):
                patch["contractNo"] = "HT25-0281"
            if not fields.get("postingDate"):
                patch["postingDate"] = "2025-12-20"
            if not fields.get("buyerName"):
                patch["buyerName"] = "华东某整车制造有限公司"
        if patch and fn:
            call(
                "PATCH",
                f"/api/v1/workflow/jobs/{jid}/documents/fields",
                {"file_name": fn, "fields": patch},
            )
    job = call("GET", f"/api/v1/workflow/jobs/{jid}")
    return list(job.get("classified") or [])


def _assert_workbook(path: Path) -> None:
    from openpyxl import load_workbook

    if not path.is_file():
        raise RuntimeError(f"导出文件不存在: {path}")
    wb = load_workbook(path)
    ws = wb.active
    if ws["F5"].value != "GOSPD01030":
        raise RuntimeError(f"F5 程序号异常: {ws['F5'].value!r}")
    m5 = ws["M5"].value
    ok_m5 = False
    if hasattr(m5, "year"):
        ok_m5 = m5.year == 2025 and m5.month == 12 and m5.day == 15
    else:
        ok_m5 = "2025-12-15" in str(m5 or "")
    if not ok_m5:
        raise RuntimeError(f"M5 期末异常: {m5!r}")
    v = str(ws["V30"].value or "")
    if not v.startswith("=") or "$M$5" not in v:
        raise RuntimeError(f"V30 须保留公式: {v!r}")
    if ws["B30"].value in (None, ""):
        raise RuntimeError("B30 无样本编号：Gate4 链未写入正式行？")
    if LOG_SHEET not in wb.sheetnames:
        raise RuntimeError(f"缺日志页 {LOG_SHEET}")
    log = wb[LOG_SHEET]
    texts = " ".join(
        str(log.cell(r, c).value or "")
        for r in range(1, min(log.max_row, 120) + 1)
        for c in range(1, 8)
    )
    if "步骤3_应收账款期间" not in texts:
        raise RuntimeError("日志缺步骤3_应收账款期间")
    ok(f"workbook F5/M5/V/B30/步骤3 OK path={path}")


def main() -> int:
    mode = "HTTP " + BASE if USE_HTTP else "TestClient(in-process)"
    step(f"health ({mode})")
    health = call("GET", "/health")
    if not health or health.get("status") != "ok":
        print("FAIL health", health)
        return 1
    ver = str(health.get("version") or "")
    ok(f"version={ver} formal_ocr={(health.get('audit') or {}).get('formal_ocr')}")

    step("job + goals gospd01030 + period_end")
    job = call("POST", "/api/v1/workflow/jobs", {"title": "01030-e2e-SO25-0281"})
    jid = job["job_id"]
    job = call(
        "PUT",
        f"/api/v1/workflow/jobs/{jid}/goals",
        {
            "goal_ids": ["gospd01030"],
            "period_end": PERIOD_END,
            "entity_name": "01030端到端签收主体",
        },
    )
    plan = job.get("plan") or {}
    assert "gospd01030" in (job.get("goal_ids") or [])
    assert "three_way_cutoff" in (plan.get("required_steps") or [])
    assert job.get("period_end") == PERIOD_END
    ok(f"job={jid} period_end={PERIOD_END}")

    raise RuntimeError("快速注入验收已移除；请运行 scripts/accept_gospd01030_ocr_e2e.py，使用明确的 PDF 夹具完成正式上传与识别。")
    classified = _ensure_cross_refs(jid, list(job.get("classified") or []))
    if len(classified) < 4:
        raise RuntimeError(f"seed 单据不足: {len(classified)}")
    types = sorted({str(c.get("doc_type")) for c in classified})
    ok(f"docs={len(classified)} types={types}")
    order = next(c for c in classified if c.get("doc_type") == "order")
    if not (order.get("fields") or {}).get("contractNo"):
        raise RuntimeError("订单缺 contractNo")
    ok("chain cross-refs present")

    step("Gate4 未确认 → preview 无正式行")
    preview0 = call("GET", f"/api/v1/workflow/jobs/{jid}/workbook-rows/preview")
    rows0 = preview0.get("rows") or []
    if rows0:
        raise RuntimeError(f"Gate4 未确认不得有 preview 行: {len(rows0)}")
    ok("preview rows=0 before gate4")

    step("HITL fields → evidence → gate4")
    job = call("POST", f"/api/v1/workflow/jobs/{jid}/hitl/fields/confirm")
    assert job.get("fields_confirmed")
    primary_id = _pick_so_chain(jid)
    job = call("PUT", f"/api/v1/workflow/jobs/{jid}/active-chain", {"chain_id": primary_id})
    ok(f"active_chain={job.get('active_chain_id')}")
    job = call("POST", f"/api/v1/workflow/jobs/{jid}/evidence-match")
    ok(f"evidence={(job.get('evidence') or {}).get('status')}")
    _clear_blockers(jid)
    job = call(
        "POST",
        f"/api/v1/workflow/jobs/{jid}/hitl/matching/confirm",
        {"reason": "01030 e2e"},
    )
    sample = ((job.get("gospd_sample_results") or {}).get(primary_id) or {})
    assert sample.get("matching_confirmed") or job.get("matching_confirmed")
    ok(f"gate4 chain={primary_id}")

    step("three-way-cutoff")
    job = call("POST", f"/api/v1/workflow/jobs/{jid}/three-way-cutoff", {})
    tw = job.get("three_way") or {}
    ok(f"three_way overall={tw.get('overall_status') or tw.get('status')}")

    step("改 period_end 须级联失效三单")
    job = call(
        "PUT",
        f"/api/v1/workflow/jobs/{jid}/goals",
        {
            "goal_ids": ["gospd01030"],
            "period_end": "2025-12-16",
            "entity_name": "01030端到端签收主体",
        },
    )
    if job.get("three_way"):
        raise RuntimeError("改期末后 three_way 应被清空")
    ok("period_end cascade cleared three_way")
    job = call(
        "PUT",
        f"/api/v1/workflow/jobs/{jid}/goals",
        {
            "goal_ids": ["gospd01030"],
            "period_end": PERIOD_END,
            "entity_name": "01030端到端签收主体",
        },
    )
    sample = ((job.get("gospd_sample_results") or {}).get(primary_id) or {})
    if not (sample.get("matching_confirmed") or job.get("matching_confirmed")):
        _clear_blockers(jid)
        job = call(
            "POST",
            f"/api/v1/workflow/jobs/{jid}/hitl/matching/confirm",
            {"reason": "01030 e2e re-gate4"},
        )
        ok("gate4 re-confirmed after period change")
    job = call("POST", f"/api/v1/workflow/jobs/{jid}/three-way-cutoff", {})
    ok(f"three_way rerun={ (job.get('three_way') or {}).get('overall_status') }")

    step("Gate5 + export")
    _clear_blockers(jid)
    sample = ((job.get("gospd_sample_results") or {}).get(primary_id) or {})
    if not (sample.get("matching_confirmed") or job.get("matching_confirmed")):
        job = call(
            "POST",
            f"/api/v1/workflow/jobs/{jid}/hitl/matching/confirm",
            {"reason": "01030 e2e gate4 before gate5"},
        )
    job = call(
        "POST",
        f"/api/v1/workflow/jobs/{jid}/hitl/conclusion/confirm",
        {"reason": "01030 e2e"},
    )
    assert job.get("conclusion_confirmed") is True
    ok("gate5 confirmed")
    job = call("POST", f"/api/v1/workflow/jobs/{jid}/workbook/export")
    wpath = job.get("workbook_path") or ""
    if not wpath:
        paths = job.get("workbook_paths") or []
        wpath = paths[0] if paths else ""
    if not wpath:
        raise RuntimeError(f"导出无路径: keys={list(job.keys())}")
    _assert_workbook(Path(wpath))

    step("in-process 单元门禁")
    r = subprocess.run(
        [sys.executable, str(ROOT / "scripts" / "accept_gospd01030_gates.py")],
        cwd=str(ROOT),
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        env={**os.environ, "PYTHONPATH": str(ROOT)},
    )
    print(r.stdout)
    if r.returncode != 0:
        print(r.stderr)
        raise RuntimeError("accept_gospd01030_gates failed")

    print("\nALL 01030 E2E CHECKS PASSED")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:  # noqa: BLE001
        print(f"\nFAIL: {exc}")
        raise
